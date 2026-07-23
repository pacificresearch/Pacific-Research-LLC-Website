#!/usr/bin/env python3
"""
SAM.gov API Capability Matcher & Screening Script
==================================================

Pulls active federal contract opportunities from the SAM.gov Opportunities API,
filters and evaluates them against the capabilities of Pacific Research Group LLC
(a Service-Disabled Veteran-Owned Small Business), and prints a Markdown
screening report to the console.

The report is organized into four tiers:
    1. Core-target opportunities  — deep capability match (primary NAICS).
    2. Actionable recommendations — top 3 core matches to pursue first.
    3. Low-barrier / "warm body" opportunities — staffing/administrative/support
       work under secondary NAICS that PRG can reasonably win without deep
       domain specialization (mainly needs eligible, qualified personnel).
    4. Subcontracting & teaming opportunities — awarded primes and large
       unrestricted solicitations where PRG's realistic path is teaming as a
       subcontractor rather than bidding as prime.

Usage:
    python3 samgov_opportunity_matcher.py

Optional command-line overrides:
    --days N        Look back N days for posted opportunities (default: 30)
    --limit N       Max opportunities to request per NAICS query (default: 100)
    --api-key KEY   Override the hardcoded API key (or set env SAM_API_KEY)

Notes:
    * The script is self-contained; only the standard library and `requests`
      are required (`pip install requests`).
    * SAM.gov requires date filters (postedFrom / postedTo) in MM/DD/YYYY
      format and caps ranges at one year. Results are paginated.
"""

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
import textwrap
import time
from collections import Counter, OrderedDict

try:
    import requests
except ImportError:  # pragma: no cover
    sys.stderr.write(
        "ERROR: The 'requests' library is required. Install it with:\n"
        "    pip install requests\n"
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# 1. INPUT DATA CONFIGURATION
# ---------------------------------------------------------------------------

# API access. Prefer the environment variable so the key is not committed to
# history, but fall back to the value provided for turnkey execution.
API_KEY = os.environ.get(
    "SAM_API_KEY", "SAM-a87c1f1f-af4a-4e49-b2cd-7e21b68022a2"
)

# SAM.gov Opportunities API (v2 is the current GA endpoint; v1 is deprecated).
API_URL = "https://api.sam.gov/opportunities/v2/search"

# Company profile ----------------------------------------------------------
COMPANY_NAME = "Pacific Research Group LLC"
COMPANY_WEBSITE = "https://pacificresearchllc.com/"
SOCIOECONOMIC_STATUS = "SDVOSB (Service-Disabled Veteran-Owned Small Business)"

# Core capabilities (used to build the keyword-matching corpus).
CORE_CAPABILITIES = [
    "Clinical trial operations, clinical research coordination, and site "
    "management (Phases II-IV).",
    "Decentralized and hybrid clinical trial models, eConsent, ePRO/eCOA, and "
    "telehealth workflows.",
    "Regulatory compliance (ICH GCP, 21 CFR, HIPAA, CITI GCP) and audit "
    "readiness.",
    "Data analytics, electronic data capture (EDC), REDCap, CTMS, OnCore, "
    "Medidata Rave management, and database optimization.",
    "Biomedical equipment management, diagnostics, maintenance, and technical "
    "operations (DoD/VA healthcare settings).",
    "Emergency and trauma medicine clinical workflows, training, and "
    "healthcare technology deployment.",
]

# Tier 1 — PRIMARY / CORE-TARGET NAICS codes (deep capability match).
TARGET_NAICS = OrderedDict(
    [
        ("541715", "R&D in the Physical, Engineering, and Life Sciences"),
        ("541714", "R&D in Biotechnology (except Nanobiotechnology)"),
        ("541690", "Other Scientific & Technical Consulting Services"),
        ("541611", "Administrative & General Management Consulting"),
        ("541990", "All Other Professional, Scientific & Technical Services"),
        ("811210", "Electronic & Precision Equipment Repair & Maintenance"),
    ]
)

# Tier 2 — LOW-BARRIER / "WARM BODY" NAICS codes.
# These are labor-, staffing-, and support-heavy service categories that PRG,
# as an eligible SDVOSB, can realistically win by fielding qualified personnel
# rather than by deep technical differentiation. They are adjacent to PRG's
# healthcare / veteran-services footprint but have a lower barrier to entry.
LOW_BARRIER_NAICS = OrderedDict(
    [
        ("561320", "Temporary Help Services (medical/admin staffing)"),
        ("561110", "Office Administrative Services"),
        ("561210", "Facilities Support Services"),
        ("561499", "All Other Business Support Services"),
        ("561410", "Document Preparation Services"),
        ("541618", "Other Management Consulting Services"),
        ("541519", "Other Computer Related Services (IT staff aug)"),
        ("611430", "Professional & Management Development Training"),
        ("621999", "All Other Misc. Ambulatory Health Care Services"),
        ("621399", "Offices of All Other Misc. Health Practitioners"),
        ("621512", "Diagnostic Imaging Centers (support staffing)"),
        ("561612", "Security Guards & Patrol Services"),
        ("561720", "Janitorial Services"),
        ("488190", "Other Support Activities for Air Transportation"),
        # Research / analysis / writing codes — rich in solo-doable knowledge work
        ("541720", "R&D in Social Sciences & Humanities"),
        ("541910", "Marketing Research & Public Opinion Polling"),
        ("611710", "Educational Support Services"),
    ]
)

# Tier 3 — VIBE-CODE / light web & software NAICS. Simple builds a solo founder
# + AI can ship without a dev team: static sites, landing pages, small web apps,
# forms, dashboards, CMS/WordPress, and Section 508 / accessibility remediation.
# PRG delivers via its no-/low-code stack (Screendot for CO-facing intros &
# capability/landing sites; Digipop for app builds).
VIBECODE_NAICS = OrderedDict(
    [
        ("541511", "Custom Computer Programming Services"),
        ("541512", "Computer Systems Design Services"),
        ("541513", "Computer Facilities Management Services"),
        ("541430", "Graphic Design Services (web/UI)"),
        ("518210", "Data Processing, Hosting & Related Services"),
        # 541519 (Other Computer Related) is already queried under low-barrier.
    ]
)

# Every NAICS we actually query the API for (primary, low-barrier, vibe-code).
ALL_QUERY_NAICS = OrderedDict()
ALL_QUERY_NAICS.update(TARGET_NAICS)
ALL_QUERY_NAICS.update(LOW_BARRIER_NAICS)
ALL_QUERY_NAICS.update(VIBECODE_NAICS)

# Vibe-code fit signals: light web/software work one person + AI can deliver.
VIBECODE_KEYWORDS = [
    "website", "web site", "web design", "web development", "web application",
    "web app", "landing page", "static site", "static website", "microsite",
    "content management", "wordpress", "drupal", "joomla", "cms",
    "front-end", "front end", "user interface", "ui/ux", "ux/ui",
    "responsive design", "web portal", "portal", "dashboard",
    "data visualization", "web form", "online form", "intake form",
    "section 508", "508 compliance", "508 remediation", "accessibility",
    "wcag", "prototype", "mvp", "minimum viable product", "single page",
    "html", "css", "javascript", "no-code", "low-code", "web-based tool",
    "web based application", "simple database", "site redesign", "web hosting",
    "landing pages", "digital form", "e-form", "online portal",
]
# What PRG uses to build & deliver these (surfaced in the report tab header).
VIBECODE_DELIVERY_NOTE = (
    "Delivery stack — Screendot (CO-facing intros, capability & landing sites) "
    "and Digipop (app builds). Scope every bid to what one person + AI can ship: "
    "static/landing sites, small web apps, forms, dashboards, CMS/WordPress, and "
    "Section 508 / accessibility remediation. Anything needing an ATO, a security "
    "package, or an embedded dev team is out of this lane.")

# Keywords that signal a technical-competency match. Kept lowercase for
# case-insensitive substring matching against title + description.
CAPABILITY_KEYWORDS = [
    "clinical research", "clinical trial", "clinical trials", "clinical study",
    "clinical studies", "clinical operations", "clinical site", "site management",
    "clinical coordinator", "research coordination", "research operations",
    "decentralized trial", "hybrid trial", "econsent", "epro", "ecoa",
    "telehealth", "telemedicine", "good clinical practice", "gcp", "ich gcp",
    "21 cfr", "hipaa", "audit readiness", "regulatory compliance",
    "data analytics", "electronic data capture", "edc", "redcap", "ctms",
    "oncore", "medidata", "rave", "database", "data management",
    "biomedical equipment", "biomedical", "medical equipment",
    "diagnostic", "diagnostics", "equipment maintenance", "equipment repair",
    "precision equipment", "healthcare technology", "health technology",
    "emergency medicine", "trauma", "trauma medicine", "medical research",
    "research and development", "life sciences", "pharmaceutical",
    "protocol", "irb", "institutional review board", "patient recruitment",
    "healthcare", "health care", "medical", "laboratory", "lab equipment",
]

# Keywords that signal a "warm body" / low-barrier staffing or support need —
# work where fielding eligible, credentialed personnel is the main requirement.
LOW_BARRIER_KEYWORDS = [
    "staffing", "staff augmentation", "temporary staffing", "personnel",
    "administrative support", "admin support", "office support", "clerical",
    "data entry", "records management", "medical records", "scheduling",
    "front desk", "reception", "call center", "help desk",
    "facilities support", "custodial", "janitorial", "housekeeping",
    "grounds", "security guard", "security services", "logistics support",
    "warehouse", "supply", "courier", "transcription", "coding",
    "credentialing", "training support", "instructor", "program support",
    "technician", "phlebotom", "medical assistant", "nursing", "cna",
    "patient transport", "escort", "labor", "operations support",
]

# Keywords that indicate a subcontracting / teaming angle.
SUBCONTRACT_KEYWORDS = [
    "subcontract", "subcontracting plan", "subcontractor", "teaming",
    "small business subcontracting", "sub-award", "subaward", "joint venture",
]

# --- "Solo-friendly" detection ------------------------------------------------
# Knowledge/office work that one capable person (with an AI assistant) can
# deliver: research, analysis, writing, data, reviews, small advisory tasks.
SOLO_FRIENDLY_KEYWORDS = [
    "research", "study", "studies", "analysis", "analytical", "report",
    "white paper", "assessment", "evaluation", "survey", "market research",
    "literature review", "systematic review", "data analysis", "data entry",
    "data management", "transcription", "translation", "editing", "proofreading",
    "technical writing", "writing", "documentation", "curriculum",
    "training material", "instructional design", "consulting", "advisory",
    "subject matter expert", "sme", "abstract", "records review", "chart review",
    "medical coding", "coding", "audit support", "program support",
    "administrative support", "grant writing", "proposal support",
    "policy analysis", "feasibility", "needs assessment", "sbir", "sttr",
    "phase i", "spreadsheet", "database design", "analyst", "review of",
]

# Signals that a task needs a crew, physical presence, or a large team —
# these push an opportunity OUT of "solo-friendly".
SOLO_EXCLUDE_KEYWORDS = [
    "construction", "renovation", "janitorial", "custodial", "security guard",
    "guard services", "landscaping", "grounds maintenance", "hvac", "plumbing",
    "roofing", "food service", "warehouse", "fleet", "installation",
    "manufacture", "manufacturing", "onsite staffing", "full-time equivalents",
    "ftes", "24/7", "around the clock", "nationwide", "multiple locations",
    "guards", "custodians", "shift work", "call center", "help desk",
]

# Firm-Fixed-Price small-buy thresholds (FAR): micro-purchase and the
# Simplified Acquisition Threshold — smaller usually means solo-doable.
MICRO_PURCHASE = 10_000
SIMPLIFIED_ACQ_THRESHOLD = 250_000

_FTE_RE = re.compile(
    r"(\d{1,3})\s*(?:\+\s*)?(?:full[-\s]?time|fte'?s?|full time equivalents?)",
    re.IGNORECASE,
)

# --- Company certifications ---------------------------------------------------
# Toggle these on as PRG earns them; each unlocks the matching set-aside as
# "eligible to bid". SDVOSB is assumed. HUBZone is location-based — flip it on
# only if PRG's principal office is in a HUBZone (check the SBA HUBZone map).
IS_HUBZONE_CERTIFIED = False
IS_8A_CERTIFIED = False
IS_WOSB_CERTIFIED = False

# Set-aside codes/labels that this company is eligible for. SAM.gov returns a
# `typeOfSetAside` code and a `typeOfSetAsideDescription`. We map the codes.
# See: https://open.gsa.gov/api/opportunities-api/  (set-aside code table)
SDVOSB_SETASIDE_CODES = {"SDVOSBC", "SDVOSBS"}          # SDVOSB set-aside / sole source
SMALL_BUSINESS_SETASIDE_CODES = {"SBA", "SBP"}          # Total Small Business
VOSB_SETASIDE_CODES = {"VSA", "VSS"}                    # VOSB (SDVOSB qualifies)
HUBZONE_CODES = {"HZC", "HZS"}
EIGHTA_CODES = {"8A", "8AN"}
WOSB_CODES = {"WOSB", "WOSBSS", "EDWOSB", "EDWOSBSS"}

# WIDE-NET sweep: every set-aside code PRG is eligible for, queried across ALL
# NAICS (not just the target list). This is the widest *relevant* net — PRG's
# structural edge is being an SDVOSB, so a set-aside anywhere is winnable
# regardless of NAICS. The screening gates then find the compelling ones and the
# NAICS-gap report flags codes PRG should add to its SAM registration.
SETASIDE_QUERY_CODES = ["SDVOSBC", "SDVOSBS", "VSA", "VSS", "SBA", "SBP"]

# Target PSC (Product/Service) codes — more specific than NAICS. Queried in
# addition to NAICS to catch work classified by service type, which pros do.
TARGET_PSC = OrderedDict([
    ("R408", "Program Management / Support Services"),
    ("R499", "Other Professional Services"),
    ("R707", "Management Support Services"),
    ("B505", "Special Studies/Analysis — Data"),
    ("B506", "Special Studies/Analysis — Medical"),
    ("Q301", "Medical — Laboratory Testing / Analysis"),
    ("AN11", "R&D — Health/Medical (applied research)"),
])

# --- Deep-screen gates (PRG capability baseline: solo clinical-research pro,
# no employees, no trade crew, no equipment/factory certs) --------------------
# GATE 4 — scope reality: work PRG's sole performer CANNOT self-deliver. A hit
# is a hard FAIL (killed), not a score-down.
DISQUALIFIER_KEYWORDS = [
    "construction", "renovation", "demolition", "build-out", "new construction",
    "janitorial", "custodial", "housekeeping", "grounds maintenance",
    "landscaping", "pest control", "snow removal", "facility maintenance",
    "facilities maintenance", "hvac", "plumbing", "electrical work",
    "electrician", "welding", "carpentry", "roofing", "painting services",
    "masonry", "boiler", "elevator", "generator maintenance", "fire alarm",
    "fire suppression", "equipment maintenance", "equipment repair",
    "preventive maintenance", "calibration services", "install and maintain",
    "installation and maintenance", "logistics", "distribution", "warehouse",
    "supply chain", "freight", "courier", "fleet", "food service", "cafeteria",
    "laundry", "security guard", "guard services", "armed guard",
    "unarmed guard", "tree removal", "fencing", "groundskeeping",
]
# Commodity supply / distribution of goods — PRG performs professional services,
# not product supply. A "supply verb" paired with a product/supply classification
# (manufacturing/wholesale NAICS, numeric or 65xx PSC) is a kill; an analytic
# "service verb" (analyze/coordinate/consult) keeps the notice in PRG's lane.
# This pair also gates the "medical keyword alone" false positive: for a supply
# classification, the VERB decides whether a medical topic is in-lane work.
SUPPLY_VERB_KEYWORDS = [
    "supply and delivery", "supply and deliver", "furnish and deliver",
    "delivery of goods", "procurement of supplies", "procurement of medical",
    "provide and deliver", "delivery order for", "brand name or equal",
    "supply of test", "supply of medical", "supply of laboratory",
]
SERVICE_VERB_KEYWORDS = [
    "analyze", "analysis", "coordinate", "coordination", "consult",
    "consulting", "advisory", "monitor", "monitoring", "assessment",
    "evaluate", "evaluation", "research", "data management", "regulatory",
    "protocol", "study design", "technical writing", "program support",
]
# GATE 3 — coverage killers: embedded/around-the-clock/guaranteed-response or
# simultaneous multi-site presence a solo performer cannot commit to.
COVERAGE_KILLER_KEYWORDS = [
    "24/7", "24x7", "24 x 7", "around the clock", "on-call", "on call",
    "emergency response", "guaranteed response", "response time",
    "embedded", "on-site daily", "daily on-site", "full-time on-site",
    "weekly on-site", "shift work", "rotating shift", "multiple simultaneous",
    "simultaneous locations", "staff on site at all times",
]
# GATE 0.5 — structural eligibility: assets/relationships the offeror must hold
# TODAY (not capability). A hit is a hard FAIL (can't be built before the due date).
STRUCTURAL_FAIL_KEYWORDS = [
    "facility clearance required", "facility security clearance",
    "must possess a facility", "requires a facility security clearance",
    "active secret clearance", "top secret clearance required",
    "must hold an active security clearance", "ts/sci",
    "gsa schedule holders", "schedule holders only", "must hold a gsa",
    "must have an active gsa", "gsa mas required", "idiq holders only",
    "idiq holder", "bpa holders", "basic ordering agreement holders",
    "prior phase i", "completed phase i", "phase i award is required",
    "must be a prior phase", "phase ii only", "successful phase i",
    "must own", "proprietary product", "existing prototype",
    "originated at a university", "university-originated",
    "must have an existing", "commercially available product must",
]
# Notice types that are NOT a bid yet → route to the WATCHLIST (prep, don't bid).
FUTURE_NOTICE_MARKERS = ("presolicitation", "sources sought", "special notice",
                         "forecast", "intent to")

# --- Stage-0 timing / notice-type gating (runs BEFORE capability scoring) -----
WATCH_NOTICE_CLASSES = {"PRESOL", "SOURCES_SOUGHT", "SPECIAL", "RFI", "FORECAST"}
SHORT_FUSE_CALENDAR_DAYS = 7      # ~5 business days


def _classify_notice(notice_type):
    """Map a SAM.gov notice type to a routing class."""
    nt = (notice_type or "").lower()
    if "amendment" in nt or "modification" in nt:
        return "AMENDMENT"
    if "award" in nt:
        return "AWARD"
    if "sources sought" in nt:
        return "SOURCES_SOUGHT"
    if "presolicitation" in nt or "pre-solicitation" in nt or "presol" in nt:
        return "PRESOL"
    if "request for information" in nt or nt.strip() == "rfi":
        return "RFI"
    if "special notice" in nt:
        return "SPECIAL"
    if "forecast" in nt:
        return "FORECAST"
    return "SOLICITATION"

# --- The single most important axis: HIREABLE credential vs UNHIREABLE
# authorization (PRG's win-and-manage model) ---------------------------------
#
# TYPE A — INDIVIDUAL CREDENTIAL (a licensed/certified PERSON). PRG can hire the
# person W-2 (their labor counts toward the 50% self-performance rule) and
# project-manage. POTENTIALLY WINNABLE → route to WATCH-TEMPLATE, never kill.
# The test: can the capability be supplied by ONE individual on PRG's payroll?
INDIVIDUAL_CREDENTIAL_TEMPLATE = [
    "professional engineer", "p.e. stamp", "pe stamp", "registered architect",
    "state medical license", "rn license", "registered nurse license",
    "physician license", "licensed physician", "pharmacist license",
    "certified industrial hygienist", "board certified", "board-certified",
    "medical physicist", "health physicist", "radiation safety officer",
    "licensed counselor", "licensed clinical", "licensed psychologist",
    "licensed social worker", "certified professional coder",
    "american board of radiology", "licensed dietitian", "licensed physicist",
    # Accredited appraisers (fine art / personal property / real estate) — a
    # hireable individual credential (ASA/AAA/ISA, USPAP-compliant), not a
    # firm license. Route to hire-to-win, never a silent score-down.
    "accredited appraiser", "certified appraiser", "qualified appraiser",
    "uspap", "appraiser",
]
# TYPE B — CORPORATE / OEM / PROPRIETARY authorization (a certified COMPANY).
# The authorization belongs to a corporation PRG cannot put on payroll; subbing
# the principal purpose to that firm breaches the 50% limitation on
# subcontracting (illegal pass-through). Cannot be solved by hiring one person →
# HARD KILL. (OEM_PASSTHROUGH_KILL is folded in below.)
CORP_AUTHORIZATION_KILL = [
    "authorized dealer", "authorized distributor of", "channel partner",
    "written authorization from the manufacturer", "manufacturer's authorization",
    "manufacturer authorization", "certified by the manufacturer",
    "sole authorized", "proprietary system", "proprietary building automation",
    "proprietary software license required",
    # named proprietary control / calibration systems that require the offeror to
    # BE a corporate-authorized provider (unhireable via one W-2 person):
    "metasys", "johnson controls", "cepheid", "fluke", "siemens desigo",
    "honeywell ebi", "niagara framework", "andover continuum",
]
# IMPROVEMENT 2 — capital / workforce that must exist BEFORE award. PRG has no
# upfront capital: it can only field ONE contingent specialist (or its own
# labor). Anything needing multiple trained/certified technicians, a trade crew,
# owned equipment / a lab, or a facility at proposal time → HARD KILL.
CAPITAL_WORKFORCE_KILL = [
    # "asse 6030"-style certs only — bare "asse" would substring-match
    # "assembly"/"assessment" and mis-kill in-lane work.
    "asse 6", "asse certified", "asse-certified", "medical gas", "nfpa 99",
    "licensed electrician", "master electrician", "journeyman",
    "licensed plumber", "hvac certification", "epa 608", "boiler operator",
    "asbestos", "lead abatement", "stationary engineer",
    "commercial driver", "cdl",
    "certified technicians", "trained technicians", "multiple technicians",
    "team of certified", "crew of", "iso 17025", "iso/iec 17025",
    "accredited calibration", "calibration laboratory", "a2la accredited",
    "must own equipment", "offeror must own", "own a laboratory",
    "furnish all equipment and personnel", "provide all tools and equipment",
]
# GATE 1 — workforce / CBA signals: incumbent-workforce takeover PRG can't staff.
WORKFORCE_KILL_KEYWORDS = [
    "collective bargaining", "wage determination", "service contract act",
    "phase-in", "phase in period", "phase-out", "phaseout",
    "right of first refusal", "incumbent workforce", "successor contractor",
    "e.o. 14055", "eo 14055", "seniority list",
]
# GATE 5 — bonding / capital requirements beyond a laptop.
BONDING_KILL_KEYWORDS = [
    "bid bond", "performance bond", "payment bond", "bid guarantee",
    "irrevocable letter of credit", "bonding requirement",
]
# Score-UP signals — first-contract-winnable (LPTA, remote, neutral PP rating).
SCORE_UP_KEYWORDS = [
    "lowest price technically acceptable", "lpta", "price will be the primary",
    "remote", "telework", "hybrid", "virtual", "neither favorably nor unfavorably",
    "neutral rating", "no minimum past performance",
]
# Past-performance wall — corporate-level PP requirements a first-timer can't meet.
PAST_PERF_WALL_KEYWORDS = [
    "cpars", "past performance questionnaire", "similar in size, scope",
    "similar size and scope", "recent and relevant past performance",
    "corporate experience", "minimum of three references",
]

# FIX 1/5 — specialty field-science / permit work PRG's solo clinical founder
# cannot do: cultural-resource/archaeology & environmental fieldwork requiring
# agency permits, Secretary-of-the-Interior professional qualification standards,
# or sustained on-site field crews. (Phrases chosen to NOT collide with clinical
# research: e.g. clinical "site visits" and a clinical "principal investigator"
# must not trip this.)
SPECIALTY_FIELD_KILL = [
    "cultural resource", "cultural resources", "class iii inventory",
    "class ii inventory", "class i inventory", "cultural resource use permit",
    "crup", "secretary of the interior", "secretary of interior", "soi-qualified",
    "36 cfr 61", "shpo", "state historic preservation", "section 106", "nhpa",
    "arpa permit", "archaeolog", "paleontolog", "field school",
    "historic preservation", "qualified biologist", "esa section 7",
    "endangered species act", "wetland delineation", "certified arborist",
    "certified forester", "habitat survey", "botanical survey", "avian survey",
    "threatened and endangered species", "biological survey",
    # Fix 5 — unambiguous physical field-crew signals (never used in clinical work)
    "pedestrian survey", "transect", "field crew", "high-clearance vehicle",
]
# "Principal investigator" only disqualifies in a field-science context — it is
# normal (and core) in clinical research, so require an archaeology/cultural cue.
_PI_FIELD_CONTEXT = ["archaeolog", "anthropolog", "cultural resource",
                     "historic preservation", "paleontolog"]

# Pass-through / OEM-specific maintenance trap — the sharpest LOS trap: brand-
# named equipment maintenance/calibration/repair to manufacturer spec needs a
# lab or manufacturer authorization PRG cannot self-perform. Hard kill
# regardless of set-aside or dollar size.
OEM_PASSTHROUGH_KILL = [
    "authorized service center", "authorized service provider",
    "designated service center", "factory authorized", "factory-authorized",
    "oem authorized", "manufacturer authorized service", "manufacturer's authorized",
    "authorized repair facility", "original equipment manufacturer service",
]
_OEM_WORK = ["maintenance", "repair", "calibration", "preventive maintenance",
             "overhaul", "refurbish"]
_OEM_SPEC = ["manufacturer specification", "manufacturer's specification",
             "oem specification", "factory specification", "per oem", "per the oem"]
# Product/software resale contexts that must NOT trip the OEM-maintenance kill.
OEM_EXEMPT_CONTEXT = ["saas", "software license", "subscription", "reseller",
                      "letter of supply", "software as a service"]
# Equipment-repair NAICS families (PRG owns no bench) — score down categorically.
EQUIP_REPAIR_NAICS_PREFIXES = ("8112", "8113")

# PRG OPPORTUNITY SCREENING RULE (mirrors CLAUDE.md on main) — five kill
# criteria, any one = NO-BID. KC2 (self-performance / pass-through), KC3
# (physical delivery), and KC4 (dead timeline) are already enforced by the
# existing gates (Type B / capital / workforce / bonding / specialty /
# coverage / FTE self-perform, and the stage-0 timing gate). KC1 and KC5
# below close the two remaining structural gaps.
#
# KC1 — NAICS/PSC families PRG can never perform (physical delivery of
# construction or manufactured goods). Construction (23xxxx, incl. 238 trades)
# and manufacturing (31-33xxxx) NAICS, and Y/Z PSC codes (construction /
# maintenance of real property), are automatic kills regardless of set-aside.
KC1_NAICS_KILL_PREFIXES = ("23", "31", "32", "33")
KC1_PSC_KILL_PREFIXES = ("Y", "Z")
# KC5 — wrong scale: a single-member LLC cannot self-perform the required
# share (LOS 50% services) of a contract above this ceiling.
KC5_MAX_SOLO_VALUE = 10_000_000
# Dollar figures that are small-business SIZE STANDARDS ("size standard:
# $34.5 million"), not contract values — stripped before the KC5 check.
_SIZE_STANDARD_RE = re.compile(r"[^.]{0,40}size\s+standard[^.]{0,80}",
                               re.IGNORECASE)

# ---------------------------------------------------------------------------
# PRG RUBRIC v2 — three-gate model + weighted scoring.
# Business model: win as prime, hire 1099/W2 labor, founder PMs the work.
# The founder does NOT need to personally perform delivery; founder
# credentials matter as (a) personal-delivery capability on a subset and
# (b) PM-credibility signal on the rest.
# ---------------------------------------------------------------------------

# GATE 0 — product purchase / subscription / license RESALE with no meaningful
# services labor. No labor to arbitrage → hard kill.
RESALE_KILL_KEYWORDS = [
    "subscription", "subscriptions", "license renewal", "renewal of license",
    "software license", "perpetual license", "saas", "reseller",
    "letter of supply", "database access", "journal access", "ebsco",
    "site license", "annual license", "licenses for", "subscription renewal",
]
# Services-labor signals that rescue a "license/subscription" notice from the
# resale kill (implementation/integration/help-desk work is real labor).
RESALE_SERVICES_EXEMPT = [
    "implementation services", "integration services", "configuration",
    "migration services", "help desk", "training services", "development",
    "customization", "support staff", "labor hours", "personnel shall",
]

# GATE 0 — firm-level professional licensure PRG does not hold (land
# surveying, PE stamp, architecture) — kill unless the solicitation allows
# the license to be held by key personnel.
FIRM_LICENSURE_KILL = [
    "professional engineer", "pe stamp", "stamped by a licensed",
    "licensed land surveyor", "professional land surveyor",
    "land surveying license", "registered land surveyor",
    "licensed architect", "registered architect", "architect-engineer",
    "a-e firm", "licensed engineering firm", "firm must be licensed",
]
KEY_PERSONNEL_LICENSE_OK = [
    "may be held by key personnel", "key personnel may hold",
    "licensed key personnel", "key personnel shall hold",
    "licensure may be met by key personnel", "held by an employee",
]

# GATE 2 — does a hirable labor market exist? (0-25)
LABOR_COMMODITY = [           # 20-25: deep 1099/W2 professional market
    "analyst", "analysts", "writer", "technical writer", "editor",
    "communications specialist", "communication specialist", "comms",
    "public affairs", "survey researcher", "survey research",
    "biostatistician", "statistician", "data scientist", "trainer",
    "instructor", "facilitator", "administrative support", "admin support",
    "research assistant", "program analyst", "management analyst",
    "graphic design", "transcription", "translator", "outreach specialist",
]
LABOR_SPECIALIZED = [         # 10-19: findable but specialized — lead time
    "biomedical equipment technician", "biomedical technician", "bmet",
    "clinical research coordinator", "clinical research associate",
    "regulatory specialist", "regulatory affairs", "epidemiologist",
    "industrial hygienist", "health physicist", "gis analyst",
    "environmental scientist", "environmental specialist", "logistician",
    "medical coder", "credentialed", "subject matter expert",
]
LABOR_THIN = [                # 1-9: thin market / geographic / OEM-cert bound
    "oem certified", "factory trained", "factory certified",
    "manufacturer certified", "remote duty station", "rotating personnel",
    "cleared personnel", "polygraph",
]

# GATE 3 — founder PM literacy (0-25). NOT "can founder do the work" — can the
# founder spec it, QA it, and talk to the CO without a translator.
FOUNDER_DOMAIN_KEYWORDS = [   # 20-25: inside a founder credential domain
    # Clinical research operations (Stanford, 1,500+ participants, NIH +
    # industry, 4x JAMA; ACRP-PM/CP, GCP)
    "clinical research", "clinical trial", "clinical study", "gcp",
    "human subjects", "irb", "clinical data", "patient recruitment",
    "decentralized trial", "telehealth", "econsent", "redcap", "ctms",
    "electronic data capture", "data management plan", "honest broker",
    # Biomedical equipment / healthcare technology (USAF 4A2X5, CBET,
    # >98% PM compliance on 200+ devices, Joint Commission/DoD readiness)
    "biomedical equipment", "medical equipment", "medical device",
    "healthcare technology", "hospital equipment", "clinical engineering",
    "accreditation readiness", "joint commission", "medical readiness",
    # Emergency medicine / EMS (Level I trauma center, NREMT, ACLS/PALS,
    # PHTLS/TCCC; CPT-I phlebotomy)
    "emergency medical", "emergency department", "triage", "ems",
    "emergency medical technician", "phlebotomy", "specimen collection",
    "vaccination", "immunization", "health screening", "covid-19 testing",
    "mass casualty", "medical surge",
    # Refugee / humanitarian / evacuation operations (Operation Allies
    # Welcome — 25K+ evacuees; interagency, CDC compliance)
    "evacuation", "evacuee", "refugee", "resettlement", "humanitarian",
    "repatriation", "displaced persons",
    # International / diplomatic (MA International Studies; Brazil, Peru,
    # China field experience)
    "global health", "international health", "diplomatic", "embassy program",
    "spanish language", "latin america", "foreign affairs",
    "international development", "capacity building",
    # Cultural heritage / humanities (BA Classics) — PM literacy for
    # collections work; the appraiser credential itself is hire-to-win.
    "cultural heritage", "heritage collection", "museum", "curatorial",
    "collections management", "archival", "archives", "fine art",
    "works of art", "art collection", "classical",
    # Research analysis / evaluation / technical writing (SAS, SPSS, R,
    # Python; Google Advanced Data Analytics)
    "program evaluation", "longitudinal study", "research analysis",
    "literature review", "systematic review", "technical writing",
    "technical documentation", "health program", "public health program",
    "statistical modeling", "study design", "protocol development",
    "data analytics", "statistical analysis",
]
ADJACENT_DOMAIN_KEYWORDS = [  # 10-19: founder can spec/QA without a translator
    "program support", "program management support", "data analysis",
    "data management", "training coordination", "training development",
    "curriculum", "communications", "communication support", "outreach",
    "stakeholder engagement", "strategic planning", "policy analysis",
    "market research", "survey", "needs assessment", "report preparation",
    "administrative support", "project management", "management consulting",
    "knowledge management", "conference support",
    # Resume-adjacent: healthcare admin, education, records/inventory, QA
    "medical records", "records management", "credentialing",
    "medical staffing", "health education", "education services",
    "instruction", "tutoring", "quality assurance", "quality control",
    "process improvement", "standard operating procedures", "inventory",
    "medical courier", "specimen transport", "laboratory support",
    "public health", "case management", "enrollment support",
    "translation services", "interpretation services",
]
LOW_LITERACY_KEYWORDS = [     # 1-9: founder cannot judge work quality
    "environmental sampling", "environmental compliance",
    "environmental remediation", "air monitoring", "soil sampling",
    "groundwater",
    "it infrastructure", "network engineering", "network infrastructure",
    "cybersecurity engineering", "cloud migration", "data center",
    "financial audit", "financial statement audit", "actuarial",
    "geotechnical", "hydrology", "utility survey", "land survey",
]

# Weighted modifiers.
DELIVERABLE_SCOPE_KEYWORDS = [   # deliverable-based vs level-of-effort
    "report", "reports", "study", "studies", "analysis", "assessment",
    "evaluation report", "white paper", "deliverables", "literature review",
    "manuscript", "curriculum", "training materials", "plan development",
    "modeling", "data set", "dataset", "final report",
]
LOE_SCOPE_KEYWORDS = [
    "full-time", "full time", "man-hours", "labor hours", "level of effort",
    "personnel shall be on-site", "staff augmentation", "seat support",
    "positions", "billets",
]
RECURRING_KEYWORDS = [
    "option year", "option years", "option period", "base year",
    "base plus", "multi-year", "multiyear", "ordering period", "idiq",
]
ONSITE_KEYWORDS = [
    "on-site", "onsite", "on site", "government facility", "duty station",
    "contractor personnel shall report",
]
_SOCAL_STATES = {"CA"}           # crude: any CA PoP skips the -5 travel hit

# Rating bands (rubric v2): Green = 75+, Yellow = 55-74, Red = below 55 or
# any Gate 0 hit.
RUBRIC_BASE = 30                 # for passing Gates 0 and 1
GREEN_MIN = 75
YELLOW_MIN = 55

# ---------------------------------------------------------------------------
# CAPTURE MODEL v3 — contract AGGREGATION, management, and execution.
# PRG primes and owns the customer relationship, compliance, QC, and
# invoicing; execution is sourced through subs, 1099s, staffing firms,
# trade contractors, distributors, and teaming partners. Opportunities are
# judged on WIN → SOURCE → PERFORM → CONTROL → INVOICE → COLLECT → PROFIT,
# not on whether the founder can personally do the labor.
# ---------------------------------------------------------------------------

# Value-added components that rescue a product/supply buy from the pure-
# commodity-resale kill (the aggregator adds margin through these).
VALUE_ADD_KEYWORDS = [
    "installation", "install and", "assembly", "commissioning", "site prep",
    "removal and disposal", "haul away", "training", "configuration",
    "integration", "maintenance plan", "warranty service",
    "delivery and setup", "setup", "white glove", "inside delivery",
]

# Delegable trade / field services — NOT kills under the capture model.
# Deep local sub markets exist; PRG primes, manages, and QCs. Flagged so the
# margin model prices sub-management (thinner) rather than professional labor.
DELEGABLE_TRADE_KEYWORDS = [
    "janitorial", "custodial", "housekeeping", "grounds maintenance",
    "landscaping", "pest control", "snow removal", "facility maintenance",
    "facilities maintenance", "hvac", "plumbing", "electrical work",
    "painting services", "roofing", "paving", "asphalt", "concrete",
    "striping", "site work", "drainage", "fencing", "tree removal",
    "welding", "carpentry", "masonry", "boiler", "elevator",
    "generator maintenance", "fire alarm", "fire suppression",
    "equipment maintenance", "equipment repair", "preventive maintenance",
    "calibration services", "install and maintain",
    "installation and maintenance", "logistics", "distribution", "warehouse",
    "supply chain", "freight", "courier", "fleet", "food service",
    "cafeteria", "laundry", "groundskeeping", "demolition", "renovation",
    "construction", "build-out", "new construction", "utilities", "pipe",
    "pumps", "water systems",
]

# Guard services need a firm-level state license (e.g., CA PPO) the PRIME
# must hold — not delegable to a sub without ostensible-subcontractor risk.
GUARD_LICENSE_KILL = [
    "security guard", "guard services", "armed guard", "unarmed guard",
    "security officer services", "protective services officer",
]

# Miller Act: performance/payment bonds are mandatory on federal construction
# above this. PRG has no bonding capacity — construction above it is a kill;
# below it, construction is a prime-with-trade-sub play (self-perform ≥15%).
MILLER_ACT_THRESHOLD = 150_000

# Capture priority weights (sum = 1.0) for the 1-10 dimension scores.
CAPTURE_WEIGHTS = {
    "p_win": 0.25, "margin": 0.25, "delegation": 0.15, "speed_cash": 0.10,
    "perf_risk": 0.10, "wc_burden": 0.05, "pp_val": 0.05, "scalability": 0.05,
}

# International / non-US-government buyers. None recognize US small-business or
# SDVOSB set-asides — PRG's main structural advantage evaporates — so score DOWN
# and flag (never first-contract material). Not a hard kill: a specific one may
# matter to PRG's international ambitions someday.
INTERNATIONAL_BUYER_KEYWORDS = [
    "nato", "ncia", "nato communications", "nato support", "nspa",
    "shape", "allied command", "united nations", "u.n. ", "unicef", "undp",
    "unhcr", "unops", "unesco", "world health organization",
    "world bank", "international organization", "foreign ministry",
    "ministry of", "embassy of", "european union", "european commission",
    "african union", "pan american health", "paho",
]
# US diplomatic posts overseas. FAR 19.000(b): small-business set-aside programs
# generally do not apply to acquisitions for use OUTSIDE the US, so PRG's SDVOSB
# edge does not carry at an embassy/consulate buying for in-country delivery.
# Score DOWN + flag (not a kill) only when paired with a foreign place of
# performance, so CONUS State Department work is unaffected.
OVERSEAS_POST_KEYWORDS = [
    "embassy", "consulate", "consulate general", "diplomatic post",
    "u.s. mission", "us mission", "overseas post",
]

# Priority agencies PRG actively targets — their OSDBU offices run active SDVOSB
# outreach. The OSDBU sites (osdbu.va.gov, osdbu.hhs.gov, dtra.mil) are forecast/
# outreach PORTALS with no clean public API, but the underlying solicitations
# post to SAM.gov (we already pull them by NAICS) and the awards/incumbents are
# in USASpending. So we filter both feeds by agency, give each its own tab, and
# link the portal for manual forecast review. `match` is checked against the
# notice's full parent-path / office / title (a sub-agency like DTRA never shows
# in the trimmed department name). `usaspending_tier` picks toptier vs subtier
# for the awarded-contract (Recompete Radar) pull.
AGENCY_TARGETS = [
    {
        "key": "VA", "label": "VA (OSDBU)",
        "match": ("veterans affairs", "veterans health", "veterans benefits",
                  "department of veterans"),
        "usaspending_name": "Department of Veterans Affairs",
        "usaspending_tier": "toptier",
        "portal": "https://osdbu.va.gov/",
        "portal_note": "VA OSDBU — VIP/SDVOSB verification, events & forecast",
    },
    {
        "key": "HHS", "label": "HHS (OSDBU)",
        "match": ("health and human services", "national institutes of health",
                  "centers for disease control", "food and drug administration",
                  "indian health service", "health resources and services",
                  "substance abuse and mental health"),
        "usaspending_name": "Department of Health and Human Services",
        "usaspending_tier": "toptier",
        "portal": "https://osdbu.hhs.gov/",
        "portal_note": "HHS OSDBU — vendor outreach, forecast & events",
    },
    {
        "key": "DTRA", "label": "DTRA",
        "match": ("defense threat reduction", "dtra"),
        "usaspending_name": "Defense Threat Reduction Agency",
        "usaspending_tier": "subtier",
        "portal": "https://www.dtra.mil/Mission/Acquisition/",
        "portal_note": "DTRA acquisition & small-business programs",
    },
]


def _agency_bucket(opp):
    """Return the AGENCY_TARGETS key (VA/HHS/DTRA) an opportunity belongs to, or
    "". Checks the full parent path / office / title, not just the trimmed
    department name, so a sub-agency (e.g. DTRA under DoD) is still detected."""
    blob = " ".join(str(opp.get(k, "")) for k in (
        "fullParentPathName", "organizationName", "departmentName",
        "subTier", "office", "officeAddress", "title")).lower()
    for a in AGENCY_TARGETS:
        if any(m in blob for m in a["match"]):
            return a["key"]
    return ""


def _agency_bucket_name(*names):
    """Bucket one or more agency-name strings (e.g. USASpending awarding agency /
    sub-agency) into an AGENCY_TARGETS key, or ""."""
    low = " ".join(str(n or "") for n in names).lower()
    for a in AGENCY_TARGETS:
        if any(m in low for m in a["match"]):
            return a["key"]
    return ""


# ---------------------------------------------------------------------------
# RESOURCE DIRECTORY — every place a federal contractor should hunt, research,
# get certified, and get FREE help. Rendered as an HTML section + Excel tab so
# PRG has one-click access to the whole ecosystem, not just this tool's output.
# URLs are stable roots; deep OSDBU pages move, so a few point at the department
# home with a note on what to click.
# ---------------------------------------------------------------------------
RESOURCE_DIRECTORY = [
    # --- Find live opportunities ---
    ("Find Opportunities", "SAM.gov — Contract Opportunities",
     "https://sam.gov/content/opportunities",
     "The system of record. Every federal RFP / RFQ / sources-sought / presol."),
    ("Find Opportunities", "SAM.gov — saved searches & alerts",
     "https://sam.gov/content/opportunities",
     "Save your NAICS/set-aside filters and get daily email alerts (free login)."),
    ("Find Opportunities", "GSA eBuy",
     "https://www.ebuy.gsa.gov/",
     "RFQs issued to GSA Schedule holders — requires a GSA Schedule to see/quote."),
    ("Find Opportunities", "DLA Internet Bid Board (DIBBS)",
     "https://www.dibbs.bsm.dla.mil/",
     "Defense Logistics Agency solicitations for supplies & services."),
    ("Find Opportunities", "Unison Marketplace",
     "https://marketplace.unisonglobal.com/",
     "Government reverse-auction micro-purchases and simplified buys."),
    ("Find Opportunities", "SBIR / STTR",
     "https://www.sbir.gov/",
     "R&D innovation awards set aside for small business; annual agency cycles."),
    ("Find Opportunities", "Grants.gov",
     "https://www.grants.gov/",
     "Federal grants & cooperative agreements (research funding, not contracts)."),
    ("Find Opportunities", "FedConnect",
     "https://www.fedconnect.net/",
     "Some agencies post opportunities & manage awards here in addition to SAM."),
    ("Find Opportunities", "eCFR / FAR (rules of the game)",
     "https://www.acquisition.gov/browse/index/far",
     "The Federal Acquisition Regulation — the rulebook every solicitation runs on."),
    # --- Research awards, incumbents & competitors ---
    ("Market Intel", "USAspending.gov — Advanced Search",
     "https://www.usaspending.gov/search",
     "Every federal award: incumbents, dollar amounts, and PoP end dates (recompetes)."),
    ("Market Intel", "USAspending API (free, no key)",
     "https://api.usaspending.gov/",
     "The keyless API this tool uses for the Recompete Radar. Build your own queries."),
    ("Market Intel", "FPDS-NG (Federal Procurement Data System)",
     "https://www.fpds.gov/",
     "Raw contract-action data — deep incumbent & agency-spend research."),
    ("Market Intel", "SAM.gov — Entity Search",
     "https://sam.gov/search",
     "Look up any registered contractor: competitors, primes, potential partners."),
    ("Market Intel", "SBA Dynamic Small Business Search (DSBS)",
     "https://web.sba.gov/pro-net/search/dsp_dsbs.cfm",
     "Find teaming partners & competitors by NAICS, certification, and location."),
    ("Market Intel", "GSA eLibrary",
     "https://www.gsaelibrary.gsa.gov/",
     "Who holds which GSA Schedules — find primes to team with or subcontract under."),
    ("Market Intel", "SBA SubNet",
     "https://web.sba.gov/subnet/search/index.cfm",
     "Subcontracting opportunities posted by large prime contractors."),
    ("Market Intel", "USAspending — Recipient Profiles",
     "https://www.usaspending.gov/recipient",
     "Full award history for any company (size up an incumbent before you bid)."),
    # --- Get & stay eligible (registrations + certifications) ---
    ("Certify & Register", "SAM.gov — Entity Registration (UEI)",
     "https://sam.gov/content/entity-registration",
     "Mandatory: get your UEI and active registration BEFORE any award. Free."),
    ("Certify & Register", "SBA Certify (8(a), WOSB/EDWOSB, HUBZone)",
     "https://certify.sba.gov/",
     "Apply for and manage SBA socioeconomic certifications."),
    ("Certify & Register", "SBA VetCert (SDVOSB / VOSB)",
     "https://veterans.certify.sba.gov/",
     "SDVOSB/VOSB verification — REQUIRED to win VA & federal SDVOSB set-asides."),
    ("Certify & Register", "SBA HUBZone Map",
     "https://maps.certify.sba.gov/hubzone/",
     "Check whether an address qualifies for the HUBZone program."),
    ("Certify & Register", "SBA Size Standards tool",
     "https://www.sba.gov/size-standards",
     "Confirm you qualify as 'small' for each NAICS you bid."),
    # --- Free help, counseling & teaming ---
    ("Free Help", "APEX Accelerators (formerly PTAC)",
     "https://www.apexaccelerators.us/",
     "FREE 1:1 government-contracting counseling: bid-matching, proposal & pricing help."),
    ("Free Help", "Veterans Business Outreach Center (VBOC)",
     "https://www.sba.gov/local-assistance/resource-partners/veterans-business-outreach-center-vboc-program",
     "SBA veteran-focused business counseling and training."),
    ("Free Help", "Small Business Development Centers (SBDC)",
     "https://americassbdc.org/",
     "Free local business advising, including federal-contracting readiness."),
    ("Free Help", "SCORE mentoring",
     "https://www.score.org/",
     "Free mentors, many with government-contracting experience."),
    ("Free Help", "SBA — Local Assistance / District Offices",
     "https://www.sba.gov/local-assistance",
     "Find your SBA district office & Procurement Center Representative (PCR)."),
    ("Free Help", "SBA — Federal Contracting guide",
     "https://www.sba.gov/federal-contracting",
     "Plain-English overviews of set-asides, joint ventures, and the bid process."),
    # --- Agency OSDBU / Small-Business offices (the outreach doors) ---
    ("Agency OSDBU", "VA OSDBU", "https://osdbu.va.gov/",
     "Veterans Affairs — SDVOSB outreach, VIP verification, events & forecast."),
    ("Agency OSDBU", "HHS OSDBU", "https://osdbu.hhs.gov/",
     "Health & Human Services (NIH, CDC, FDA) vendor outreach & forecast."),
    ("Agency OSDBU", "DoD Office of Small Business Programs", "https://business.defense.gov/",
     "Department of Defense gateway to all military-service small-business offices."),
    ("Agency OSDBU", "DTRA — Acquisition & Small Business", "https://www.dtra.mil/",
     "Defense Threat Reduction Agency acquisition & small-business programs."),
    ("Agency OSDBU", "Army OSBP", "https://www.sellingtoarmy.army.mil/",
     "Department of the Army small-business office."),
    ("Agency OSDBU", "Navy OSBP", "https://www.secnav.navy.mil/smallbusiness/",
     "Department of the Navy / Marine Corps small-business office."),
    ("Agency OSDBU", "Air Force Small Business", "https://www.airforcesmallbiz.af.mil/",
     "Department of the Air Force / Space Force small-business office."),
    ("Agency OSDBU", "DHS OSDBU", "https://www.dhs.gov/osdbu",
     "Homeland Security small-business office (CBP, FEMA, TSA, USCG, etc.)."),
    ("Agency OSDBU", "DOE OSDBU", "https://www.energy.gov/osdbu",
     "Department of Energy and national labs small-business office."),
    ("Agency OSDBU", "USDA OSDBU", "https://www.dm.usda.gov/smallbus/",
     "Department of Agriculture small-business office."),
    ("Agency OSDBU", "DOI OSDBU", "https://www.doi.gov/osdbu",
     "Department of the Interior (BLM, NPS, USGS, etc.) small-business office."),
    ("Agency OSDBU", "Commerce OSDBU", "https://www.commerce.gov/oam",
     "Department of Commerce (NOAA, Census, NIST) small-business office."),
    ("Agency OSDBU", "DOT OSDBU", "https://www.transportation.gov/osdbu",
     "Department of Transportation (FAA, FHWA) small-business office."),
    ("Agency OSDBU", "DOJ Small Business", "https://www.justice.gov/jmd/small-business",
     "Department of Justice (FBI, BOP, DEA) small-business office."),
    ("Agency OSDBU", "State Dept OSDBU", "https://www.state.gov/",
     "Department of State — search 'OSDBU' for small-business & embassy buys."),
    ("Agency OSDBU", "Treasury OSDBU", "https://home.treasury.gov/services/small-business-programs",
     "Department of the Treasury (IRS, Mint) small-business programs."),
    ("Agency OSDBU", "Education OSDBU", "https://www.ed.gov/",
     "Department of Education — search 'OSDBU' (IES/SBIR research buys)."),
    ("Agency OSDBU", "Labor OSDBU", "https://www.dol.gov/agencies/oasam/centers-offices/osdbu",
     "Department of Labor small-business office."),
    ("Agency OSDBU", "HUD OSDBU", "https://www.hud.gov/program_offices/osdbu",
     "Housing & Urban Development small-business office."),
    ("Agency OSDBU", "EPA Office of Small Business Programs", "https://www.epa.gov/osbp",
     "Environmental Protection Agency small-business office."),
    ("Agency OSDBU", "NASA Office of Small Business Programs", "https://www.nasa.gov/osbp/",
     "National Aeronautics & Space Administration small-business office."),
    ("Agency OSDBU", "GSA — Small Business", "https://www.gsa.gov/small-business",
     "General Services Administration OSDBU + Schedules on-ramp."),
    ("Agency OSDBU", "SSA OSDBU", "https://www.ssa.gov/oag/acq_osdbu.htm",
     "Social Security Administration small-business office."),
    ("Agency OSDBU", "NSF Small Business", "https://www.nsf.gov/od/oecr/",
     "National Science Foundation — research & SBIR/STTR buys."),
]

_RESOURCE_COLS = [
    ("Category", "category"),
    ("Resource", "name"),
    ("What it's for", "note"),
    ("URL", "url"),
]


def _resource_rows():
    """RESOURCE_DIRECTORY tuples -> row dicts for the Excel/HTML renderers."""
    return [{"category": c, "name": n, "url": u, "note": note}
            for (c, n, u, note) in RESOURCE_DIRECTORY]


def _intl_source_rows():
    """INTERNATIONAL_SOURCES tuples -> row dicts."""
    return [{"category": c, "name": n, "url": u, "note": note}
            for (c, n, u, note) in INTERNATIONAL_SOURCES]


def _sled_source_rows():
    """SLED_MARKETPLACE_SOURCES tuples -> row dicts."""
    return [{"category": c, "name": n, "url": u, "note": note}
            for (c, n, u, note) in SLED_MARKETPLACE_SOURCES]


# Columns for the other-data-bank tabs (grants, SBIR, international directory).
_GRANTS_COLS = [
    ("Title", "title"),
    ("Agency", "agency"),
    ("Opportunity #", "number"),
    ("Status", "status"),
    ("Opens", "open_date"),
    ("Closes", "close_date"),
    ("ALN / CFDA", "aln"),
    ("URL", "link"),
]
_SBIR_COLS = [
    ("Solicitation", "title"),
    ("Program", "program"),
    ("Agency", "agency"),
    ("Branch", "branch"),
    ("Solicitation #", "number"),
    ("Opens", "open_date"),
    ("Closes", "close_date"),
    ("Topics", "topics"),
    ("URL", "link"),
]
_INTL_SRC_COLS = [
    ("Category", "category"),
    ("Source (register here)", "name"),
    ("What it covers / fit for PRG", "note"),
    ("URL", "url"),
]
_RELIEFWEB_COLS = [
    ("Title", "title"),
    ("Organization", "org"),
    ("Country", "country"),
    ("Category", "category"),
    ("Experience", "experience"),
    ("Closes", "closing"),
    ("URL", "link"),
]


# FIX 2 — the only credentials PRG actually holds (satisfy a "must hold" req).
HELD_CREDENTIALS = ["acrp-cp", "acrp-pm", "sdvosb", "sam registration", "sam.gov"]
# Contexts where "must hold X" is a product/reseller ASSET, not a personnel
# credential — do NOT kill (e.g. SaaS / software-license resale).
CREDENTIAL_EXEMPT_CONTEXT = [
    "reseller", "letter of supply", "oem authorization", "authorized distributor",
    "value added reseller", "software license", "saas", "subscription",
    "business license", "insurance", "manufacturer authorization",
]
# FIX 2 — generic "shall/must hold/possess a <permit|license|certification|
# registration> …" detector, so we stop playing blocklist whack-a-mole.
_CRED_REQUIRE_RE = re.compile(
    r"\b(shall|must|will)\b[^.\n]{0,60}?\b(hold|possess|maintain|have)\b"
    r"[^.\n]{0,70}?\b(permit|licens\w*|certification|registration|credential|"
    r"accreditation)\b", re.IGNORECASE)

# FIX 6 — content-hash cache so identical packages short-circuit as "already vetted".
CACHE_PATH = os.path.join(os.path.expanduser("~"), ".prg_screen_cache.json")

# Soft signals: score down / flag, never fail on their own.
BONDING_KEYWORDS = ["bid bond", "performance bond", "payment bond", "bonding"]
INCUMBENT_SOFT_KEYWORDS = ["incumbent", "recompete", "currently being performed",
                           "current contractor"]

# NAICS that are catch-all codes — flag PASS items as "verify scope in PWS".
CATCHALL_NAICS = {"541990", "561499", "541618"}

# Notice types that represent a completed award (a prime now exists to sub under).
AWARD_NOTICE_MARKERS = ("award",)

# USASpending.gov public API (keyless) — used for the Recompete Radar: finds
# existing contracts under PRG's NAICS that are expiring soon (upcoming rebids).
USASPENDING_URL = "https://api.usaspending.gov/api/v2/search/spending_by_award/"


# ---------------------------------------------------------------------------
# 2. API QUERY CONSTRUCTION
# ---------------------------------------------------------------------------

# Run-level flags surfaced to the user at the end (e.g. partial results).
_RUN_STATE = {"rate_limited": False}


def _is_http(url):
    """True only for http(s) URLs — guards against javascript:/data: in hrefs."""
    return bool(url) and str(url).lower().startswith(("http://", "https://"))


def _http_retry(fn, *args, **kwargs):
    """Call requests.get/post with up to 3 tries and exponential backoff on
    transient connection/timeout errors (not on HTTP status errors)."""
    delay = 1.0
    for attempt in range(3):
        try:
            return fn(*args, **kwargs)
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout):
            if attempt == 2:
                raise
            time.sleep(delay)
            delay *= 2

def fetch_opportunities(api_key, code, posted_from, posted_to, limit,
                        code_param="ncode"):
    """Query the SAM.gov Opportunities API for a single NAICS or PSC code.

    `code_param` is "ncode" for a NAICS code or "ccode" for a PSC (Product/
    Service) code. Returns a list of opportunity dicts (possibly empty).
    Network / API errors are caught and surfaced as an empty list plus a
    stderr warning so that one bad request does not abort the whole run.
    """
    params = {
        "api_key": api_key,
        code_param: code,             # ncode=NAICS or ccode=PSC
        "postedFrom": posted_from,    # MM/DD/YYYY (required)
        "postedTo": posted_to,        # MM/DD/YYYY (required)
        "limit": min(limit, 1000),    # API max page size is 1000
        "offset": 0,
        # Notice types: solicitation, presol, combined, sources sought, special,
        # and award notices ('a') so we can surface subcontracting/teaming targets.
        "ptype": "o,p,k,r,s,a",
    }

    collected = []
    try:
        while True:
            resp = _http_retry(requests.get, API_URL, params=params, timeout=30)
            if resp.status_code == 429:
                _RUN_STATE["rate_limited"] = True
                sys.stderr.write(
                    "WARNING: Rate limited by SAM.gov (HTTP 429). "
                    "Results are PARTIAL. Try again later or reduce --limit.\n"
                )
                break
            resp.raise_for_status()
            payload = resp.json()

            batch = payload.get("opportunitiesData") or []
            collected.extend(batch)

            total = payload.get("totalRecords", len(collected))
            params["offset"] += len(batch)
            if not batch or params["offset"] >= total or params["offset"] >= limit:
                break
    except requests.exceptions.HTTPError as exc:
        sys.stderr.write(
            f"WARNING: HTTP error querying {code_param} {code}: {exc}\n"
            f"         Response: {getattr(exc.response, 'text', '')[:300]}\n"
        )
    except requests.exceptions.RequestException as exc:
        sys.stderr.write(
            f"WARNING: Network error querying {code_param} {code}: {exc}\n"
        )
    except ValueError as exc:  # JSON decode error
        sys.stderr.write(
            f"WARNING: Could not decode JSON for {code_param} {code}: {exc}\n"
        )

    return collected


def fetch_recompetes(naics_codes, months_ahead=18, max_pages=10, agency=None):
    """Recompete Radar: find existing federal contracts under PRG's NAICS codes
    whose period of performance ends within `months_ahead` months — i.e. the
    upcoming rebids to position for early. Uses the keyless USASpending.gov API.

    When `agency` (an AGENCY_TARGETS dict) is given, the pull is filtered to that
    awarding agency (top- or sub-tier) so the incumbents + recompetes at a
    priority agency (VA / HHS / DTRA) are surfaced even if they'd fall outside
    the general soonest-first page window.

    Returns a list of dicts sorted by soonest end date. Network/parse errors
    degrade to an empty list so the rest of the run is unaffected.
    """
    today = dt.date.today()
    horizon = today + dt.timedelta(days=int(months_ahead * 30.4))
    # Sort by period-of-performance end date ASCENDING and page through, keeping
    # only contracts whose end date lands in [today, horizon]. Because results
    # are ascending, we can stop as soon as we pass the horizon. Restricting the
    # action-date window to the last ~3 years keeps the already-expired prefix
    # small. (Requesting only well-known field names avoids a 422 that would
    # zero out the whole feature.)
    filters = {
        "award_type_codes": ["A", "B", "C", "D"],   # definitive contracts
        "naics_codes": list(naics_codes),
        "time_period": [{
            "start_date": (today - dt.timedelta(days=365 * 3)).isoformat(),
            "end_date": today.isoformat(),
        }],
    }
    if agency:
        filters["agencies"] = [{
            "type": "awarding",
            "tier": agency.get("usaspending_tier", "toptier"),
            "name": agency["usaspending_name"],
        }]
    base = {
        "filters": filters,
        "fields": [
            "Award ID", "Recipient Name", "Award Amount",
            "Period of Performance Current End Date",
            "Awarding Agency", "Awarding Sub Agency", "NAICS",
        ],
        "sort": "Period of Performance Current End Date",
        "order": "asc",
        "limit": 100,
    }
    out = []
    try:
        for page in range(1, max_pages + 1):
            body = dict(base, page=page)
            resp = _http_retry(requests.post, USASPENDING_URL, json=body, timeout=45)
            resp.raise_for_status()
            results = resp.json().get("results") or []
            if not results:
                break
            passed_horizon = False
            for a in results:
                end_raw = (a.get("Period of Performance Current End Date")
                           or a.get("End Date") or "")
                try:
                    end = dt.date.fromisoformat(str(end_raw)[:10])
                except ValueError:
                    continue
                if end < today:
                    continue          # already expired — skip
                if end > horizon:
                    passed_horizon = True
                    break             # ascending: everything after is further out
                amt = a.get("Award Amount")
                try:
                    amt = float(amt) if amt not in (None, "") else None
                except (TypeError, ValueError):
                    amt = None
                # USASpending may return NAICS as a string or a
                # {"code": ..., "description": ...} dict — normalize to the code.
                naics_raw = a.get("NAICS")
                if isinstance(naics_raw, dict):
                    naics_val = naics_raw.get("code") or "—"
                else:
                    naics_val = naics_raw or "—"
                out.append({
                    "award_id": a.get("Award ID") or "N/A",
                    "recipient": a.get("Recipient Name") or "N/A",
                    "amount": amt,
                    "amount_display": _format_currency(amt) if amt else "Not stated",
                    "end_date": end.isoformat(),
                    "months_left": round((end - today).days / 30.4, 1),
                    "agency": a.get("Awarding Agency")
                              or a.get("Awarding Sub Agency") or "N/A",
                    "agency_bucket": agency["key"] if agency else _agency_bucket_name(
                        a.get("Awarding Agency"), a.get("Awarding Sub Agency")),
                    "naics": naics_val,
                })
            if passed_horizon or len(results) < 100:
                break
    except requests.exceptions.RequestException as exc:
        sys.stderr.write(f"WARNING: Recompete Radar (USASpending) unavailable: {exc}\n")
    except ValueError as exc:
        sys.stderr.write(f"WARNING: Recompete Radar JSON decode failed: {exc}\n")
    out.sort(key=lambda r: r["end_date"])
    return out


# --- Other opportunity banks with FREE, keyless APIs -------------------------
GRANTS_URL = "https://api.grants.gov/v1/api/search2"
SBIR_URL = "https://api.www.sbir.gov/public/api/solicitations"
RELIEFWEB_URL = "https://api.reliefweb.int/v1/jobs"

# Keywords for the Grants.gov pull — PRG's research / health / global-health lane.
GRANTS_KEYWORDS = [
    "clinical", "health", "global health", "research", "biomedical",
    "public health", "medical", "veteran", "emergency", "surveillance",
    "health security", "data", "evaluation",
]


def fetch_grants(keywords=None, max_pages=6, rows=100):
    """Pull open federal GRANT / cooperative-agreement opportunities from the
    keyless Grants.gov Search2 API (where State Dept global-health APS and other
    assistance awards live — these never touch SAM.gov). Grants are a different
    instrument than contracts, so they get their own tab; returns [] on any
    error so the rest of the run is unaffected."""
    kw = " ".join(keywords or GRANTS_KEYWORDS[:3])
    out, seen = [], set()
    try:
        for page in range(max_pages):
            body = {"keyword": kw, "oppStatuses": "forecasted|posted",
                    "rows": rows, "startRecordNum": page * rows}
            resp = _http_retry(requests.post, GRANTS_URL, json=body, timeout=45)
            resp.raise_for_status()
            data = (resp.json() or {}).get("data") or {}
            hits = data.get("oppHits") or []
            if not hits:
                break
            for h in hits:
                oid = str(h.get("id") or h.get("number") or "")
                if oid in seen:
                    continue
                seen.add(oid)
                aln = h.get("alnist") or h.get("cfdaList") or []
                out.append({
                    "number": h.get("number") or "N/A",
                    "title": _clean(h.get("title") or "Untitled", 80),
                    "agency": _clean(h.get("agency") or h.get("agencyCode") or "N/A", 40),
                    "open_date": h.get("openDate") or "",
                    "close_date": h.get("closeDate") or "",
                    "status": h.get("oppStatus") or "",
                    "aln": ", ".join(aln) if isinstance(aln, list) else str(aln),
                    "link": (f"https://grants.gov/search-results-detail/{oid}"
                             if oid else "https://grants.gov/search-grants"),
                })
            if len(hits) < rows:
                break
    except requests.exceptions.RequestException as exc:
        sys.stderr.write(f"WARNING: Grants.gov unavailable: {exc}\n")
    except ValueError as exc:
        sys.stderr.write(f"WARNING: Grants.gov JSON decode failed: {exc}\n")
    return out


def fetch_sbir(rows=200):
    """Pull OPEN SBIR/STTR solicitations from the keyless SBIR.gov API — the R&D
    innovation awards set aside for small business, squarely in PRG's research
    lane. Returns [] on any error."""
    out = []
    try:
        resp = _http_retry(requests.get, SBIR_URL,
                           params={"open": 1, "rows": rows}, timeout=45)
        resp.raise_for_status()
        payload = resp.json()
        items = payload if isinstance(payload, list) else (payload.get("data") or [])
        for s in items:
            if not isinstance(s, dict):
                continue
            topics = s.get("solicitation_topics") or []
            out.append({
                "title": _clean(s.get("solicitation_title") or "Untitled", 80),
                "number": s.get("solicitation_number") or "N/A",
                "program": s.get("program") or "SBIR/STTR",
                "agency": _clean(s.get("agency") or "N/A", 40),
                "branch": _clean(s.get("branch") or "", 30),
                "open_date": s.get("open_date") or "",
                "close_date": s.get("close_date") or "",
                "topics": (str(len(topics)) + " topic(s)"
                           if isinstance(topics, list) else ""),
                "link": (s.get("sbir_solicitation_link")
                         or s.get("solicitation_agency_url")
                         or "https://www.sbir.gov/solicitations"),
            })
    except requests.exceptions.RequestException as exc:
        sys.stderr.write(f"WARNING: SBIR.gov unavailable: {exc}\n")
    except ValueError as exc:
        sys.stderr.write(f"WARNING: SBIR.gov JSON decode failed: {exc}\n")
    return out


def fetch_reliefweb(rows=100):
    """Pull short-term international consultancies / jobs from the keyless
    ReliefWeb API (api.reliefweb.int) — UN & humanitarian postings, filtered to
    PRG's clinical/health/research lane. One of the few international banks with
    an open API. Returns [] on any error."""
    out = []
    body = {
        "limit": rows,
        "profile": "list",
        "fields": {"include": ["title", "source", "country", "date.closing",
                               "url", "url_alias", "career_categories",
                               "experience", "type"]},
        "query": {"value": ("health OR clinical OR research OR consultant OR "
                            "medical OR epidemiolog OR \"data management\" OR "
                            "monitoring OR evaluation"),
                  "operator": "OR"},
        "sort": ["date.created:desc"],
    }
    try:
        resp = _http_retry(requests.post, RELIEFWEB_URL,
                           params={"appname": "prg-opportunity-matcher"},
                           json=body, timeout=45)
        resp.raise_for_status()
        for item in (resp.json() or {}).get("data") or []:
            f = item.get("fields") or {}
            src = f.get("source") or []
            srcs = ", ".join(s.get("name", "") for s in src if isinstance(s, dict))
            ctry = f.get("country") or []
            ctrys = ", ".join(c.get("name", "") for c in ctry if isinstance(c, dict))
            cats = f.get("career_categories") or []
            cat = ", ".join(c.get("name", "") for c in cats if isinstance(c, dict))
            exp = f.get("experience") or []
            if isinstance(exp, list):
                exp = ", ".join(e.get("name", "") if isinstance(e, dict) else str(e)
                                for e in exp)
            closing = (f.get("date") or {}).get("closing") or ""
            out.append({
                "title": _clean(f.get("title") or "Untitled", 80),
                "org": _clean(srcs or "N/A", 40),
                "country": _clean(ctrys or "—", 30),
                "category": _clean(cat, 30),
                "experience": _clean(str(exp), 20),
                "closing": str(closing)[:10],
                "link": f.get("url_alias") or f.get("url")
                        or "https://reliefweb.int/jobs",
            })
    except requests.exceptions.RequestException as exc:
        sys.stderr.write(f"WARNING: ReliefWeb unavailable: {exc}\n")
    except ValueError as exc:
        sys.stderr.write(f"WARNING: ReliefWeb JSON decode failed: {exc}\n")
    return out


# --- International opportunity SOURCES (no clean public API → curated links) --
# Short-term international consultancies live on portals that don't talk to each
# other and mostly require a login, so they can't be live-pulled — but knowing
# WHERE to register is the whole game. Fit-notes are tuned to PRG's founder:
# ACRP clinical creds, Spanish, veteran + evacuation-ops background.
INTERNATIONAL_SOURCES = [
    ("UN system", "UNGM — UN Global Marketplace", "https://www.ungm.org/",
     "ONE registration covers UNDP, UNOPS, WHO, UNICEF, IOM. Short-term individual "
     "consultancies post constantly. WHO consultant rosters fit ACRP clinical "
     "creds; IOM/UNOPS fit evacuation-ops. **Register this week.**"),
    ("UN system", "WHO — Careers / Consultant rosters",
     "https://www.who.int/careers", "Consultancies & expert rosters — direct fit "
     "for clinical-research credentials."),
    ("UN system", "ReliefWeb — Jobs & Tenders", "https://reliefweb.int/jobs",
     "Humanitarian short-term assignments and consultancies across UN/NGOs "
     "(searchable; has a public API for the technically inclined)."),
    ("Development banks", "World Bank — eConsultant2",
     "https://wbgeconsult2.worldbank.org/", "Individual consultant contracts, "
     "often 20–60 day assignments."),
    ("Development banks", "Asian Development Bank — CMS",
     "https://cms.adb.org/", "ADB Consultant Management System — individual "
     "consultant assignments across Asia-Pacific."),
    ("Development banks", "Inter-American Development Bank (IDB)",
     "https://www.iadb.org/en/how-we-can-work-together/consultants",
     "Latin America / Caribbean consulting — **Spanish fluency is an edge here.**"),
    ("Aggregators", "Devex", "https://www.devex.com/",
     "The aggregator for the whole development sector: tenders, STTA gigs, "
     "consultant calls. Membership optional — wait until roster flow proves out."),
    ("Aggregators", "DevelopmentAid", "https://www.developmentaid.org/",
     "Cheaper Devex alternative — tenders, grants, and consultant opportunities."),
    ("Implementing partners", "Chemonics", "https://www.chemonics.com/careers/",
     "Holds big State/foreign-assistance contracts; staffs surge work from a "
     "consultant database. **Register once, get called.**"),
    ("Implementing partners", "DAI", "https://www.dai.com/careers",
     "Global development prime — consultant & short-term technical rosters."),
    ("Implementing partners", "Tetra Tech (International Development)",
     "https://www.tetratech.com/careers/", "Prime on health/dev contracts — "
     "surge staffing from rosters."),
    ("Implementing partners", "Abt Global", "https://www.abtglobal.com/careers",
     "Health-systems & research prime — consultant database."),
    ("Implementing partners", "FHI 360", "https://www.fhi360.org/careers/",
     "Global health implementer — clinical/research short-term assignments."),
    ("Implementing partners", "Palladium", "https://thepalladiumgroup.com/careers",
     "Health & data-for-development prime — consultant roster."),
    ("Implementing partners", "RTI International", "https://www.rti.org/careers",
     "Research institute + global health prime — STTA and consultancies."),
    ("Implementing partners", "Jhpiego (Johns Hopkins)", "https://www.jhpiego.org/careers/",
     "Global clinical/health implementer — direct fit for ACRP credentials."),
    ("Implementing partners", "MSH — Management Sciences for Health",
     "https://msh.org/careers/", "Health-systems strengthening — consultant calls."),
    ("US assistance", "Grants.gov (also pulled live in this report)",
     "https://grants.gov/search-grants",
     "State Dept global-health APS + assistance awards that never touch SAM.gov."),
]

# SLED (state/local/education) + federal marketplace sources. None expose a
# clean keyless API the way SAM.gov does, so these are REGISTER-HERE directory
# entries with a priority call — not live pulls. Ordered by leverage for PRG.
SLED_MARKETPLACE_SOURCES = [
    ("Federal — register first",
     "VA OSDBU — Forecast of Contracting Opportunities",
     "https://www.va.gov/osdbu/",
     "PRG's single highest-leverage agency: under 38 USC 8127 (Kingdomware) "
     "the VA MUST set aside for SDVOSB/VOSB when two capable firms exist. "
     "The forecast shows next-FY buys BEFORE they hit SAM.gov — preposition, "
     "call the OSDBU, respond to the sources sought. **Check monthly.**"),
    ("Federal — register first", "SBA SUBNet",
     "https://subnet.sba.gov/",
     "Subcontracting opportunities posted by large primes who NEED small-"
     "business/SDVOSB participation to meet FAR 52.219-9 plan goals — feeds "
     "the PURSUE AS SUBCONTRACTOR lane directly. Browsable without a login. "
     "**Sweep weekly.**"),
    ("Federal — register first", "Unison Marketplace (formerly FedBid)",
     "https://marketplace.unisonglobal.com/",
     "Reverse-auction simplified buys, many SDVOSB-set-aside, mostly under "
     "$250K — fast-cash past-performance builders. Discipline required: it's "
     "a price race, so bid only value-added supply/services where you hold a "
     "real supplier quote first. Free seller registration."),
    ("California / SLED — home turf",
     "Cal eProcure + CA DVBE certification (DGS)",
     "https://caleprocure.ca.gov/",
     "California's SAM.gov. **Highest-leverage state move: get the DVBE "
     "certification** — CA statute sets a 3% DVBE participation goal, so "
     "state primes actively hunt DVBE subs and agencies set contracts aside. "
     "Federal SDVOSB verification does most of the paperwork lift. Both "
     "prime AND sub demand, zero federal competition politics."),
    ("California / SLED — home turf", "PlanetBids",
     "https://www.planetbids.com/",
     "The bid portal for dozens of SoCal cities, counties, and special "
     "districts (one free vendor registration per agency). Start with the "
     "LA-metro agencies and register under services + supply commodity "
     "codes; DVBE preference often applies locally too."),
    ("California / SLED — home turf", "OpenGov Procurement",
     "https://procurement.opengov.com/",
     "Portal platform for a growing set of CA/AZ/NV municipalities — "
     "searchable public listings, register per agency. Sweep monthly for "
     "janitorial/facilities/program-support scopes in PRG's lanes."),
    ("California / SLED — home turf", "Bonfire (Euna Solutions)",
     "https://vendor.bonfirehub.com/",
     "One vendor account follows many agencies (cities, water/transit "
     "districts, universities). Good coverage of CA special districts — "
     "recurring facilities and professional-services RFPs."),
    ("Low priority — skip for now", "DLA DIBBS",
     "https://www.dibbs.bsm.dla.mil/",
     "Mostly exact-NSN manufactured-part supply governed by the "
     "nonmanufacturer rule — the pure no-value-add resale this report's own "
     "Gate 0 kills. Revisit only if PRG later builds a value-added medical "
     "supply niche with a distributor relationship."),
]


# ---------------------------------------------------------------------------
# 3. EVALUATION LOGIC
# ---------------------------------------------------------------------------

def naics_tier(code):
    """Classify a NAICS code into 'primary', 'low_barrier', or 'other'."""
    code = (code or "").strip()
    if code in TARGET_NAICS:
        return "primary"
    if code in LOW_BARRIER_NAICS:
        return "low_barrier"
    return "other"


def classify_setaside(opp):
    """Return (flag_label, eligible_bool, is_sdvosb_bool).

    flag_label   -> short human-readable set-aside label for the report.
    eligible     -> True if PRG (an SDVOSB small business) can bid.
    is_sdvosb    -> True if it is specifically an SDVOSB set-aside (best fit).
    """
    code = (opp.get("typeOfSetAside") or "").strip().upper()
    desc = (opp.get("typeOfSetAsideDescription") or "").strip()

    if code in SDVOSB_SETASIDE_CODES:
        return ("SDVOSB", True, True)
    if code in VOSB_SETASIDE_CODES:
        return ("VOSB", True, False)          # SDVOSB qualifies for VOSB set-asides
    if code in SMALL_BUSINESS_SETASIDE_CODES:
        return ("Total SB", True, False)
    if code in HUBZONE_CODES:
        return ("HUBZone", IS_HUBZONE_CERTIFIED, False)
    if code in EIGHTA_CODES:
        return ("8(a)", IS_8A_CERTIFIED, False)
    if code in WOSB_CODES:
        return ("WOSB", IS_WOSB_CERTIFIED, False)
    if not code:
        return ("None", True, False)  # Unrestricted -> anyone (incl. SB) may bid
    # Unknown / other restricted set-aside.
    return (desc or code, False, False)


def _keyword_hits(haystack, keywords):
    """Return the list of keywords (in order) that appear in haystack."""
    matched = []
    for kw in keywords:
        if kw in haystack and kw not in matched:
            matched.append(kw)
    return matched


def _haystack(opp):
    """Lowercased title + description + codes for keyword matching."""
    parts = [
        opp.get("title", ""),
        opp.get("description", ""),
        opp.get("naicsCode", ""),
        opp.get("classificationCode", ""),
    ]
    return " ".join(str(p) for p in parts).lower()


def _classification_needs_verb_check(opp):
    """True when the PSC/NAICS class is one where a 'medical'-type keyword must
    be corroborated by an analytic verb rather than the topic alone: PSC Q
    (medical *services* — often staffing/O&M), PSC 65xx (medical/dental supplies
    & equipment), or NAICS 339x (medical device / supplies manufacturing)."""
    psc = (opp.get("classificationCode") or "").strip().upper()
    naics = (opp.get("naicsCode") or "").strip()
    return psc[:1] == "Q" or psc[:2] == "65" or naics[:3] == "339"


def _is_commodity_supply(opp):
    """True when the notice is the supply/distribution of goods (a product buy),
    not a professional service. Requires a product/supply classification so a
    clinical notice that merely mentions 'supplies' is not killed: manufacturing
    (NAICS 31-33) or wholesale (NAICS 42) codes, a 65xx PSC, or a numeric
    (Federal Supply Classification = product) PSC — AND a supply verb with no
    dominant analytic service verb."""
    text = _haystack(opp)
    naics = (opp.get("naicsCode") or "").strip()
    psc = (opp.get("classificationCode") or "").strip().upper()
    product_code = (naics[:2] in ("31", "32", "33", "42")
                    or psc[:2] == "65"
                    or (psc[:1].isdigit()))
    supply_verb = any(v in text for v in SUPPLY_VERB_KEYWORDS)
    service_verb = any(v in text for v in SERVICE_VERB_KEYWORDS)
    return product_code and supply_verb and not service_verb


def match_capabilities(opp):
    """Return (is_match, matched_keywords) for technical-competency scoring.

    Classification-code awareness: for medical-services / medical-supply codes
    (PSC Q, PSC 65xx, NAICS 339x) a 'medical'-family keyword does NOT auto-
    qualify the notice as in-lane — the verb decides. A supply/deliver verb with
    no analytic verb means the work is product supply, not PRG's service lane.
    """
    matched = _keyword_hits(_haystack(opp), CAPABILITY_KEYWORDS)
    if matched and _classification_needs_verb_check(opp):
        text = _haystack(opp)
        if (any(v in text for v in SUPPLY_VERB_KEYWORDS)
                and not any(v in text for v in SERVICE_VERB_KEYWORDS)):
            return (False, [])
    return (len(matched) > 0, matched)


def match_low_barrier(opp):
    """Return (is_match, matched_keywords) for warm-body / staffing signals."""
    matched = _keyword_hits(_haystack(opp), LOW_BARRIER_KEYWORDS)
    return (len(matched) > 0, matched)


def match_vibecode(opp):
    """Return (is_match, matched_keywords) for light web/software builds a solo
    founder + AI can ship (websites, landing pages, small web apps, forms,
    dashboards, CMS, 508 remediation)."""
    matched = _keyword_hits(_haystack(opp), VIBECODE_KEYWORDS)
    return (len(matched) > 0, matched)


def classify_subcontracting(opp, setaside_label, tier, tech_match):
    """Decide whether an opportunity is a realistic subcontracting/teaming target.

    Returns (is_subcontracting, notice_type, angle_reason).

    The realistic sub/teaming paths for a small SDVOSB are:
      * Award Notices  -> a prime has been selected; approach them to sub.
      * Large UNRESTRICTED solicitations -> the prime must meet FAR 52.219-9
        small-business subcontracting goals, so an SDVOSB is an attractive sub.
      * Any notice whose text explicitly mentions subcontracting/teaming.
    We only flag notices with some NAICS relevance so the list stays focused.
    """
    notice_type = (opp.get("type") or opp.get("baseType") or "").strip()
    nt_lower = notice_type.lower()
    text = _haystack(opp)

    is_award = any(m in nt_lower for m in AWARD_NOTICE_MARKERS)
    sub_kw = _keyword_hits(text, SUBCONTRACT_KEYWORDS)
    mentions_sub = len(sub_kw) > 0
    unrestricted_prime = setaside_label == "None"
    relevant = tier in ("primary", "low_barrier") or tech_match

    is_sub = relevant and (is_award or mentions_sub or unrestricted_prime)

    if not is_sub:
        return (False, notice_type, "")

    angle_bits = []
    if is_award:
        awardee = _extract_awardee(opp)
        angle_bits.append(
            f"awarded prime{f' ({awardee})' if awardee else ''} — pursue as sub"
        )
    if mentions_sub:
        angle_bits.append(f"notice cites subcontracting/teaming ({sub_kw[0]})")
    if unrestricted_prime and not is_award:
        angle_bits.append(
            "unrestricted prime must meet SB subcontracting goals (FAR 52.219-9)"
        )
    return (True, notice_type or "N/A", "; ".join(angle_bits) + ".")


def evaluate(opp):
    """Evaluate a single opportunity and return a structured result dict."""
    setaside_label, setaside_eligible, is_sdvosb = classify_setaside(opp)
    tech_match, matched_kw = match_capabilities(opp)
    lb_match, lb_kw = match_low_barrier(opp)
    vibe_match, vibe_kw = match_vibecode(opp)

    naics = (opp.get("naicsCode") or "").strip()
    tier = naics_tier(naics)
    naics_targeted = tier == "primary"

    is_sub, notice_type, sub_angle = classify_subcontracting(
        opp, setaside_label, tier, tech_match
    )
    # Award notices are already-won contracts — biddable only as a subcontractor,
    # never as prime. Detect them so they stay OUT of the pursue-as-prime lists.
    award = opp.get("award") or {}
    is_awarded = ("award" in (notice_type or "").lower()) or bool(
        (isinstance(award, dict) and (award.get("awardee") or award.get("amount")))
    )
    deadline_days = _deadline_days(opp)   # computed once, reused below
    is_expired = deadline_days is not None and deadline_days < 0

    # Prime eligibility for the CORE table = allowed to bid AND technically relevant.
    eligible = setaside_eligible and tech_match

    # Warm-body eligibility = allowed to bid; deep tech match NOT required.
    low_barrier_eligible = setaside_eligible

    reason = _build_reason(
        eligible, setaside_label, setaside_eligible, is_sdvosb,
        tech_match, matched_kw, naics, naics_targeted,
    )
    lb_reason = _build_low_barrier_reason(
        setaside_eligible, setaside_label, is_sdvosb, tier,
        lb_match, lb_kw, tech_match, matched_kw, naics,
    )

    # Score used for ranking recommendations (higher = better fit).
    score = 0
    if is_sdvosb:
        score += 5
    elif setaside_eligible and setaside_label != "None":
        score += 3
    elif setaside_eligible:
        score += 1
    if naics_targeted:
        score += 3
    score += min(len(matched_kw), 6)  # cap keyword contribution

    # Separate score for ranking low-barrier plays (staffing signal + eligibility).
    lb_score = 0
    if is_sdvosb:
        lb_score += 4
    elif setaside_eligible and setaside_label != "None":
        lb_score += 3
    elif setaside_eligible:
        lb_score += 1
    lb_score += min(len(lb_kw), 5)
    if tech_match:
        lb_score += 1  # bonus if it also brushes core capability

    value_display, value_num = _extract_value(opp)

    # Richer decision fields.
    fte = _estimate_fte(_haystack(opp))
    personnel = str(fte) if fte else "Not stated"
    (disposition, kill_gate, kill_reason, soft_flags, needs_scope_check,
     score_up_bonus, pp_wall, credential_needed, template_action) = screen_gates(
        opp, fte, deadline_days, naics, setaside_label, notice_type)
    disqualified = disposition == "FAIL"
    is_future = disposition == "FUTURE"
    is_watch_template = disposition == "WATCH-TEMPLATE"
    # Capital/workforce required upfront (Improvement 2): flag Type B / capital /
    # bonding kills so the report can call out "no upfront capital" explicitly.
    capital_required = disqualified and (
        "Type B" in kill_gate or "capital/workforce" in kill_gate
        or "bonding" in kill_gate)
    location_display, country_code, is_international = _extract_location(opp)
    location_type = "International" if is_international else "CONUS / US"
    staffing = _staffing_type(opp)
    incumbent = ("inherit" in staffing.lower()) or ("take over" in staffing.lower())
    timeframe = _timeframe_note(opp)
    fit = _fit_label(tier, tech_match, matched_kw)
    if setaside_eligible:
        llc_note = "Yes — LLC may bid" + (" (SDVOSB cert req'd)" if is_sdvosb else "")
    else:
        llc_note = "No — set-aside excludes PRG"
    # PRG rubric v2 — three-gate model + weighted scoring (replaces the old
    # set-aside-weighted composite that rated undeliverable contracts Green).
    rub = prg_rubric(opp, setaside_label, setaside_eligible, is_sdvosb,
                     value_num, fte, incumbent, deadline_days,
                     disqualified, kill_gate, kill_reason)
    win_score = rub["win_score"]
    win_band, win_emoji, win_note = (rub["win_band"], rub["win_emoji"],
                                     rub["win_note"])
    # "Solo" (rubric v2) = the founder can personally deliver the entire
    # scope — NOT "single-person contract".
    is_solo = rub["founder_can_deliver"]
    solo_score = win_score
    solo_reason = rub["solo_reason"]
    # Apply first-contract-winnable score-up signals (LPTA, remote, neutral PP).
    if not disqualified and score_up_bonus:
        win_score = min(100, win_score + score_up_bonus)
        if win_score >= GREEN_MIN and win_band != "Green":
            win_band, win_emoji, win_note = "Green", "🟢", "Best bet — pursue"

    # Categorical score-down: electronic/commercial equipment repair & maintenance
    # (NAICS 8112xx / 8113xx). PRG owns no bench, so these are structurally weak
    # even when the notice text alone doesn't trip a hard OEM gate.
    if str(naics).startswith(EQUIP_REPAIR_NAICS_PREFIXES):
        win_score = max(0, win_score - 18)
        soft_flags = list(soft_flags) + [
            "equipment-repair NAICS (8112/8113) — PRG owns no bench"]
        if win_score < GREEN_MIN and win_band == "Green":
            win_band, win_emoji, win_note = "Yellow", "🟡", "On the fence — verify scope"
        if win_score < YELLOW_MIN and win_band == "Yellow":
            win_band, win_emoji, win_note = "Red", "🔴", "Weak fit — likely NO-BID"

    # Buyer type: international / non-US-government buyers recognize no US small-
    # business or SDVOSB preference, so PRG's structural edge disappears. Score
    # DOWN and flag — never a hard kill (a specific one may fit PRG's
    # international ambitions later, but none are first-contract material).
    buyer_type, buyer_is_international, buyer_hit = _buyer_type(opp)
    if buyer_is_international and not disqualified:
        # Founder-consultancy lane: when the scope sits inside a founder
        # credential domain (gate3 >= 20), an international/overseas-post
        # buyer is PRG's separate consultancy lane — the set-aside edge never
        # existed there, so it costs a nudge, not a cliff.
        if rub["gate3_score"] >= 20:
            win_score = max(0, win_score - 8)
            soft_flags = list(soft_flags) + [
                f"international buyer ({buyer_hit}) — founder-consultancy "
                "lane; no set-aside applies"]
        else:
            win_score = max(0, win_score - 20)
            soft_flags = list(soft_flags) + [
                f"international/non-US buyer ({buyer_hit}) — no SDVOSB preference"]
        if win_score < GREEN_MIN and win_band == "Green":
            win_band, win_emoji, win_note = "Yellow", "🟡", "On the fence — no US set-aside edge"
        if win_score < YELLOW_MIN and win_band == "Yellow":
            win_band, win_emoji, win_note = "Red", "🔴", "Weak fit — non-US buyer"

    # LEAD-TIME GATE (Gate 0): a scope whose fulfillment needs hiring or
    # teaming cannot be responsibly bid inside its minimum lead time — a
    # 3-day fuse on a crew contract is not an opportunity, it's a trap.
    min_lead = rub["min_lead_days"]
    if credential_needed:
        # Hire-to-win credential = contingent partner to recruit → teaming lead.
        min_lead = max(min_lead, 10)
    if (not disqualified and not is_expired and not is_awarded
            and not rub["founder_can_deliver"]
            and deadline_days is not None and 0 <= deadline_days < min_lead):
        disqualified = True
        is_watch_template = False
        kill_gate = "Gate 0 (lead time)"
        kill_reason = (f"only {deadline_days}d to respond but fulfillment "
                       f"needs hiring/teaming — minimum {min_lead}d lead")
        win_score, win_band, win_emoji = 0, "Red", "🔴"
        win_note = "Gate 0 — " + kill_reason
        is_solo = False
        rub["priority"] = 0

    # --- STAGE 0: timing + notice-type gating (takes precedence over scoring) ---
    notice_class = _classify_notice(notice_type)
    short_fuse = (deadline_days is not None
                  and 0 <= deadline_days <= SHORT_FUSE_CALENDAR_DAYS)
    text_l = _haystack(opp)
    is_sbir = ("sbir" in text_l or "sttr" in text_l
               or "small business innovation research" in text_l)
    cycle_program = (f"{_extract_agency(opp)} SBIR/STTR" if is_sbir else "")
    # RESPOND-RECOMMENDED only when the RFI/sources-sought subject sits inside
    # PRG's self-perform lane: the whole value of responding is shaping a
    # requirement you'll actually bid and getting on the CO's radar as a credible
    # source. Responding outside the lane (e.g. warm-body staffing PRG can't self-
    # deliver, or a non-US buyer that ignores set-asides) costs hours and signals
    # nothing — worse than silence.
    respond_recommended = (
        notice_class in ("SOURCES_SOUGHT", "SPECIAL", "RFI")
        and (tier == "primary" or is_solo)
        and not disqualified
        and not buyer_is_international)

    # Verdict precedence: timing → notice-type → hire-to-win → kills → urgency → fit.
    is_watch = notice_class in WATCH_NOTICE_CLASSES
    if is_expired:
        verdict = "EXPIRED"
    elif is_awarded or notice_class == "AWARD":
        verdict = "MARKET-INTEL"
    elif is_watch:
        verdict = "WATCH"
    elif is_watch_template:
        verdict = "WATCH-TEMPLATE"    # Type A individual credential — hire-to-win
    elif disqualified:
        verdict = "NO-BID"
    elif short_fuse:
        verdict = "SHORT-FUSE"
    elif pp_wall and incumbent:
        verdict = "RESEARCH"          # PP wall + incumbent = long odds first time
    elif setaside_label == "None" and win_score < YELLOW_MIN:
        verdict = "RESEARCH"          # open competition, weak edge
    else:
        verdict = "BID"
    # Route WATCH items to the watchlist; keep awarded → subcontracting/market-intel.
    is_future = is_watch and not is_expired and not is_awarded
    # WATCH-TEMPLATE items go to their own hire-to-win bucket (not a bid, not a kill).
    watch_template = is_watch_template and not is_expired and not is_awarded
    # Vibe-code lane: pursuable light web/software builds PRG can ship solo + AI.
    is_vibecode = (vibe_match and setaside_eligible and not disqualified
                   and not is_awarded and not is_expired and not is_future
                   and not watch_template)
    # Actionable per-row fields (min qualifications, gaps, services, paid route).
    readiness = _bid_readiness(
        tier, matched_kw, vibe_kw, lb_kw, setaside_label, is_sdvosb, pp_wall,
        credential_needed, verdict, notice_class, is_sub, is_vibecode)
    poc_list = _extract_poc(opp)
    poc_display = ""
    if poc_list:
        c = poc_list[0]
        poc_display = " / ".join(b for b in (c.get("name"), c.get("email"),
                                             c.get("phone")) if b)
    # Capture v3 pursuit role (the six-way decision). A hire-to-win
    # credential (WATCH-TEMPLATE) is a concrete teaming play — it outranks
    # the generic "unrestricted prime → subcontracting angle" routing.
    if disqualified or is_expired:
        role = "PASS"
    elif is_watch_template:
        role = "PRIME WITH TEAMING PARTNER"
    elif is_awarded or is_sub or notice_class == "AWARD":
        role = "PURSUE AS SUBCONTRACTOR"
    elif is_watch:
        role = ("RESPOND TO SOURCES SOUGHT" if respond_recommended
                else "MONITOR / PREPOSITION")
    elif verdict == "RESEARCH":
        role = "MONITOR / PREPOSITION"
    else:
        role = rub["role_hint"]

    loe = _estimate_loe(opp, value_num, notice_type)
    deadline_headline = ""
    if verdict == "SHORT-FUSE":
        deadline_headline = (f"⏰ {deadline_days}d left — "
                             f"due {str(opp.get('responseDeadLine') or '').split('T')[0]}")

    return {
        "solicitation": (
            opp.get("solicitationNumber")
            or opp.get("noticeId")
            or "N/A"
        ),
        "title": _clean(opp.get("title", "Untitled")),
        "agency": _extract_agency(opp),
        "agency_bucket": _agency_bucket(opp),
        "naics": naics or "N/A",
        "naics_tier": tier,
        "naics_in_profile": naics in ALL_QUERY_NAICS,
        "setaside": setaside_label,
        "value_display": value_display,
        "value_num": value_num,
        "eligible": eligible,
        "low_barrier_eligible": low_barrier_eligible,
        "reason": reason,
        "lb_reason": lb_reason,
        "matched_keywords": matched_kw,
        "lb_keywords": lb_kw,
        "is_vibecode": is_vibecode,
        "vibe_keywords": vibe_kw,
        "vibe_kw_display": ", ".join(vibe_kw[:4]) if vibe_kw else "",
        "notice_id": opp.get("noticeId") or "",
        "poc_display": poc_display,
        "min_qualifications": readiness["min_qualifications"],
        "gaps_to_close": readiness["gaps_to_close"],
        "services_to_pitch": readiness["services_to_pitch"],
        "fastest_paid_route": readiness["fastest_paid_route"],
        "founder_pp_usable": readiness["founder_pp_usable"],
        "is_sdvosb": is_sdvosb,
        "naics_targeted": naics_targeted,
        "tech_match": tech_match,
        "lb_match": lb_match,
        "is_subcontracting": is_sub,
        "is_awarded": is_awarded,
        "is_expired": is_expired,
        "disposition": disposition,
        "verdict": verdict,
        "notice_class": notice_class,
        "short_fuse": short_fuse,
        "deadline_headline": deadline_headline,
        "is_sbir": is_sbir,
        "cycle_program": cycle_program,
        "respond_recommended": respond_recommended,
        "loe": loe,
        "pp_wall": pp_wall,
        "content_hash": hashlib.sha1(
            (str(opp.get("solicitationNumber")) + str(opp.get("title"))
             + str(opp.get("description"))).encode("utf-8", "ignore")
        ).hexdigest()[:16],
        "previously_vetted": "",
        "disqualified": disqualified,
        "is_future": is_future,
        "watch_template": watch_template,
        "credential_needed": credential_needed,
        "template_action": template_action,
        "capital_required": capital_required,
        "kill_gate": kill_gate,
        "kill_reason": kill_reason,
        "future_reason": kill_reason if is_future else "",
        "soft_flags": soft_flags,
        "soft_flags_display": "; ".join(soft_flags) if soft_flags else "",
        "needs_scope_check": needs_scope_check,
        "notice_type": notice_type or "N/A",
        "sub_angle": sub_angle,
        "is_solo": is_solo,
        "solo_score": solo_score,
        "solo_reason": solo_reason,
        "personnel": personnel,
        "fte_num": fte,
        "location": location_display,
        "location_type": location_type,
        "is_international": is_international,
        "buyer_type": buyer_type,
        "buyer_is_international": buyer_is_international,
        "staffing_type": staffing,
        "timeframe": timeframe,
        "fit": fit,
        "llc_eligible": llc_note,
        "win_score": win_score,
        "win_band": win_band,
        "win_emoji": win_emoji,
        "win_note": win_note,
        "labor_plan": rub["labor_plan"],
        "pp_value": rub["pp_value"],
        "passthrough_risk": rub["passthrough_risk"],
        "bonding_required": rub["bonding_required"],
        "gate1_notes": rub["gate1_notes"],
        "gate2_score": rub["gate2_score"],
        "gate3_score": rub["gate3_score"],
        "modifier_notes": "; ".join(rub["modifier_notes"]),
        "role": role,
        "priority": rub["priority"],
        "fulfillment_model": rub["fulfillment_model"],
        "margin_est": rub["margin_est"],
        "min_lead_days": rub["min_lead_days"],
        "lead_display": rub["lead_display"],
        "poc": _extract_poc(opp),
        "psc": (opp.get("classificationCode") or "").strip(),
        "posted": (opp.get("postedDate") or "").split("T")[0],
        "deadline_days": deadline_days,
        "naics_desc": TARGET_NAICS.get(naics) or LOW_BARRIER_NAICS.get(naics) or "",
        "score": score,
        "lb_score": lb_score,
        "link": opp.get("uiLink", ""),
        "response_deadline": opp.get("responseDeadLine", "N/A"),
        "raw_setaside_eligible": setaside_eligible,
    }


def _build_reason(eligible, setaside_label, setaside_eligible, is_sdvosb,
                  tech_match, matched_kw, naics, naics_targeted):
    """Compose a short reason / gap-analysis sentence for the core table."""
    if eligible:
        bits = []
        if is_sdvosb:
            bits.append("SDVOSB set-aside aligns with veteran status")
        elif setaside_label == "None":
            bits.append("unrestricted (SB may compete)")
        else:
            bits.append(f"{setaside_label} set-aside is open to PRG")
        if naics_targeted:
            bits.append(f"NAICS {naics} is a primary target")
        if matched_kw:
            top = ", ".join(matched_kw[:3])
            bits.append(f"matches capabilities ({top})")
        return "Core match: " + "; ".join(bits) + "."

    # Not eligible -> explain the gap.
    gaps = []
    if not setaside_eligible:
        gaps.append(
            f"set-aside '{setaside_label}' excludes PRG (not SDVOSB/SB-eligible)"
        )
    if not tech_match:
        gaps.append("description does not overlap core clinical/biomedical scope")
    if not gaps:
        gaps.append("outside core capabilities")
    return "Gap: " + "; ".join(gaps) + "."


def _build_low_barrier_reason(setaside_eligible, setaside_label, is_sdvosb, tier,
                              lb_match, lb_kw, tech_match, matched_kw, naics):
    """Compose the reason string for the low-barrier / warm-body table."""
    if not setaside_eligible:
        return (
            f"Gap: set-aside '{setaside_label}' excludes PRG — not open to bid."
        )
    bits = []
    if is_sdvosb:
        bits.append("SDVOSB set-aside")
    elif setaside_label == "None":
        bits.append("unrestricted — SB may compete")
    else:
        bits.append(f"{setaside_label} set-aside open to PRG")
    if tier == "low_barrier":
        bits.append(f"low-barrier NAICS {naics}")
    if lb_kw:
        bits.append(f"staffing/support signals ({', '.join(lb_kw[:3])})")
    elif not lb_match:
        bits.append("standard service scope — win with qualified personnel")
    if tech_match:
        bits.append(f"bonus capability overlap ({matched_kw[0]})")
    return "Warm-body play: " + "; ".join(bits) + "."


# PRG's standing minimum to credibly submit any bid (assets it can have in hand
# before award). Opportunity-specific items get appended per row.
PRG_MIN_QUALIFICATIONS = (
    "Founder résumé + capability statement; active SAM.gov registration + SDVOSB "
    "self-cert; E&O/GL insurance bound before award")


def _bid_readiness(tier, matched_kw, vibe_kw, lb_kw, setaside_label, is_sdvosb,
                   pp_wall, credential_needed, verdict, notice_class, is_sub,
                   is_vibecode):
    """Derive the actionable per-row fields (min qualifications, gaps to close,
    services to pitch, fastest paid route, founder past-performance usability) —
    the columns that turn a match into a next step."""
    # Services to pitch — concrete offerings, chosen by what the notice matched.
    services = []
    if is_vibecode or vibe_kw:
        services.append("website / landing page / web app, Section 508 remediation, "
                        "dashboards (Screendot + Digipop)")
    if any(k in matched_kw for k in (
            "clinical research", "clinical trial", "clinical trials", "protocol",
            "regulatory compliance", "gcp", "monitoring", "patient recruitment")):
        services.append("clinical research coordination, monitoring & regulatory/"
                        "protocol support")
    if any(k in matched_kw for k in (
            "data management", "data analytics", "redcap", "edc", "ctms",
            "database", "medidata")):
        services.append("data management & analysis (REDCap/EDC), reporting dashboards")
    if not services:
        if tier == "low_barrier" or lb_kw:
            services.append("qualified-personnel staffing + administrative / program "
                            "support")
        else:
            services.append("scientific & technical writing, program/project support, "
                            "research & analysis")
    services_to_pitch = "; ".join(services[:3])

    # Gaps PRG must close to bid credibly.
    gaps = ["E&O/GL insurance (bind before award)"]
    if pp_wall:
        gaps.append("corporate CPARS — cite founder's individual past performance or team")
    if credential_needed:
        gaps.append(f"{credential_needed} (contingent, on-award hire)")
    if setaside_label == "None":
        gaps.append("no set-aside edge — differentiate on price/approach")
    gaps_to_close = "; ".join(gaps)

    # Fastest route to getting paid.
    if is_sub or verdict == "MARKET-INTEL":
        fastest = "Team as subcontractor under the prime"
    elif verdict == "WATCH-TEMPLATE":
        fastest = "Line up the credentialed partner, then bid"
    elif verdict in ("WATCH", "FUTURE") or notice_class in (
            "SOURCES_SOUGHT", "RFI", "PRESOL", "SPECIAL", "FORECAST"):
        fastest = "Respond to shape the requirement; bid at RFP"
    elif verdict == "RESEARCH":
        fastest = "Research incumbent/scope, then bid"
    elif verdict in ("BID", "SHORT-FUSE"):
        fastest = "Bid / quote now"
    else:
        fastest = "Assess, then bid"

    founder_pp = ("Individual experience only — no corporate CPARS yet; team or cite "
                  "founder PP" if pp_wall else "Yes — as founder's individual experience")

    min_quals = PRG_MIN_QUALIFICATIONS
    if credential_needed:
        min_quals += f"; plus a {credential_needed}"

    return {
        "min_qualifications": min_quals,
        "gaps_to_close": gaps_to_close,
        "services_to_pitch": services_to_pitch,
        "fastest_paid_route": fastest,
        "founder_pp_usable": founder_pp,
    }


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _clean(text, max_len=70):
    """Collapse whitespace and truncate long strings for table cells."""
    text = " ".join(str(text).split())
    text = text.replace("|", "\\|")  # escape pipes so Markdown tables don't break
    if len(text) > max_len:
        text = text[: max_len - 1].rstrip() + "…"
    return text or "N/A"


def _extract_agency(opp):
    """Pull the best available agency / department name."""
    for key in ("fullParentPathName", "organizationName", "departmentName"):
        val = opp.get(key)
        if val:
            # fullParentPathName is a dotted hierarchy; take the top department.
            return _clean(str(val).split(".")[0], max_len=40)
    return "N/A"


def _buyer_type(opp):
    """Classify the BUYER (not the place of performance). Returns
    (label, is_international, matched_keyword).

    A non-US / international-organization buyer (NATO/NCIA, UN agencies, foreign
    ministries) does not recognize US small-business / SDVOSB preferences, so
    PRG's structural edge disappears. Detected from the agency/department name
    first (strong signal); a short list of unambiguous org acronyms is also
    checked against the notice text.
    """
    agency = (_extract_agency(opp) or "").lower()
    office = " ".join(str(opp.get(k, "")) for k in
                      ("organizationName", "fullParentPathName",
                       "officeAddress")).lower()
    text = _haystack(opp)
    hit = next((k for k in INTERNATIONAL_BUYER_KEYWORDS
                if k in agency or k in office), None)
    if not hit:
        # Text-level check limited to unambiguous international-org signals.
        strong = ("nato", "ncia", "nspa", "united nations", "unicef", "undp",
                  "unhcr", "unops", "world health organization",
                  "european commission", "african union")
        hit = next((k for k in strong if k in text), None)
    if hit:
        return ("International / non-US buyer", True, hit)
    # US diplomatic post buying for overseas delivery — no US SB set-aside abroad.
    _, _, pop_international = _extract_location(opp)
    post = next((k for k in OVERSEAS_POST_KEYWORDS
                 if k in office or k in text), None)
    if post and pop_international:
        return ("US overseas post (no SB set-aside abroad)", True, post)
    return ("US Federal", False, "")


def _extract_awardee(opp):
    """Return the awardee company name for an award notice, if present."""
    award = opp.get("award") or {}
    awardee = award.get("awardee") or {}
    name = awardee.get("name") if isinstance(awardee, dict) else None
    return _clean(name, max_len=30) if name else ""


def _extract_poc(opp):
    """Return government points of contact as a list of {name,email,phone,type}."""
    out = []
    for p in (opp.get("pointOfContact") or []):
        if not isinstance(p, dict):
            continue
        out.append({
            "name": (p.get("fullName") or p.get("name") or "").strip(),
            "email": (p.get("email") or "").strip(),
            "phone": (p.get("phone") or "").strip(),
            "type": (p.get("type") or "").strip(),
        })
    return out


# Dollar-amount parsing for the estimated-value column. Matches figures like
# "$1,200,000", "$1.2M", "$500K", "$3 million".
_DOLLAR_RE = re.compile(
    r"\$\s?([\d,]+(?:\.\d+)?)\s?(billion|million|thousand|bn|mil|b|m|k)?\b",
    re.IGNORECASE,
)
_DOLLAR_MULTIPLIER = {
    "billion": 1e9, "bn": 1e9, "b": 1e9,
    "million": 1e6, "mil": 1e6, "m": 1e6,
    "thousand": 1e3, "k": 1e3,
}


def _format_currency(amount):
    """Human-friendly currency formatting: $1.2M, $500K, $12,000."""
    try:
        amount = float(amount)
    except (TypeError, ValueError):
        return "Not stated"
    if amount >= 1e9:
        return f"${amount / 1e9:.1f}B"
    if amount >= 1e6:
        return f"${amount / 1e6:.1f}M"
    if amount >= 1e3:
        return f"${amount / 1e3:.0f}K"
    return f"${amount:,.0f}"


def _parse_dollar_amounts(text):
    """Return a list of dollar figures (floats) found in free text."""
    amounts = []
    for num, unit in _DOLLAR_RE.findall(text or ""):
        try:
            value = float(num.replace(",", ""))
        except ValueError:
            continue
        value *= _DOLLAR_MULTIPLIER.get((unit or "").lower(), 1)
        if value >= 1000:  # ignore trivial figures / page refs
            amounts.append(value)
    return amounts


def _extract_value(opp):
    """Estimate the contract's dollar value (i.e. the revenue it would generate).

    Priority:
      1. Award amount, when the notice is an award (the real contract value).
      2. The largest dollar figure mentioned in the notice title/description.
      3. "Not stated" — SAM rarely publishes a value for open solicitations.

    Returns (display_string, numeric_value_or_None).
    """
    award = opp.get("award") or {}
    raw = award.get("amount") if isinstance(award, dict) else None
    if raw not in (None, "", 0, "0"):
        try:
            amt = float(str(raw).replace("$", "").replace(",", "").strip())
            if amt > 0:
                return (_format_currency(amt), amt)
        except ValueError:
            pass

    candidates = _parse_dollar_amounts(_haystack(opp))
    if candidates:
        amt = max(candidates)
        return (f"~{_format_currency(amt)} (from notice text)", amt)

    return ("Not stated", None)


def screen_gates(opp, fte, deadline_days, naics, setaside_label, notice_type):
    """Keyword-level pass of PRG's deep-screen gates (a first-pass triage — the
    definitive screen still requires reading the PWS/attachments).

    Returns (disposition, gate, reason, soft_flags, needs_scope_check,
    score_up_bonus, pp_wall, credential_needed, template_action) where
    disposition is one of PASS / FAIL / FUTURE / WATCH-TEMPLATE. Gates apply in
    order; the first failure stops and is recorded.
    """
    text = _haystack(opp)
    nt = (notice_type or "").lower()
    # Conditional-gate flags accumulated on the way to a PASS (capture v3).
    pre_flags = []

    def _fail(gate, reason):
        return ("FAIL", gate, reason, [], False, 0, False, "", "")

    def _template(cred):
        # Type A individual credential — winnable IF a contingent credentialed
        # teaming partner is secured. Route to the WATCH-TEMPLATE bucket, not kill.
        action = ("ACTION: secure a contingent (on-award) commitment letter from "
                  f"a {cred}")
        return ("WATCH-TEMPLATE", "Gate 2 (Type A individual credential)",
                f"hireable — needs contingent teaming partner ({cred})",
                [], False, 0, False, cred, action)

    # GATE 0.5 — structural eligibility (must hold an asset/vehicle today).
    st = _keyword_hits(text, STRUCTURAL_FAIL_KEYWORDS)
    if st:
        return _fail("Gate 0.5 (structural)",
                     f"requires an asset/vehicle PRG lacks — '{st[0]}'")

    # GATE 0 — pure resale with NO value-added component (capture v3: value-
    # added resale — install, config, training, maintenance — is a real
    # aggregator play; zero-value-add subscription/license renewals are not).
    rs = _keyword_hits(text, RESALE_KILL_KEYWORDS)
    if (rs and not _keyword_hits(text, RESALE_SERVICES_EXEMPT)
            and not _keyword_hits(text, VALUE_ADD_KEYWORDS)):
        return _fail("Gate 0 (resale, no value-add)",
                     f"pure resale with no value-added component ('{rs[0]}')")

    # GATE 0 — firm-level professional licensure PRG does not hold (surveying,
    # PE stamp, architecture, guard-firm licenses), unless key personnel may
    # hold the license. Not delegable: licensure attaches to the PRIME.
    fl = _keyword_hits(text, FIRM_LICENSURE_KILL + GUARD_LICENSE_KILL)
    if fl and not _keyword_hits(text, KEY_PERSONNEL_LICENSE_OK):
        return _fail("Gate 0 (firm licensure)",
                     f"requires firm-level licensure PRG does not hold "
                     f"('{fl[0]}')")

    # TYPE B — corporate / OEM / proprietary-system authorization (UNHIREABLE).
    # The authorization belongs to a company PRG can't put on payroll; subbing the
    # principal purpose is an illegal pass-through (>50% subcontracted) → HARD
    # KILL. Exempt pure software/product resale (SaaS/license/reseller), where
    # "OEM authorization" is an ASSET to acquire, not bench labor to perform.
    if not any(x in text for x in OEM_EXEMPT_CONTEXT):
        tb = _keyword_hits(text, OEM_PASSTHROUGH_KILL + CORP_AUTHORIZATION_KILL)
        if tb:
            return _fail("Gate 2 (Type B corp authorization)",
                         f"unhireable — corporate/OEM authorization pass-through ('{tb[0]}')")
        if any(w in text for w in _OEM_WORK) and any(s in text for s in _OEM_SPEC):
            return _fail("Gate 2 (Type B corp authorization)",
                         "brand-specific equipment maintenance to manufacturer spec — no bench")

    # IMPROVEMENT 2 — capital / workforce required BEFORE award (PRG has none):
    # multiple trained/certified technicians, a trade crew, owned equipment/lab,
    # or a facility that must exist at proposal time → HARD KILL. The winnable
    # structure is ONE contingent specialist, not speculative pre-award hiring.
    cap = _keyword_hits(text, CAPITAL_WORKFORCE_KILL)
    if cap:
        return _fail("Gate 2 (capital/workforce upfront)",
                     f"requires pre-award workforce/equipment PRG can't pre-fund ('{cap[0]}')")

    # Stated dollar scale (size-standard mentions stripped so "size standard:
    # $34.5 million" never counts as a contract value).
    scale_vals = _parse_dollar_amounts(_SIZE_STANDARD_RE.sub("", text))
    scale = max(scale_vals) if scale_vals else None

    # CONSTRUCTION (capture v3) — no longer an automatic kill. PRG can prime
    # small construction/trade work through licensed local subs (LoS: prime
    # self-performs ≥15% of general construction — PM/QC/admin counts). The
    # Miller Act makes performance/payment bonds mandatory above $150K, and
    # PRG has no bonding capacity, so bonded or >$150K construction is still
    # a hard kill.
    psc = (opp.get("classificationCode") or "").strip().upper()
    is_construction = bool(
        (naics and naics.startswith("23"))
        or (psc and psc.startswith(KC1_PSC_KILL_PREFIXES)))
    if is_construction:
        if _keyword_hits(text, BONDING_KILL_KEYWORDS):
            return _fail("Gate 0 (bonding required)",
                         "construction with bid/performance bonding — no "
                         "bonding capacity (flag: bonding required)")
        if scale is not None and scale > MILLER_ACT_THRESHOLD:
            return _fail("Gate 0 (construction > Miller Act)",
                         f"~{_format_currency(scale)} construction — bonds "
                         "mandatory above $150K; no bonding capacity")
        pre_flags.append("construction — prime via licensed trade sub; "
                         "self-perform ≥15% (PM/QC/admin); verify bond waiver")

    # KC5 — wrong scale for a first/early contract: beyond PRG's realistic
    # management, working-capital, and limitations-on-subcontracting capacity.
    if scale is not None and scale > KC5_MAX_SOLO_VALUE:
        return _fail("Gate 0 (KC5 wrong scale)",
                     f"~{_format_currency(scale)} — beyond a first-contract "
                     "prime's management and LoS capacity")

    # TYPE A — a specific individual professional credential (a licensed/certified
    # PERSON). PRG can hire that person W-2, so this is winnable with a contingent
    # partner → WATCH-TEMPLATE (checked before scope so a hireable credential is
    # never mis-killed by a generic scope word like "calibration").
    cred_a = _keyword_hits(text, INDIVIDUAL_CREDENTIAL_TEMPLATE)
    if cred_a:
        return _template(cred_a[0])

    # FIX 1/5 — specialty field-science / permit work (cultural, environmental,
    # field-crew). Degree/experience adjacency never satisfies these (FIX 4).
    sf = _keyword_hits(text, SPECIALTY_FIELD_KILL)
    if sf:
        return _fail("Gate 2 (specialty permit/field)",
                     f"specialty field-science / permit work — '{sf[0]}'")
    if "principal investigator" in text and any(c in text for c in _PI_FIELD_CONTEXT):
        return _fail("Gate 2 (specialty PI)",
                     "requires a field-science principal investigator PRG lacks")

    # GATE 1 — workforce / CBA takeover a solo shop can't staff.
    wf = _keyword_hits(text, WORKFORCE_KILL_KEYWORDS)
    if wf:
        return _fail("Gate 1 (workforce/CBA)",
                     f"incumbent-workforce / CBA signal — '{wf[0]}'")

    # GATE 0 — bonding (bid/performance bonds). Hard kill, but flagged
    # explicitly as "bonding required" rather than silently scored.
    bond = _keyword_hits(text, BONDING_KILL_KEYWORDS)
    if bond:
        return _fail("Gate 0 (bonding required)",
                     f"bonding required — '{bond[0]}'")

    # COMMODITY SUPPLY (capture v3) — a pure product buy with NO value-added
    # component is excluded (no margin to aggregate). With a value-add
    # component (install, config, training, maintenance) it's a viable
    # supply-and-services play, flagged for the nonmanufacturer rule: on a
    # set-aside supply contract PRG must furnish the product of a small U.S.
    # manufacturer (or hold an SBA class waiver).
    if _is_commodity_supply(opp):
        if not _keyword_hits(text, VALUE_ADD_KEYWORDS):
            return _fail("Gate 4 (commodity, no value-add)",
                         "pure product supply with no value-added component")
        pre_flags.append("supply + value-add — verify nonmanufacturer rule "
                         "(small U.S. manufacturer or SBA waiver)")

    # DELEGABLE TRADE / FIELD SERVICES (capture v3) — NOT a kill: deep local
    # sub markets exist and PRG primes/manages/QCs. Flagged so the margin
    # model prices sub-management and LoS is checked (similarly situated subs
    # or W2 crews for the 50% services cap).
    dq = _keyword_hits(text, DELEGABLE_TRADE_KEYWORDS)
    if dq:
        pre_flags.append(f"trade/field scope ('{dq[0]}') — sub-fulfilled; "
                         "check LoS: similarly situated subs or W2 crew")

    # COVERAGE (capture v3) — 24/7 / embedded / guaranteed-response is
    # staffable through subs/staffing firms, but it is a performance-risk and
    # working-capital flag (payroll float, deduction exposure), not a kill.
    cov = _keyword_hits(text, COVERAGE_KILLER_KEYWORDS)
    if cov:
        pre_flags.append(f"coverage commitment ('{cov[0]}') — staffing-firm "
                         "fulfillment; deduction/payroll-float risk")

    # NOTE (rubric v2): a multi-FTE team is NOT a kill — PRG's model is win as
    # prime and hire W2/1099 labor with the founder as PM. W2 hires count as
    # self-performance under the limitations-on-subcontracting rule; the labor
    # plan (Gate 1) and labor-market score (Gate 2) price the staffing risk.

    # TYPE A (generic) — a "shall/must hold a <license/certification>" that
    # survived every kill above. An individual can hold it, so it's a hire-to-win
    # (WATCH-TEMPLATE), not a kill — unless PRG already holds it or it's a
    # product/reseller asset rather than a personnel credential.
    if (_CRED_REQUIRE_RE.search(text)
            and not any(x in text for x in CREDENTIAL_EXEMPT_CONTEXT)
            and not any(h in text for h in HELD_CREDENTIALS)):
        return _template("licensed/certified individual")

    # Soft signals — score down / flag, never fail on their own.
    soft = list(pre_flags)
    if fte == 2:
        soft.append("~2 FTE — would need a teaming partner")
    if deadline_days is not None and 0 <= deadline_days < 5:
        soft.append(f"only {deadline_days}d left — very tight")
    elif deadline_days is not None and 0 <= deadline_days < 10:
        soft.append(f"only {deadline_days}d to respond")
    if _keyword_hits(text, INCUMBENT_SOFT_KEYWORDS):
        soft.append("incumbent present")
    if setaside_label == "None":
        soft.append("open competition — no SDVOSB preference")

    # Past-performance wall — a first-timer with zero CPARS record can't clear
    # a corporate-PP evaluation, especially against an incumbent → RESEARCH.
    pp_hits = _keyword_hits(text, PAST_PERF_WALL_KEYWORDS)
    pp_wall = bool(pp_hits)
    if pp_wall:
        soft.append(f"past-performance wall ({pp_hits[0]}) — first-timer risk")

    # Score-up: first-contract-winnable signals (LPTA, remote, neutral PP).
    up = _keyword_hits(text, SCORE_UP_KEYWORDS)
    score_up_bonus = min(len(up) * 4, 12)
    if up:
        soft.append(f"first-win signals ({', '.join(up[:2])})")

    needs_scope_check = naics in CATCHALL_NAICS

    # GATE 0 — disposition: market-research notices aren't biddable yet →
    # route to the WATCHLIST (prep for the eventual RFP), don't PASS or kill.
    if any(k in nt for k in FUTURE_NOTICE_MARKERS):
        return ("FUTURE", "Gate 0 (timing)",
                "market research / pre-solicitation — prep for the eventual RFP",
                soft, needs_scope_check, score_up_bonus, pp_wall, "", "")

    return ("PASS", "", "", soft, needs_scope_check, score_up_bonus, pp_wall,
            "", "")


def _estimate_fte(text):
    """Largest full-time-equivalent headcount mentioned in the text, or None."""
    counts = []
    for m in _FTE_RE.findall(text or ""):
        try:
            counts.append(int(m))
        except ValueError:
            continue
    return max(counts) if counts else None


def prg_rubric(opp, setaside_label, setaside_eligible, is_sdvosb, value_num,
               fte, incumbent, deadline_days, disqualified, kill_gate,
               kill_reason):
    """PRG rubric v2 — three-gate model plus weighted scoring.

    Gate 0 (hard disqualifiers, applied in screen_gates or here) → score 0,
    Red, never shown in Green/Yellow lists. Gate 1 = can PRG legally prime it
    (eligibility + limitations-on-subcontracting labor plan + ostensible-
    subcontractor flag). Gate 2 = does a hirable labor market exist (0-25).
    Gate 3 = can the founder credibly SUPERVISE it (0-25 — PM literacy only,
    NOT "can the founder do the work"). Weighted modifiers after gates.
    Bands: Green = 75+, Yellow = 55-74, Red = below 55 or any Gate 0 hit.
    """
    text = _haystack(opp)
    out = {
        "win_score": 0, "win_band": "Red", "win_emoji": "🔴", "win_note": "",
        "labor_plan": "n/a — killed at Gate 0", "pp_value": "Low",
        "passthrough_risk": "low", "founder_can_deliver": False,
        "solo_reason": "", "gate1_notes": "", "gate2_score": 0,
        "gate3_score": 0, "modifier_notes": [],
        "bonding_required": "bonding" in (kill_gate or "").lower(),
        "priority": 0, "fulfillment_model": "n/a",
        "margin_est": "n/a", "role_hint": "PASS",
        "min_lead_days": 0, "lead_display": "n/a",
    }

    # --- GATE 0: hard disqualifiers (kills from screen_gates + ineligible
    # set-asides like 8(a)/HUBZone/WOSB). Score 0, Red, stop.
    if disqualified:
        out["win_note"] = f"Gate 0 — {kill_reason or 'hard disqualifier'}"
        return out
    if not setaside_eligible:
        out["labor_plan"] = "n/a — ineligible set-aside"
        out["win_note"] = "Gate 0 — set-aside PRG is ineligible for"
        return out
    if deadline_days is not None and deadline_days < 0:
        out["labor_plan"] = "n/a — expired"
        out["win_note"] = "Response deadline has passed"
        return out

    # --- GATE 2: does a hirable labor market exist? (0-25)
    thin = _keyword_hits(text, LABOR_THIN)
    commodity = _keyword_hits(text, LABOR_COMMODITY)
    specialized = _keyword_hits(text, LABOR_SPECIALIZED)
    trade_market = _keyword_hits(text, DELEGABLE_TRADE_KEYWORDS)
    if thin:
        gate2, specialty = 5, "specialist"
        g2_note = f"thin/constrained labor market ('{thin[0]}')"
    elif commodity:
        gate2 = min(25, 22 + max(0, len(commodity) - 1))
        specialty = commodity[0]
        g2_note = f"commodity professional labor ({commodity[0]}) — deep 1099 market"
    elif specialized:
        gate2 = min(19, 15 + max(0, len(specialized) - 1))
        specialty = specialized[0]
        g2_note = f"specialized but findable ({specialized[0]}) — hire with lead time"
    elif trade_market:
        # Aggregator play: local trade/field-service subs are one of the
        # DEEPEST sourcing markets there is (every metro has dozens).
        gate2, specialty = 20, f"{trade_market[0]} sub"
        g2_note = (f"deep local trade-sub market ({trade_market[0]}) — "
                   "source 3+ quotes")
    else:
        gate2, specialty = 12, "professional labor"
        g2_note = "labor market unclear — verify before bid"

    # --- GATE 3: can the founder credibly supervise it? (0-25, PM literacy)
    founder_hits = _keyword_hits(text, FOUNDER_DOMAIN_KEYWORDS)
    adjacent_hits = _keyword_hits(text, ADJACENT_DOMAIN_KEYWORDS)
    low_hits = _keyword_hits(text, LOW_LITERACY_KEYWORDS)
    if low_hits and not founder_hits:
        gate3 = 5
        g3_note = f"founder cannot judge work quality ({low_hits[0]})"
    elif founder_hits:
        gate3 = min(25, 23 + max(0, len(founder_hits) - 2))
        g3_note = f"founder credential domain ({founder_hits[0]})"
    elif adjacent_hits:
        gate3 = min(19, 15 + max(0, len(adjacent_hits) - 2))
        g3_note = f"adjacent domain — founder can spec/QA ({adjacent_hits[0]})"
    elif trade_market:
        # Routine trade scopes are inspectable by a competent PM: acceptance
        # is a QASP walk-through, not specialist judgment (a clean floor, a
        # patched roof, a mowed field). Supervision literacy is adequate.
        gate3 = 15
        g3_note = (f"routine trade scope ({trade_market[0]}) — QASP/"
                   "checklist supervisable")
    else:
        gate3 = 10
        g3_note = "PM literacy unclear — verify scope"

    # --- GATE 1: can PRG legally prime it? Labor plan under the limitations-
    # on-subcontracting rule (services: ≤50% of the government's payment may
    # go to non-similarly-situated subs; W2 hires always count as PRG).
    if is_sdvosb or setaside_label == "SDVOSB":
        labor_plan = (f"W2 {specialty} (SDVOSB LoS 50% cap — W2 hires or "
                      "SDVOSB 1099s)")
    elif setaside_label == "VOSB":
        labor_plan = (f"W2 {specialty} (VOSB LoS 50% cap — W2 hires or "
                      "VOSB 1099s)")
    elif setaside_label == "Total SB":
        labor_plan = (f"1099 {specialty} (Total SB — sole-prop 1099s are "
                      "similarly situated, no cap issue)")
    elif setaside_label == "None":
        labor_plan = f"W2 or any 1099 {specialty} (unrestricted, no LoS cap)"
    else:
        labor_plan = f"W2 {specialty} (verify LoS for {setaside_label})"
    # Ostensible-subcontractor risk: one indivisible specialty likely performed
    # end-to-end by a single sub the founder can't meaningfully manage.
    passthrough_risk = "high" if (gate3 <= 9 and not commodity) else "low"
    gate1_notes = "eligible to prime"
    if passthrough_risk == "high":
        gate1_notes += "; pass-through risk: high (indivisible specialty scope)"

    # --- Weighted modifiers (applied after gates).
    deliverable_hits = _keyword_hits(text, DELIVERABLE_SCOPE_KEYWORDS)
    loe_hits = _keyword_hits(text, LOE_SCOPE_KEYWORDS)
    crew_hits = _keyword_hits(text, SOLO_EXCLUDE_KEYWORDS)
    small_buy = ((value_num is not None
                  and value_num < SIMPLIFIED_ACQ_THRESHOLD)
                 or "simplified acquisition" in text)
    # "Solo" (rubric v2) = the FOUNDER can personally deliver the entire scope
    # (founder credential domain + deliverable-based + no crew/LOE signals +
    # small enough), NOT merely "single-person contract".
    founder_can_deliver = bool(
        founder_hits and deliverable_hits and not crew_hits and not loe_hits
        and (fte is None or fte <= 1)
        and (value_num is None or value_num <= 2 * SIMPLIFIED_ACQ_THRESHOLD))

    mods = []
    if founder_can_deliver:
        mods.append((15, "founder can personally deliver (Solo: Yes)"))
    if small_buy:
        mods.append((10, "under $250K / simplified acquisition — PP builder"))
    if deliverable_hits and not loe_hits:
        mods.append((10, "deliverable-based scope (not level-of-effort)"))
    if _keyword_hits(text, RECURRING_KEYWORDS):
        mods.append((5, "recurring / multi-year (base plus options)"))
    if is_sdvosb:
        mods.append((5, "SDVOSB set-aside — thinner competition"))
    if (deadline_days is not None and 0 <= deadline_days < 5
            and not founder_can_deliver):
        mods.append((-10, "under 5 days AND needs hiring/teaming first"))
    if incumbent:
        mods.append((-10, "identifiable incumbent recompete, no opening"))
    pop = opp.get("placeOfPerformance") or {}
    pop_state = ""
    if isinstance(pop, dict):
        s = pop.get("state") or {}
        pop_state = ((s.get("code") or s.get("name") or "")
                     if isinstance(s, dict) else str(s or ""))
    if (_keyword_hits(text, ONSITE_KEYWORDS) and pop_state
            and pop_state.strip().upper() not in _SOCAL_STATES):
        mods.append((-5, "majority on-site outside Southern California"))

    score = RUBRIC_BASE + gate2 + gate3 + sum(d for d, _ in mods)
    score = max(0, min(100, score))
    if score >= GREEN_MIN:
        band, emoji, note = "Green", "🟢", "Best bet — pursue"
    elif score >= YELLOW_MIN:
        band, emoji, note = "Yellow", "🟡", "On the fence — worth a look"
    else:
        band, emoji, note = "Red", "🔴", "Do not pursue — low odds"

    # Past-performance strategic value: small fast wins are worth more than
    # face value to a first-time prime.
    if small_buy or founder_can_deliver:
        pp_value = "High"
    elif value_num is None or value_num <= 1_000_000:
        pp_value = "Med"
    else:
        pp_value = "Low"

    if founder_can_deliver:
        solo_reason = ("Founder-deliverable: "
                       f"{founder_hits[0]} scope, deliverable-based, no "
                       "crew/LOE signals — higher margin, zero staffing risk")
    else:
        why = ("outside founder credential domains" if not founder_hits
               else "needs hired labor (crew/LOE/team signals or size)")
        solo_reason = f"Not founder-deliverable — {why}; PM + hired labor model"

    # --- CAPTURE MODEL v3: fulfillment classification, margin estimate,
    # 1-10 dimension scores, and the weighted priority (0-100).
    r_naics = (opp.get("naicsCode") or "").strip()
    r_psc = (opp.get("classificationCode") or "").strip().upper()
    is_constr = bool(r_naics.startswith("23")
                     or (r_psc and r_psc.startswith(KC1_PSC_KILL_PREFIXES)))
    trade_hits = (_keyword_hits(text, DELEGABLE_TRADE_KEYWORDS)
                  or (["construction"] if is_constr else []))
    supply_hits = (_keyword_hits(text, SUPPLY_VERB_KEYWORDS)
                   and _keyword_hits(text, VALUE_ADD_KEYWORDS))
    coverage_hits = _keyword_hits(text, COVERAGE_KILLER_KEYWORDS)
    if founder_can_deliver:
        model, margin_est, m_margin = "founder-delivered", "40-60%", 9
    elif supply_hits:
        model, margin_est, m_margin = "value-added supply", "10-20%", 3
    elif trade_hits:
        model, margin_est, m_margin = "trade-sub management", "8-15%", 4
    elif loe_hits or coverage_hits:
        model, margin_est, m_margin = "staffing (W2/1099)", "15-30%", 5
    else:
        model, margin_est, m_margin = "professional services", "30-50%", 8

    nt = (opp.get("type") or "").lower()
    simplified = small_buy or "combined synopsis" in nt or "rfq" in text
    d = {
        "p_win": max(1, min(10, round(score / 10)
                            + (1 if is_sdvosb else 0)
                            - (2 if setaside_label == "None" else 0))),
        "margin": m_margin,
        "delegation": (9 if commodity else 7 if trade_hits or supply_hits
                       else 6 if specialized else 2 if thin else 5),
        "perf_risk": max(1, 8 - (2 if trade_hits else 0)
                         - (3 if coverage_hits else 0)
                         - (2 if passthrough_risk == "high" else 0)
                         - (1 if loe_hits else 0)),
        "wc_burden": (9 if small_buy else
                      4 if (supply_hits or trade_hits) else
                      5 if (loe_hits or coverage_hits) else 7),
        "speed_cash": (8 if simplified else 5) - (1 if loe_hits else 0),
        "pp_val": {"High": 8, "Med": 5, "Low": 3}[pp_value],
        "scalability": (8 if _keyword_hits(text, RECURRING_KEYWORDS) else 4),
    }
    priority = round(sum(CAPTURE_WEIGHTS[k] * d[k] for k in CAPTURE_WEIGHTS)
                     * 10)

    # Role hint (final role also considers notice class in evaluate()):
    # teaming when execution rides on subs/staffing or pass-through risk is
    # high; prime when PRG's PM + hired labor carries it.
    needs_team = (trade_hits or supply_hits or coverage_hits
                  or passthrough_risk == "high"
                  or (fte is not None and fte >= 5))
    role_hint = ("PRIME WITH TEAMING PARTNER" if needs_team
                 else "BID AS PRIME")

    # Minimum responsible lead time before the deadline, by fulfillment
    # complexity: founder-deliverable scopes can turn fast; anything needing
    # a hire gets a week; teaming/multiple hires get ten days. Below the
    # minimum, evaluate() kills the row (Gate 0 lead time) instead of
    # letting it masquerade as an actionable SHORT-FUSE.
    if founder_can_deliver:
        min_lead_days, lead_display = 2, "2d (founder-deliverable)"
    elif needs_team:
        min_lead_days, lead_display = 10, "10d (teaming / multiple hires)"
    else:
        min_lead_days, lead_display = 7, "7d (single specialist hire)"

    out.update({
        "priority": priority, "fulfillment_model": model,
        "margin_est": margin_est, "role_hint": role_hint,
        "min_lead_days": min_lead_days, "lead_display": lead_display,
    })

    out.update({
        "win_score": score, "win_band": band, "win_emoji": emoji,
        "win_note": note + f" — G2 labor {gate2}/25 ({g2_note}); "
                           f"G3 supervision {gate3}/25 ({g3_note})",
        "labor_plan": labor_plan, "pp_value": pp_value,
        "passthrough_risk": passthrough_risk,
        "founder_can_deliver": founder_can_deliver,
        "solo_reason": solo_reason, "gate1_notes": gate1_notes,
        "gate2_score": gate2, "gate3_score": gate3,
        "modifier_notes": [f"{'+' if d > 0 else ''}{d} {label}"
                           for d, label in mods],
    })
    return out


# --- Location, staffing, timeframe, fit, and win-likelihood --------------------

_DOMESTIC_COUNTRY = {"USA", "US", "UNITED STATES", "USA ", ""}

# Signals that an incumbent already performs the work (you'd take over / may
# inherit staff) vs. a brand-new requirement (you hire from scratch).
_INCUMBENT_SIGNALS = [
    "incumbent", "currently being performed", "current contractor",
    "predecessor", "transition-in", "transition in", "phase-in", "phase in",
    "seamless transition", "right of first refusal", "currently performed",
    "existing contract", "follow-on", "recompete",
]
_NEW_REQ_SIGNALS = [
    "new requirement", "new contract", "stand up", "stand-up", "no incumbent",
    "newly established", "first time",
]


def _extract_location(opp):
    """Return (display, country_code, is_international)."""
    pop = opp.get("placeOfPerformance") or {}
    if not isinstance(pop, dict):
        return ("Not stated", "", False)

    def _name(key):
        v = pop.get(key)
        if isinstance(v, dict):
            return v.get("name") or v.get("code") or ""
        return v or ""

    city = _name("city")
    state = pop.get("state") or {}
    state_code = state.get("code") or state.get("name") if isinstance(state, dict) else state
    country = pop.get("country") or {}
    country_code = ""
    country_name = ""
    if isinstance(country, dict):
        country_code = (country.get("code") or "").upper()
        country_name = country.get("name") or ""
    else:
        country_code = str(country).upper()

    is_intl = bool(country_code) and country_code not in _DOMESTIC_COUNTRY \
        and "UNITED STATES" not in country_name.upper()

    parts = [p for p in [city, state_code] if p]
    display = ", ".join(parts)
    if is_intl:
        display = (display + " · " if display else "") + (country_name or country_code)
    if not display:
        display = country_name or country_code or "Not stated"
    return (_clean(display, 34), country_code, is_intl)


def _staffing_type(opp):
    """Guess whether you'd hire new staff or take over existing personnel."""
    text = _haystack(opp)
    if _keyword_hits(text, _INCUMBENT_SIGNALS):
        return "Take over / may inherit incumbent staff"
    if _keyword_hits(text, _NEW_REQ_SIGNALS):
        return "New requirement — hire fresh"
    return "Unclear — verify in notice"


def _timeframe_note(opp):
    """Actionable timeframe: response deadline + any period-of-performance hint."""
    deadline = opp.get("responseDeadLine") or "Not stated"
    text = _haystack(opp)
    options = len(re.findall(r"option\s+(?:year|period)", text))
    has_base = "base period" in text or "base year" in text
    pop = ""
    if has_base and options:
        pop = f"1 base + {options} option yr(s)"
    elif options:
        pop = f"{options} option period(s) referenced"
    elif has_base:
        pop = "base period (see notice)"
    resp = str(deadline).split("T")[0] if deadline else "Not stated"
    return f"Respond by {resp}" + (f"; POP: {pop}" if pop else "")


def _deadline_days(opp):
    """Days from today until the response deadline, or None if unparseable."""
    raw = opp.get("responseDeadLine") or ""
    if not raw:
        return None
    try:
        d = dt.date.fromisoformat(str(raw)[:10])
    except ValueError:
        return None
    return (d - dt.date.today()).days


def _next_cycle_window(due_iso):
    """Infer next year's release/due window for a recurring program."""
    try:
        d = dt.date.fromisoformat(str(due_iso)[:10])
        nxt_due = d.replace(year=d.year + 1)
    except Exception:
        return "next cycle ~annually (date TBD)"
    nxt_rel = nxt_due - dt.timedelta(days=60)
    return f"~release {nxt_rel.isoformat()}, due {nxt_due.isoformat()}"


def _estimate_loe(opp, value_num, notice_type):
    """Rough level-of-effort to propose, for triage/pricing."""
    text = _haystack(opp)
    nt = (notice_type or "").lower()
    if any(k in nt for k in ("sources sought", "presolicitation", "special notice")):
        return "Low — capability statement"
    heavy = any(k in text for k in (
        "technical volume", "past performance volume", "oral presentation",
        "page limit", "volume i", "factor 1", "written narrative"))
    if heavy or (value_num and value_num > 1_000_000):
        return "High — multi-volume proposal"
    if ("request for quot" in text or "rfq" in nt or "quotation" in text
            or (value_num and value_num <= SIMPLIFIED_ACQ_THRESHOLD)):
        return "Low–Medium — quote / short response"
    return "Medium — standard proposal"


def _fit_label(tier, tech_match, matched_kw):
    """Plain-English fit rating for PRG's capabilities."""
    if tier == "primary" and len(matched_kw) >= 3:
        return "Strong"
    if tech_match:
        return "Moderate"
    if tier == "low_barrier":
        return "Adjacent"
    return "Weak / off-target"


def _date_range(days_back):
    """Return (postedFrom, postedTo) as MM/DD/YYYY strings."""
    today = dt.date.today()
    start = today - dt.timedelta(days=days_back)
    fmt = "%m/%d/%Y"
    return start.strftime(fmt), today.strftime(fmt)


# ---------------------------------------------------------------------------
# 4. OUTPUT GENERATION
# ---------------------------------------------------------------------------

def _opportunity_table(rows, eligible_key, reason_key):
    """Render a standard reviewed-opportunity Markdown table from result rows.

    Rows are ordered by win score (best first) so the strongest bets sit on top.
    """
    lines = [
        "| Priority | Win | Rating | Role | Solicitation # | Title | Agency "
        "| Set-Aside | Est. Value | Fulfillment | Est. GM% | Solo? "
        "| Labor Plan | PP Value | Location | Short Reason |",
        "| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- "
        "| :--- | :--- | :--- | :--- | :--- | :--- | :--- |",
    ]
    for r in sorted(rows, key=lambda x: (x.get("priority", 0), x["win_score"]),
                    reverse=True):
        solo = "Yes" if r["is_solo"] else "No"   # Solo = founder-deliverable
        lines.append(
            f"| {r.get('priority', 0)} | {r['win_score']} "
            f"| {r['win_emoji']} {r['win_band']} | {r.get('role', '')} "
            f"| {r['solicitation']} | {r['title']} | {r['agency']} "
            f"| {r['setaside']} | {r['value_display']} "
            f"| {r.get('fulfillment_model', '')} | {r.get('margin_est', '')} "
            f"| {solo} | {r.get('labor_plan', '')} | {r.get('pp_value', '')} "
            f"| {r['location']} | {r[reason_key]} |"
        )
    return lines


def render_report(results):
    """Print the full Markdown screening report to stdout."""
    lines = []
    lines.append(f"## SAM.gov Contract Screening Report for {COMPANY_NAME}")
    lines.append("")
    lines.append(
        f"*Profile:* {SOCIOECONOMIC_STATUS} • "
        f"[{COMPANY_WEBSITE}]({COMPANY_WEBSITE}) • "
        f"Generated {dt.date.today().isoformat()}"
    )
    lines.append("")

    # --- Capture executive summary (aggregation model) ---
    def _best(rows, key=lambda r: (r.get("priority", 0), r["win_score"])):
        rows = sorted(rows, key=key, reverse=True)
        return rows[0] if rows else None

    live = [r for r in results if not r["is_expired"] and not r["is_awarded"]]
    passed = [r for r in live if not r["disqualified"]]
    biddable = [r for r in passed if not r["is_future"]
                and not r.get("watch_template")]
    b_prime = _best([r for r in biddable if r.get("role") == "BID AS PRIME"])
    b_team = _best([r for r in biddable
                    if r.get("role") == "PRIME WITH TEAMING PARTNER"])
    b_sub = _best([r for r in results
                   if r.get("role") == "PURSUE AS SUBCONTRACTOR"])
    b_pp = _best([r for r in biddable if r.get("pp_value") == "High"])
    lines.append("### A. Capture Executive Summary")
    lines.append("")
    lines.append(f"- **Reviewed:** {len(results)} | **Passed screen:** "
                 f"{len(passed)} | **Biddable now:** {len(biddable)}")
    for label, row in (("Best immediate prime", b_prime),
                       ("Best teaming play", b_team),
                       ("Best subcontract target", b_sub),
                       ("Best past-performance builder", b_pp)):
        if row:
            lines.append(
                f"- **{label}:** {row['title']} ({row['solicitation']}, "
                f"{row['agency']}) — priority {row.get('priority', 0)}, "
                f"{row.get('fulfillment_model', '')}, "
                f"GM {row.get('margin_est', '')}")
        else:
            lines.append(f"- **{label}:** none this window")
    lines.append("")

    # Award notices (already won) only belong in the subcontracting view.
    prime = [r for r in results if not r["is_awarded"] and not r["is_expired"]
             and not r["disqualified"] and not r["is_future"]
             and not r.get("watch_template")]
    primary = [r for r in prime if r["naics_tier"] == "primary"]
    low_barrier = [r for r in prime if r["naics_tier"] == "low_barrier"]
    subcontract = [r for r in results if r["is_subcontracting"]]
    solo = sorted([r for r in prime if r["is_solo"]],
                  key=lambda r: r["solo_score"], reverse=True)

    # --- Section 1: complete list (core-target NAICS) --------------------
    lines.append("### 1. Complete List of Reviewed Opportunities (Core-Target NAICS)")
    lines.append("")
    if not primary:
        lines.append(
            "_No opportunities were returned for the primary target NAICS codes "
            "in the selected date window. Widen the `--days` window or verify "
            "the API key, then re-run._"
        )
        lines.append("")
    else:
        lines.extend(_opportunity_table(primary, "eligible", "reason"))
        lines.append("")

    # --- Section 2: recommendations --------------------------------------
    lines.append("### 2. Actionable Recommendations")
    lines.append("")

    top3 = sorted(
        [r for r in primary if r["eligible"]],
        key=lambda r: r["score"],
        reverse=True,
    )[:3]

    if not top3:
        lines.append(
            "*   No eligible, on-target prime opportunities were identified in "
            "this run. Recommended actions:"
        )
        lines.append(
            "    *   Broaden the search window (`--days 60` or `--days 90`) to "
            "capture recently posted notices."
        )
        lines.append(
            "    *   Review the low-barrier and subcontracting lists below for "
            "near-term, winnable work."
        )
        lines.append(
            "    *   Confirm the SAM.gov API key is active and the entity's "
            "SDVOSB certification is current in SAM.gov and VetCert."
        )
    else:
        lines.append(
            f"Top {len(top3)} best-matching prime opportunit"
            f"{'y' if len(top3) == 1 else 'ies'} for {COMPANY_NAME}:"
        )
        lines.append("")
        for i, r in enumerate(top3, start=1):
            _render_recommendation(lines, i, r)
    lines.append("")

    # --- Section 3: low-barrier / warm-body ------------------------------
    lines.append('### 3. Low-Barrier / "Warm Body" Opportunities')
    lines.append("")
    lines.append(
        "_Adjacent staffing, administrative, and support work under secondary "
        "NAICS codes that PRG can reasonably win by fielding eligible, "
        "qualified personnel — lower barrier to entry, less domain "
        "specialization required._"
    )
    lines.append("")
    lb_eligible = sorted(
        [r for r in low_barrier if r["low_barrier_eligible"]],
        key=lambda r: r["lb_score"],
        reverse=True,
    )
    if not lb_eligible:
        lines.append(
            "_No eligible low-barrier opportunities were returned in this window. "
            "Widen `--days`, or register for Sources Sought under the secondary "
            "NAICS codes to shape upcoming staffing/support set-asides._"
        )
    else:
        lines.extend(
            _opportunity_table(lb_eligible, "low_barrier_eligible", "lb_reason")
        )
    lines.append("")

    # --- Section 4: subcontracting / teaming -----------------------------
    lines.append("### 4. Subcontracting & Teaming Opportunities")
    lines.append("")
    lines.append(
        "_Awarded primes and large unrestricted solicitations where PRG's "
        "realistic path is teaming as a subcontractor rather than bidding as "
        "prime. Unrestricted primes must meet FAR 52.219-9 small-business "
        "subcontracting goals, making an SDVOSB an attractive teammate._"
    )
    lines.append("")
    subs_sorted = sorted(subcontract, key=lambda r: r["score"], reverse=True)
    if not subs_sorted:
        lines.append(
            "_No subcontracting/teaming targets were identified in this window._"
        )
    else:
        lines.append(
            "| Solicitation # | Title | Agency | NAICS | Set-Aside | Est. Value "
            "| Notice Type | Subcontracting Angle |"
        )
        lines.append("| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |")
        for r in subs_sorted:
            lines.append(
                f"| {r['solicitation']} | {r['title']} | {r['agency']} "
                f"| {r['naics']} | {r['setaside']} | {r['value_display']} "
                f"| {r['notice_type']} | {r['sub_angle']} |"
            )
    lines.append("")

    # --- Section 5: solo-friendly (one-person) ---------------------------
    lines.append("### 5. Solo: Founder-Deliverable Opportunities")
    lines.append("")
    lines.append(
        "_Small-scale knowledge work — research, analysis, writing, data, and "
        "reviews — that one capable person could realistically deliver without "
        "a team. Ranked by fit; verify scope in the full notice before bidding._"
    )
    lines.append("")
    if not solo:
        lines.append(
            "_No clearly solo-doable opportunities surfaced in this window. "
            "Widen `--days`, or add research/analysis NAICS (e.g. 541720, "
            "541990, 611710) and re-run._"
        )
    else:
        lines.append(
            "| Win | Rating | Solicitation # | Title | Agency | Set-Aside "
            "| Est. Value | Timeframe | Location | Why One-Person Doable |"
        )
        lines.append(
            "| :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |"
        )
        for r in sorted(solo, key=lambda x: x["win_score"], reverse=True):
            lines.append(
                f"| {r['win_score']} | {r['win_emoji']} {r['win_band']} "
                f"| {r['solicitation']} | {r['title']} | {r['agency']} "
                f"| {r['setaside']} | {r['value_display']} | {r['timeframe']} "
                f"| {r['location']} | {r['solo_reason']} |"
            )
        if len(solo) < 10:
            lines.append("")
            lines.append(
                f"_Only {len(solo)} solo-friendly match"
                f"{'' if len(solo) == 1 else 'es'} this window — widen "
                f"`--days` to 120 or 180 to surface more._"
            )
    lines.append("")

    # --- Summary footer ---------------------------------------------------
    total = len(results)
    n_elig = sum(1 for r in primary if r["eligible"])
    n_sdvosb = sum(1 for r in results if r["is_sdvosb"])

    # Total stated value across everything PRG could pursue (prime-eligible core
    # + low-barrier plays), where a dollar figure was actually published.
    pursuable = {id(r): r for r in
                 [x for x in primary if x["eligible"]] + lb_eligible}.values()
    valued = [r["value_num"] for r in pursuable if r["value_num"]]
    pipeline_total = sum(valued)

    lines.append("---")
    lines.append(
        f"*Screened **{total}** opportunit"
        f"{'y' if total == 1 else 'ies'} across "
        f"{len(TARGET_NAICS)} core + {len(LOW_BARRIER_NAICS)} low-barrier NAICS "
        f"codes — **{n_elig}** core-eligible, "
        f"**{len(lb_eligible)}** low-barrier plays, "
        f"**{len(subs_sorted)}** subcontracting/teaming targets, "
        f"**{n_sdvosb}** SDVOSB set-asides.*"
    )
    if valued:
        lines.append("")
        lines.append(
            f"*Estimated pipeline value (where a dollar figure was published): "
            f"**{_format_currency(pipeline_total)}** across {len(valued)} "
            f"opportunit{'y' if len(valued) == 1 else 'ies'}. Most open "
            f"solicitations do not publish a value, so this is a floor, not a "
            f"total — treat all figures as rough estimates.*"
        )

    print("\n".join(lines))


def _render_recommendation(lines, i, r):
    """Append one formatted top-pick recommendation to `lines`."""
    naics_desc = TARGET_NAICS.get(r["naics"], "")
    kw = ", ".join(r["matched_keywords"][:5]) or "n/a"
    deadline = r.get("response_deadline", "N/A")
    lines.append(
        f"*   **#{i}. {r['title']}** (Solicitation `{r['solicitation']}`)"
    )
    lines.append(f"    *   **Agency:** {r['agency']}")
    lines.append(
        f"    *   **NAICS:** {r['naics']}"
        + (f" — {naics_desc}" if naics_desc else "")
    )
    lines.append(
        f"    *   **Set-Aside:** {r['setaside']}"
        + ("  ✅ SDVOSB direct fit" if r["is_sdvosb"] else "")
    )
    lines.append(f"    *   **Est. Contract Value (revenue):** {r['value_display']}")
    lines.append(f"    *   **Why it fits:** capability signals → {kw}.")
    lines.append(f"    *   **Response Deadline:** {deadline}")
    if r.get("link"):
        lines.append(f"    *   **Link:** {r['link']}")
    lines.append("")


# ---------------------------------------------------------------------------
# 5. SPREADSHEET EXPORT (Excel / CSV)
# ---------------------------------------------------------------------------

# Column layouts for each worksheet: (header, result-key-or-callable).
# Full decision-dashboard columns, ordered for go/no-go review.
_RICH_COLS = [
    ("Go / No-Go", "verdict"),
    ("Recommended Role", "role"),
    ("Priority (capture)", "priority"),
    ("Win Score", "win_score"),
    ("Rating", lambda r: f"{r['win_emoji']} {r['win_band']}"),
    ("Fulfillment Model", "fulfillment_model"),
    ("Est. GM%", "margin_est"),
    ("Est. LOE", "loe"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("Fit for PRG", "fit"),
    ("Set-Aside", "setaside"),
    ("Eligible as LLC?", "llc_eligible"),
    ("Est. Value (Revenue)", "value_display"),
    ("Personnel (FTE)", "personnel"),
    ("Hire New or Take Over?", "staffing_type"),
    # "Solo" (rubric v2) = founder can personally deliver the entire scope.
    ("Solo (Founder-Deliverable)?", lambda r: "Yes" if r["is_solo"] else "No"),
    ("Labor Plan", "labor_plan"),
    ("PP Value", "pp_value"),
    ("Pass-Through Risk", "passthrough_risk"),
    ("Lead Time Needed", "lead_display"),
    ("Timeframe", "timeframe"),
    ("Location", "location"),
    ("Intl / CONUS", "location_type"),
    ("Buyer Type", "buyer_type"),
    ("Fastest Paid Route", "fastest_paid_route"),
    ("Services to Pitch", "services_to_pitch"),
    ("Gaps to Close (before award)", "gaps_to_close"),
    ("Min Qualifications", "min_qualifications"),
    ("Founder PP Usable?", "founder_pp_usable"),
    ("POC (from notice)", "poc_display"),
    ("NAICS", "naics"),
    ("Notice ID", "notice_id"),
    ("Solicitation #", "solicitation"),
    ("Link", "link"),
]
_CORE_COLS = _RICH_COLS
_LOWBAR_COLS = _RICH_COLS + [("Why It's Winnable", "lb_reason")]
_SUB_COLS = [
    ("Win Score", "win_score"),
    ("Rating", lambda r: f"{r['win_emoji']} {r['win_band']}"),
    ("Solicitation #", "solicitation"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("NAICS", "naics"),
    ("Set-Aside", "setaside"),
    ("Est. Value (Revenue)", "value_display"),
    ("Notice Type", "notice_type"),
    ("Subcontracting Angle", "sub_angle"),
    ("Link", "link"),
]
_SOLO_COLS = _RICH_COLS + [("Why One-Person Doable", "solo_reason")]

# Vibe-code (light web/software) tab — the "what to build" columns up front.
_VIBECODE_COLS = [
    ("Win Score", "win_score"),
    ("Rating", lambda r: f"{r['win_emoji']} {r['win_band']}"),
    ("Go / No-Go", "verdict"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("What to Build (matched)", "vibe_kw_display"),
    ("Services to Pitch", "services_to_pitch"),
    ("Set-Aside", "setaside"),
    ("Est. Value (Revenue)", "value_display"),
    ("Est. LOE", "loe"),
    ("Respond By", "response_deadline"),
    ("Gaps to Close (before award)", "gaps_to_close"),
    ("POC (from notice)", "poc_display"),
    ("NAICS", "naics"),
    ("Solicitation #", "solicitation"),
    ("Link", "link"),
]


def _cell(row, spec):
    """Resolve a column spec (key or callable) against a result row."""
    value = spec(row) if callable(spec) else row.get(spec, "")
    # Un-escape the Markdown pipe-escaping for spreadsheet cells.
    return str(value).replace("\\|", "|")


def _split_sections(results):
    """Return the (core, low_barrier, subcontracting, solo, international) groups.

    Award notices (already-won contracts) are excluded from every
    pursue-as-prime group and appear only under Subcontracting.
    """
    prime = [r for r in results if not r["is_awarded"] and not r["is_expired"]
             and not r["disqualified"] and not r["is_future"]
             and not r.get("watch_template")]
    primary = [r for r in prime if r["naics_tier"] == "primary"]
    low_barrier = sorted(
        [r for r in prime
         if r["naics_tier"] == "low_barrier" and r["low_barrier_eligible"]],
        key=lambda r: r["lb_score"], reverse=True,
    )
    subs = sorted(
        [r for r in results if r["is_subcontracting"]],
        key=lambda r: r["score"], reverse=True,
    )
    solo = sorted(
        [r for r in prime if r["is_solo"]],
        key=lambda r: r["solo_score"], reverse=True,
    )
    international = sorted(
        [r for r in prime if r["is_international"]],
        key=lambda r: r["win_score"], reverse=True,
    )
    return primary, low_barrier, subs, solo, international


_RECOMPETE_COLS = [
    ("Ends In", lambda r: f"{r['months_left']} mo"),
    ("PoP End Date", "end_date"),
    ("Incumbent (who to beat / team with)", "recipient"),
    ("Award Amount", "amount_display"),
    ("Agency", "agency"),
    ("Priority Agency", lambda r: r.get("agency_bucket") or "—"),
    ("NAICS", "naics"),
    ("Award ID", "award_id"),
]

_TOP10_COLS = [
    ("Rank", "rank"),
    ("Go / No-Go", "verdict"),
    ("Win Score", "win_score"),
    ("Est. LOE", "loe"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("Set-Aside", "setaside"),
    ("Est. Value", "value_display"),
    ("Solo?", lambda r: "Yes" if r["is_solo"] else "No"),
    ("Flags", "soft_flags_display"),
    ("Respond By", "response_deadline"),
    ("Solicitation #", "solicitation"),
    ("Link", "link"),
]


_KILL_COLS = [
    ("Failed Gate", "kill_gate"),
    ("Why Killed", "kill_reason"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("NAICS", "naics"),
    ("Set-Aside", "setaside"),
    ("Solicitation #", "solicitation"),
    ("Link", "link"),
]

_WATCHLIST_COLS = [
    ("Notice Type", "notice_type"),
    ("Why (prep, don't bid yet)", "future_reason"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("NAICS", "naics"),
    ("Set-Aside", "setaside"),
    ("Fit for PRG", "fit"),
    ("Respond By", "response_deadline"),
    ("Solicitation #", "solicitation"),
    ("Link", "link"),
]

# WATCH-TEMPLATE (hire-to-win): Type A individual-credential opportunities PRG
# could win IF a contingent credentialed teaming partner is secured.
_TEMPLATE_COLS = [
    ("Credential Needed", "credential_needed"),
    ("ACTION", "template_action"),
    ("Title", "title"),
    ("Agency", "agency"),
    ("Set-Aside", "setaside"),
    ("NAICS", "naics"),
    ("Respond By", "response_deadline"),
    ("Solicitation #", "solicitation"),
    ("Link", "link"),
]


_SUMMARY_COLS = [
    ("Metric", "metric"),
    ("Value", "value"),
    ("Detail", "detail"),
]


def _summary_rows(results, recompetes=None):
    """A one-glance Summary tab: run totals, verdict mix, and kill-reason
    breakdown — so the workbook opens on 'what did this run find' before the
    1,000-row tabs."""
    recompetes = recompetes or []
    from collections import Counter
    verdicts = Counter(r["verdict"] for r in results)
    kills = Counter(r["kill_gate"] for r in results if r["disqualified"])
    primary, low_barrier, subs, solo, international = _split_sections(results)
    rows = [
        {"metric": "Official source", "value": "SAM.gov Opportunities API + "
         "USASpending.gov", "detail": "live federal opportunities & awards"},
        {"metric": "Opportunities reviewed (after dedupe)", "value": len(results),
         "detail": "unique notices screened this run"},
        {"metric": "BID (pursue as prime)", "value": verdicts.get("BID", 0),
         "detail": "eligible + in-lane + winnable now"},
        {"metric": "SHORT-FUSE", "value": verdicts.get("SHORT-FUSE", 0),
         "detail": "due within ~5 business days"},
        {"metric": "RESEARCH", "value": verdicts.get("RESEARCH", 0),
         "detail": "winnable but needs prep / weak edge"},
        {"metric": "WATCH-TEMPLATE (hire-to-win)",
         "value": sum(1 for r in results if r.get("watch_template")),
         "detail": "individual credential — line up a contingent partner"},
        {"metric": "WATCH (shape the RFP)",
         "value": sum(1 for r in results if r["is_future"]),
         "detail": "pre-sol / sources sought — respond to shape"},
        {"metric": "NO-BID (screened out)", "value": verdicts.get("NO-BID", 0),
         "detail": "auto-killed — see KILL LOG"},
        {"metric": "EXPIRED / MARKET-INTEL",
         "value": verdicts.get("EXPIRED", 0) + verdicts.get("MARKET-INTEL", 0),
         "detail": "past deadline or already awarded"},
        {"metric": "—", "value": "—", "detail": "— lane counts —"},
        {"metric": "Top 10 shortlist", "value": len(_top10(results)),
         "detail": "do these first"},
        {"metric": "Solo-friendly (1-person)", "value": len(solo),
         "detail": "one person + AI can deliver"},
        {"metric": "Vibe-code (web/software)",
         "value": sum(1 for r in results if r.get("is_vibecode")),
         "detail": "Screendot + Digipop builds"},
        {"metric": "Core opportunities", "value": len(primary),
         "detail": "clinical/research in-lane"},
        {"metric": "Low-barrier (warm body)", "value": len(low_barrier),
         "detail": "win with qualified personnel"},
        {"metric": "International (consulting)", "value": len(international),
         "detail": "overseas — flagged, no US SB edge"},
        {"metric": "Subcontracting / teaming", "value": len(subs),
         "detail": "team under a prime"},
        {"metric": "Recompete Radar", "value": len(recompetes),
         "detail": "incumbents + upcoming rebids"},
    ]
    if kills:
        rows.append({"metric": "—", "value": "—", "detail": "— why killed —"})
        for gate, n in kills.most_common():
            rows.append({"metric": gate, "value": n, "detail": "auto-screened kills"})
    return rows


_NAICS_GAP_COLS = [
    ("NAICS (not in PRG profile)", "naics"),
    ("# Winnable Opportunities", "count"),
    ("Best Win Score", "best_win"),
    ("Top Set-Aside", "setaside"),
    ("Example Opportunity", "example_title"),
    ("Example Agency", "example_agency"),
    ("Recommendation", "recommendation"),
]


def _naics_gap_rows(results):
    """Wide-net payoff: NAICS codes NOT in PRG's profile that carry set-aside-
    eligible, un-killed opportunities — i.e. codes worth ADDING to the SAM
    registration. Grouped by NAICS, ranked by how many winnable notices sit
    there."""
    from collections import defaultdict
    groups = defaultdict(list)
    for r in results:
        if (not r.get("naics_in_profile") and r["raw_setaside_eligible"]
                and not r["disqualified"] and not r["is_awarded"]
                and not r["is_expired"] and not r["is_future"]
                and not r.get("watch_template")):
            groups[r["naics"]].append(r)
    rows = []
    for code, rs in groups.items():
        rs.sort(key=lambda r: r["win_score"], reverse=True)
        best = rs[0]
        rows.append({
            "naics": code,
            "count": len(rs),
            "best_win": best["win_score"],
            "setaside": best["setaside"],
            "example_title": best["title"],
            "example_agency": best["agency"],
            "recommendation": "Add this NAICS to SAM.gov (Reps & Certs) to be "
                              "listed and to match agency market research",
        })
    rows.sort(key=lambda x: (x["count"], x["best_win"]), reverse=True)
    return rows


def _top10(results):
    """The single best-bet shortlist: highest-scoring eligible, pursuable
    opportunities (not awarded, not expired), ranked, capped at 10."""
    pool = [r for r in results if r["raw_setaside_eligible"]
            and not r["is_awarded"] and not r["is_expired"]
            and not r["disqualified"] and not r["is_future"]
            and not r.get("watch_template")]
    pool.sort(key=lambda r: r["win_score"], reverse=True)
    # Return shallow copies with a rank, so we don't mutate the shared results.
    top = []
    for i, r in enumerate(pool[:10], 1):
        row = dict(r)
        row["rank"] = i
        top.append(row)
    return top


def export_spreadsheet(results, path, recompetes=None, grants=None, sbir=None,
                       reliefweb=None):
    """Write results to an Excel workbook (.xlsx) if openpyxl is available,
    otherwise fall back to a set of CSV files. Returns the path actually written.
    """
    primary, low_barrier, subs, solo, international = _split_sections(results)
    killed = sorted([r for r in results if r["disqualified"]
                     and not r["is_awarded"] and not r["is_expired"]],
                    key=lambda r: r["kill_gate"])
    watchlist = sorted([r for r in results if r["is_future"]],
                       key=lambda r: r["win_score"], reverse=True)
    vibecode = sorted([r for r in results if r.get("is_vibecode")],
                      key=lambda r: r["win_score"], reverse=True)
    sheets = [
        ("Summary", _SUMMARY_COLS, _summary_rows(results, recompetes)),
        ("Top 10 - Do These First", _TOP10_COLS, _top10(results)),
        ("Solo-Friendly (1-Person)", _SOLO_COLS, solo),
        ("Vibe-Code (Web-Software)", _VIBECODE_COLS, vibecode),
        ("Core Opportunities", _CORE_COLS, primary),
        ("Low-Barrier (Warm Body)", _LOWBAR_COLS, low_barrier),
        ("International (Consulting)", _CORE_COLS, international),
    ]
    # Priority-agency tabs (VA / HHS / DTRA): live, still-open opportunities at
    # each target agency, drawn from the same SAM.gov pull, ranked by win score.
    for a in AGENCY_TARGETS:
        agency_rows = sorted(
            [r for r in results if r.get("agency_bucket") == a["key"]
             and not r["is_awarded"] and not r["is_expired"]],
            key=lambda r: r["win_score"], reverse=True,
        )
        sheets.append((a["label"], _CORE_COLS, agency_rows))
    template_rows = sorted([r for r in results if r.get("watch_template")],
                           key=lambda r: r["win_score"], reverse=True)
    sheets += [
        ("Subcontracting", _SUB_COLS, subs),
        ("Recompete Radar", _RECOMPETE_COLS, recompetes or []),
        ("WATCH-TEMPLATE (hire-to-win)", _TEMPLATE_COLS, template_rows),
        ("WATCHLIST (prep, future)", _WATCHLIST_COLS, watchlist),
        ("NAICS Gap - Add to SAM", _NAICS_GAP_COLS, _naics_gap_rows(results)),
        ("Grants.gov (Assistance $)", _GRANTS_COLS, grants or []),
        ("SBIR-STTR (Open R&D)", _SBIR_COLS, sbir or []),
        ("ReliefWeb (Intl Gigs)", _RELIEFWEB_COLS, reliefweb or []),
        ("International Sources", _INTL_SRC_COLS, _intl_source_rows()),
        ("SLED + Marketplaces", _INTL_SRC_COLS, _sled_source_rows()),
        ("KILL LOG (screened out)", _KILL_COLS, killed),
        ("Resources (Hunt & Help)", _RESOURCE_COLS, _resource_rows()),
    ]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return _export_csv(sheets, path)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    wrap = Alignment(vertical="top", wrap_text=True)
    band_fill = {
        "Green": PatternFill("solid", fgColor="C6E7D0"),
        "Yellow": PatternFill("solid", fgColor="FBEBB0"),
        "Red": PatternFill("solid", fgColor="F3C2BC"),
    }

    for sheet_name, cols, rows in sheets:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel caps sheet names at 31
        headers = [h for h, _ in cols]
        ws.append(headers)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = wrap
        # Best bets first.
        rows = sorted(rows, key=lambda r: r.get("win_score", 0), reverse=True)
        rag_cols = [i for i, h in enumerate(headers, start=1)
                    if h in ("Win Score", "Rating")]
        link_col = next((i for i, h in enumerate(headers, 1) if h == "Link"), None)
        sol_col = next((i for i, h in enumerate(headers, 1)
                        if h == "Solicitation #"), None)
        url_col = next((i for i, h in enumerate(headers, 1) if h == "URL"), None)
        link_font = Font(color="0563C1", underline="single")
        for r in rows:
            ws.append([_cell(r, spec) for _, spec in cols])
            fill = band_fill.get(r.get("win_band"))
            if fill:
                for ci in rag_cols:
                    ws.cell(row=ws.max_row, column=ci).fill = fill
            # Make the SAM.gov notice clickable from the Link and Solicitation cells.
            url = r.get("link")
            if _is_http(url):
                if link_col:
                    c = ws.cell(row=ws.max_row, column=link_col)
                    c.value = "Open in SAM.gov"
                    c.hyperlink = url
                    c.font = link_font
                if sol_col:
                    c = ws.cell(row=ws.max_row, column=sol_col)
                    c.hyperlink = url
                    c.font = link_font
            # A literal "URL" column (Resources tab) hyperlinks to its own value.
            if url_col:
                c = ws.cell(row=ws.max_row, column=url_col)
                if _is_http(c.value):
                    c.hyperlink = c.value
                    c.font = link_font
        # Column widths + wrapping.
        widths = {
            "Title": 45, "Agency": 26, "Reason / Gap Analysis": 60,
            "Why It's Winnable": 50, "Subcontracting Angle": 50, "Link": 32,
            "Est. Value (Revenue)": 20, "Why One-Person Doable": 55,
            "Win Score": 10, "Rating": 12, "Verdict": 26, "Fit for PRG": 14,
            "Go / No-Go": 12, "Est. LOE": 28, "Flags": 44,
            "Set-Aside": 14, "Eligible as LLC?": 24, "Personnel (FTE)": 13,
            "Hire New or Take Over?": 34, "Solo-Doable?": 12, "Timeframe": 30,
            "Location": 26, "Intl / CONUS": 13, "Buyer Type": 24,
            "Solicitation #": 20, "Category": 20, "Resource": 42,
            "What it's for": 66, "URL": 52, "Priority Agency": 16,
            "Credential Needed": 30, "ACTION": 62, "Respond By": 20,
            "Fastest Paid Route": 26, "Services to Pitch": 52,
            "Gaps to Close (before award)": 46, "Min Qualifications": 52,
            "Founder PP Usable?": 30, "POC (from notice)": 34,
            "Notice ID": 30, "What to Build (matched)": 34,
            "Metric": 34, "Value": 12, "Detail": 44,
            "NAICS (not in PRG profile)": 24, "# Winnable Opportunities": 22,
            "Best Win Score": 14, "Top Set-Aside": 22, "Example Opportunity": 46,
            "Example Agency": 28, "Recommendation": 54,
            "Opportunity #": 22, "ALN / CFDA": 16, "Opens": 14, "Closes": 14,
            "Solicitation": 46, "Program": 14, "Branch": 22, "Topics": 12,
            "Source (register here)": 40, "What it covers / fit for PRG": 70,
            "Status": 14, "Organization": 34, "Country": 22, "Experience": 16,
        }
        for idx, (h, _) in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = widths.get(h, 16)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrap
        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    if not path.lower().endswith(".xlsx"):
        path = path + ".xlsx"
    try:
        wb.save(path)
    except PermissionError:
        # File is open in Excel (locked). Save under a timestamped name instead
        # so a long scan is never lost.
        alt = _timestamped_alt(path)
        wb.save(alt)
        sys.stderr.write(
            f"NOTE: '{os.path.basename(path)}' was open/locked — saved as "
            f"'{os.path.basename(alt)}' instead.\n"
        )
        return alt
    return path


def _export_csv(sheets, path):
    """Fallback: write one CSV per section next to `path`."""
    import csv
    base = path[:-5] if path.lower().endswith(".xlsx") else path
    base = base[:-4] if base.lower().endswith(".csv") else base
    written = []
    for sheet_name, cols, rows in sheets:
        slug = sheet_name.lower().replace(" ", "_").replace("(", "").replace(")", "")
        fname = f"{base}_{slug}.csv"
        with open(fname, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow([h for h, _ in cols])
            for r in rows:
                writer.writerow([_cell(r, spec) for _, spec in cols])
        written.append(fname)
    return ", ".join(written)


def _default_output_path():
    """Best-guess path for the output workbook: PRG_Contracts.xlsx on the
    user's Desktop (handles OneDrive-redirected Desktops), falling back to the
    home directory if no Desktop folder is found.
    """
    desktop = _resolve_desktop()
    return os.path.join(desktop, "PRG_Contracts.xlsx")


def _resolve_desktop():
    """Return the user's REAL, visible Desktop folder.

    On Windows we ask the registry for the actual Desktop known-folder path
    (this correctly returns the OneDrive-redirected desktop when that is what
    the user sees). Elsewhere / on failure we fall back to the best guess,
    preferring a OneDrive Desktop over the hidden literal one.
    """
    # Authoritative on Windows: the Shell Folders registry key holds the fully
    # resolved Desktop path, redirection included.
    try:
        import winreg  # Windows only
        key = r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders"
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, key) as k:
            val, _ = winreg.QueryValueEx(k, "Desktop")
            val = os.path.expandvars(val)
            if os.path.isdir(val):
                return val
    except Exception:
        pass

    home = os.path.expanduser("~")
    profile = os.environ.get("USERPROFILE", home)
    for folder in (
        os.path.join(profile, "OneDrive", "Desktop"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(profile, "Desktop"),
        os.path.join(home, "Desktop"),
    ):
        if os.path.isdir(folder):
            return folder
    return home


def _desktop_path(filename):
    """Return `filename` on the user's Desktop (or home dir fallback)."""
    base = _default_output_path()
    return os.path.join(os.path.dirname(base), filename)


def _load_vet_cache():
    """FIX 6 — load the content-hash vetting cache (empty on any failure)."""
    try:
        with open(CACHE_PATH, encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def _save_vet_cache(cache):
    try:
        with open(CACHE_PATH, "w", encoding="utf-8") as fh:
            json.dump(cache, fh)
    except Exception:
        pass


def apply_vet_cache(results):
    """FIX 6 — annotate opportunities whose content hash matches a prior run as
    'already vetted [date]: [verdict]', and update the cache. Returns how many
    were previously vetted."""
    cache = _load_vet_cache()
    today_iso = dt.date.today().isoformat()
    seen_before = 0
    for r in results:
        h = r.get("content_hash")
        if not h:
            continue
        prior = cache.get(h)
        if prior and prior.get("verdict") == r["verdict"]:
            r["previously_vetted"] = (
                f"already vetted {prior.get('date')}: {prior.get('verdict')}")
            seen_before += 1
        cache[h] = {"date": today_iso, "verdict": r["verdict"],
                    "sol": r["solicitation"]}
    _save_vet_cache(cache)
    return seen_before


def _timestamped_alt(path):
    """Insert a HH-MM-SS stamp before the extension, e.g. for locked files."""
    root, ext = os.path.splitext(path)
    stamp = dt.datetime.now().strftime("%H-%M-%S")
    return f"{root}_{stamp}{ext}"


# ---------------------------------------------------------------------------
# 6. EXECUTIVE HTML REPORT
# ---------------------------------------------------------------------------

def _html_escape(text):
    return (str(text).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def _svg_hbars(items, value_fmt, default_color="#1E3A5F", width=680):
    """Horizontal bar chart as inline SVG. `items` = list of (label, value)
    or (label, value, color). Returns an SVG string."""
    items = [it for it in items if (it[1] or 0) > 0]
    if not items:
        return '<p class="empty">No published values to chart yet.</p>'
    maxv = max(it[1] for it in items) or 1
    row_h, pad, label_w = 34, 10, 200
    bar_max = width - label_w - 96
    height = pad * 2 + row_h * len(items)
    out = [f'<svg viewBox="0 0 {width} {height}" class="chart" '
           f'preserveAspectRatio="xMinYMin meet" role="img">']
    y = pad
    for it in items:
        label, val = it[0], it[1]
        color = it[2] if len(it) > 2 else default_color
        bw = max(2.0, (val / maxv) * bar_max)
        yc = y + row_h / 2
        out.append(f'<text x="0" y="{yc + 4:.0f}" class="c-lab">'
                   f'{_html_escape(_clean(label, 26))}</text>')
        out.append(f'<rect x="{label_w}" y="{y + 6:.0f}" width="{bw:.1f}" '
                   f'height="{row_h - 14}" rx="3" fill="{color}"/>')
        out.append(f'<text x="{label_w + bw + 8:.1f}" y="{yc + 4:.0f}" '
                   f'class="c-val">{_html_escape(value_fmt(val))}</text>')
        y += row_h
    out.append("</svg>")
    return "".join(out)


_RAG_HEX = {"Green": "#2E7D4F", "Yellow": "#D9A62A", "Red": "#C0392B"}


def _key_findings_html(ranked, eligible):
    """Build the 'Key Findings & Things to Keep in Mind' review callout."""
    if not eligible:
        return ('<section class="findings"><h2>Key Findings</h2>'
                '<p class="empty">No eligible opportunities this run — widen the '
                'search window and check again.</p></section>')

    closing = sorted(
        [r for r in eligible
         if r["deadline_days"] is not None and 0 <= r["deadline_days"] <= 14],
        key=lambda r: r["deadline_days"],
    )
    solos = [r for r in ranked if r["is_solo"]]
    valued = [r for r in eligible if r["value_num"]]
    top_value = max(valued, key=lambda r: r["value_num"]) if valued else None
    green_yellow = [r for r in ranked if r["win_band"] in ("Green", "Yellow")]

    items = []

    # Urgent deadlines first.
    if closing:
        names = "; ".join(
            f'{_html_escape(_clean(r["title"], 40))} '
            f'(<b>{r["deadline_days"]}d</b>, {_html_escape(r["solicitation"])})'
            for r in closing[:4])
        items.append(('🔴', 'Closing soon — act now',
                      f'{len(closing)} eligible opportunit'
                      f'{"y" if len(closing) == 1 else "ies"} close within 14 days: '
                      f'{names}.'))

    # Best bets.
    if green_yellow:
        names = "; ".join(
            f'{_html_escape(_clean(r["title"], 40))} '
            f'({r["win_emoji"]} {r["win_score"]})' for r in green_yellow[:3])
        items.append(('⭐', 'Strongest fits to pursue',
                      f'{names}.'))

    # Best solo play.
    if solos:
        r = solos[0]
        items.append(('🧑‍💻', 'Best solo (you + AI) play',
                      f'{_html_escape(_clean(r["title"], 50))} '
                      f'({_html_escape(r["solicitation"])}) — {r["win_emoji"]} '
                      f'{r["win_score"]}, {_html_escape(r["value_display"])}.'))

    # Biggest opportunity.
    if top_value:
        items.append(('💰', 'Largest published value',
                      f'{_html_escape(_clean(top_value["title"], 50))} — '
                      f'{_html_escape(top_value["value_display"])} '
                      f'({_html_escape(top_value["agency"])}).'))

    # Standing reminders.
    items.append(('✅', 'Keep in mind',
                  'Respond same-day where you can; keep SDVOSB certification '
                  'active through award; SDVOSB set-asides require ≥50% '
                  'self-performance (FAR 52.219-14); a Sources Sought response '
                  'is cheap and can shape the eventual solicitation.'))

    rows = "".join(
        f'<div class="find"><span class="find-i">{ic}</span>'
        f'<div><b>{_html_escape(t)}</b> — {body}</div></div>'
        for ic, t, body in items)
    return (f'<section class="findings"><h2>Key Findings &amp; Things to Keep '
            f'in Mind</h2>{rows}</section>')


_ROLE_SHORT = {
    "BID AS PRIME": "Prime",
    "PRIME WITH TEAMING PARTNER": "Prime + team",
    "PURSUE AS SUBCONTRACTOR": "Sub",
    "RESPOND TO SOURCES SOUGHT": "Respond (SS)",
    "MONITOR / PREPOSITION": "Monitor",
    "PASS": "Pass",
}


def _role_short(role):
    """Compact label for the six-way pursuit role, for narrow table cells."""
    return _ROLE_SHORT.get(role, role or "—")


def _role_filter_key(role):
    """Map a pursuit role to the matrix filter bucket it belongs to."""
    if role == "BID AS PRIME":
        return "prime"
    if role == "PRIME WITH TEAMING PARTNER":
        return "team"
    if role == "PURSUE AS SUBCONTRACTOR":
        return "sub"
    return ""


def export_html_report(results, path, days, recompetes=None, grants=None,
                       sbir=None, reliefweb=None):
    """Write a self-contained executive HTML report of the opportunity pipeline."""
    recompetes = recompetes or []
    grants = grants or []
    sbir = sbir or []
    reliefweb = reliefweb or []
    # Pursue-as-prime view: eligible AND not already awarded (awarded ones live
    # in the Subcontracting category only).
    eligible = [r for r in results if r["raw_setaside_eligible"]
                and not r["is_awarded"] and not r["is_expired"]
                and not r["disqualified"] and not r["is_future"]
                and not r.get("watch_template")]
    # Rank by capture priority first (win score breaks ties) — the whole
    # report is ordered by the aggregation model, not the raw win score.
    ranked = sorted(eligible, key=lambda r: (r.get("priority", 0),
                                             r["win_score"]), reverse=True)
    top10 = ranked[:10]
    killed = [r for r in results if r["disqualified"]
                     and not r["is_awarded"] and not r["is_expired"]]
    watchlist = sorted([r for r in results if r["is_future"]],
                       key=lambda r: (r.get("priority", 0), r["win_score"]),
                       reverse=True)

    # --- Capture-model role groups (best-of, for the executive summary) ---
    def _best_of(role):
        rows = [r for r in ranked if r.get("role") == role]
        return rows[0] if rows else None
    best_prime = _best_of("BID AS PRIME")
    best_team = _best_of("PRIME WITH TEAMING PARTNER")
    best_sub = next((r for r in sorted(
        results, key=lambda r: (r.get("priority", 0), r["win_score"]),
        reverse=True) if r.get("role") == "PURSUE AS SUBCONTRACTOR"), None)
    best_pp = next((r for r in ranked if r.get("pp_value") == "High"), None)

    n_prime = sum(1 for r in ranked if r.get("role") == "BID AS PRIME")
    n_team = sum(1 for r in ranked
                 if r.get("role") == "PRIME WITH TEAMING PARTNER")
    n_sub = sum(1 for r in results
                if r.get("role") == "PURSUE AS SUBCONTRACTOR")

    # --- KPIs ---
    total_value = sum(r["value_num"] for r in eligible if r["value_num"])
    n_active = len(eligible)
    n_solo = sum(1 for r in eligible if r["is_solo"])
    n_sdvosb = sum(1 for r in eligible if r["is_sdvosb"])
    n_green = sum(1 for r in eligible if r["win_band"] == "Green")

    agency_counts = Counter(r["agency"] for r in eligible if r["agency"] != "N/A")

    # --- Chart data ---
    agency_value = Counter()
    for r in eligible:
        if r["value_num"]:
            agency_value[r["agency"]] += r["value_num"]
    chart_agency = [(a, v) for a, v in agency_value.most_common(8)]
    model_counts = Counter(r.get("fulfillment_model", "—") for r in ranked)
    _model_hex = {"founder-delivered": "#2E7D4F",
                  "professional services": "#1E5F9E",
                  "staffing (W2/1099)": "#7A5BA6",
                  "trade-sub management": "#9A6B1C",
                  "value-added supply": "#C0392B"}
    chart_models = [(m, n, _model_hex.get(m, "#5A6B7E"))
                    for m, n in model_counts.most_common(6)]

    n_intl = sum(1 for r in eligible if r["is_international"])
    kpis = [
        ("Pipeline Value", _format_currency(total_value) if total_value else "N/A",
         "sum of published values"),
        ("Prime Plays", str(n_prime), "bid as prime"),
        ("Teaming Plays", str(n_team), "prime + trade/staffing sub"),
        ("Subcontract Targets", str(n_sub), "pursue as sub"),
        ("Best Bets (Green)", str(n_green), "highest priority"),
        ("SDVOSB Set-Asides", str(n_sdvosb), "your direct lane"),
    ]

    # --- Assemble HTML ---
    p = ['<!doctype html><html lang="en"><head><meta charset="utf-8">',
         '<meta name="viewport" content="width=device-width, initial-scale=1">',
         f'<title>PRG Pipeline Executive Report — {dt.date.today().isoformat()}</title>',
         "<style>", _REPORT_CSS, "</style></head><body>"]
    p.append('<div class="wrap">')

    # Header
    p.append('<header><p class="eyebrow">Executive Pipeline Report</p>'
             f'<h1>{COMPANY_NAME} — Federal Opportunity Pipeline</h1>'
             f'<p class="sub">{SOCIOECONOMIC_STATUS} · generated '
             f'{dt.date.today().isoformat()} · last {days} days · '
             f'{len(results)} notices screened</p></header>')

    # KPI cards — each count card is a clickable filter for the matrix below.
    kpi_filter = {
        "Prime Plays": "role:prime",
        "Teaming Plays": "role:team",
        "Subcontract Targets": "role:sub",
        "Best Bets (Green)": "band:Green",
        "SDVOSB Set-Asides": "sdvosb",
    }
    p.append('<section class="kpis">')
    for label, val, sub in kpis:
        filt = kpi_filter.get(label)
        attrs = (f' class="kpi clickable" data-filter="{filt}" tabindex="0" '
                 f'role="button"' if filt else ' class="kpi"')
        hint = ' <span class="kpi-go">▸ click to filter</span>' if filt else ''
        p.append(f'<div{attrs}><div class="kpi-v">{_html_escape(val)}</div>'
                 f'<div class="kpi-l">{_html_escape(label)}</div>'
                 f'<div class="kpi-s">{_html_escape(sub)}{hint}</div></div>')
    p.append('</section>')

    # Capture Executive Summary — the aggregation-model headline (best prime,
    # teaming, subcontract, and past-performance plays this window).
    def _summary_line(label, r, sub_role=False):
        if not r:
            return (f'<div class="find"><div class="find-i">—</div><div>'
                    f'<b>{label}:</b> <span class="muted">none this window'
                    f'</span></div></div>')
        title = _html_escape(_clean(r["title"], 66))
        if _is_http(r["link"]):
            title = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                     f'rel="noopener">{title}</a>')
        meta = (f'priority <b>{r.get("priority", 0)}</b> · '
                f'{_html_escape(r.get("fulfillment_model", ""))} · '
                f'GM {_html_escape(r.get("margin_est", ""))} · '
                f'{_html_escape(r["agency"])}')
        dl = str(r["response_deadline"]).split("T")[0]
        if r["deadline_days"] is not None and 0 <= r["deadline_days"] <= 7:
            meta += f' · <b class="urgent">due {dl} ({r["deadline_days"]}d)</b>'
        return (f'<div class="find"><div class="find-i">★</div><div>'
                f'<b>{label}:</b> {title}<br><span class="muted" '
                f'style="font-size:12px">{meta}</span></div></div>')
    p.append('<section class="findings"><h2>Capture Executive Summary</h2>')
    p.append(f'<p class="muted" style="font-size:12.5px;margin:2px 0 8px">'
             f'Aggregation model: PRG primes and manages; execution is sourced '
             f'through subs, staffing, trades, and distributors. '
             f'{n_prime} prime · {n_team} teaming · {n_sub} subcontract '
             f'targets from {len(results)} notices screened.</p>')
    p.append(_summary_line("Best immediate prime", best_prime))
    p.append(_summary_line("Best teaming play", best_team))
    p.append(_summary_line("Best subcontract target", best_sub))
    p.append(_summary_line("Best past-performance builder", best_pp))
    p.append('</section>')

    # Key findings / review
    p.append(_key_findings_html(ranked, eligible))

    # Top 10 — Do These First
    p.append('<section><h2>🎯 Top 10 — Do These First</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Your highest-priority, still-open opportunities — ranked by the '
             'capture priority score (win probability, margin, delegation, '
             'speed to cash). Act on these first.</p>')
    if top10:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>#</th><th>Priority</th><th>Rating</th><th>Role</th>'
                 '<th>Title</th><th>Agency</th><th>Fulfillment</th>'
                 '<th>GM%</th><th>Est. Value</th><th>Respond By</th>'
                 '</tr></thead><tbody>')
        for i, r in enumerate(top10, 1):
            chip = (f'<span class="chip" style="background:'
                    f'{_RAG_HEX[r["win_band"]]}">{r["win_band"]}</span>')
            title = _html_escape(_clean(r["title"], 54))
            if _is_http(r["link"]):
                title = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                         f'rel="noopener">{title}</a>')
            dl = str(r["response_deadline"]).split("T")[0]
            if r["deadline_days"] is not None and 0 <= r["deadline_days"] <= 7:
                dl = f'<b class="urgent">{dl} · {r["deadline_days"]}d 🔴</b>'
            p.append(f"<tr><td class='num'><b>{i}</b></td>"
                     f"<td class='num'><b>{r.get('priority', 0)}</b></td>"
                     f"<td>{chip}</td>"
                     f"<td>{_html_escape(_role_short(r.get('role', '')))}</td>"
                     f"<td>{title}</td>"
                     f"<td>{_html_escape(r['agency'])}</td>"
                     f"<td>{_html_escape(r.get('fulfillment_model', ''))}</td>"
                     f"<td class='num'>{_html_escape(r.get('margin_est', ''))}</td>"
                     f"<td class='num'>{_html_escape(r['value_display'])}</td>"
                     f"<td>{dl}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No eligible open opportunities this run.</p>')
    p.append('</section>')

    # Charts
    p.append('<section class="charts">')
    p.append('<div class="card"><h3>Pipeline Value by Agency</h3>'
             + _svg_hbars(chart_agency, _format_currency, "#1E3A5F") + '</div>')
    p.append('<div class="card"><h3>Opportunities by Fulfillment Model</h3>'
             + _svg_hbars(chart_models, lambda v: str(int(v))) + '</div>')
    p.append('</section>')

    # Contract matrix (interactive: filter buttons + live search).
    p.append('<section><h2>Contract / Opportunity Matrix</h2>')
    p.append('<div class="filterbar">'
             '<button class="fb active" data-filter="all">All</button>'
             '<button class="fb" data-filter="role:prime">Prime</button>'
             '<button class="fb" data-filter="role:team">Prime + team</button>'
             '<button class="fb" data-filter="role:sub">Subcontract</button>'
             '<button class="fb" data-filter="band:Green">🟢 Best bets</button>'
             '<button class="fb" data-filter="band:Yellow">🟡 On the fence</button>'
             '<button class="fb" data-filter="solo">Founder-deliverable</button>'
             '<button class="fb" data-filter="intl">🌍 International</button>'
             '<button class="fb" data-filter="sdvosb">SDVOSB</button>'
             '<input id="q" class="fsearch" type="search" '
             'placeholder="Search title / agency / #…">'
             '<span id="count" class="count"></span></div>')
    p.append('<div class="scroll"><table id="matrix"><thead><tr>'
             '<th>Priority</th><th>Rating</th><th>Role</th>'
             '<th>Solicitation #</th><th>Agency</th><th>Fulfillment</th>'
             '<th>GM%</th><th>Est. Value</th><th>Set-Aside</th>'
             '<th>NAICS / PSC</th><th>Location</th><th>Respond By</th>'
             '</tr></thead><tbody id="matrixBody">')
    for r in ranked:
        chip = (f'<span class="chip" style="background:{_RAG_HEX[r["win_band"]]}">'
                f'{r["win_band"]}</span>')
        deadline = str(r["response_deadline"]).split("T")[0]
        dd = r["deadline_days"]
        if dd is not None and 0 <= dd <= 7:
            deadline = f'<b class="urgent">{deadline} · {dd}d 🔴</b>'
        elif dd is not None and 0 <= dd <= 14:
            deadline = f'{deadline} · {dd}d 🟠'
        naics_psc = r["naics"] + (f" / {r['psc']}" if r["psc"] else "")
        sol = _html_escape(r["solicitation"])
        if _is_http(r["link"]):
            sol = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                   f'rel="noopener">{sol}</a>')
        searchtext = _html_escape(
            f"{r['solicitation']} {r['title']} {r['agency']} {r['naics']}".lower())
        p.append(
            f'<tr data-band="{r["win_band"]}" '
            f'data-role="{_role_filter_key(r.get("role", ""))}" '
            f'data-solo="{1 if r["is_solo"] else 0}" '
            f'data-intl="{1 if r["is_international"] else 0}" '
            f'data-sdvosb="{1 if r["is_sdvosb"] else 0}" '
            f'data-text="{searchtext}">'
            f"<td class='num'><b>{r.get('priority', 0)}</b></td>"
            f"<td>{chip}</td>"
            f"<td>{_html_escape(_role_short(r.get('role', '')))}</td>"
            f"<td>{sol}</td>"
            f"<td>{_html_escape(r['agency'])}</td>"
            f"<td>{_html_escape(r.get('fulfillment_model', ''))}</td>"
            f"<td class='num'>{_html_escape(r.get('margin_est', ''))}</td>"
            f"<td class='num'>{_html_escape(r['value_display'])}</td>"
            f"<td>{_html_escape(r['setaside'])}</td>"
            f"<td>{_html_escape(naics_psc)}</td>"
            f"<td>{_html_escape(r['location'])}</td>"
            f"<td>{deadline}</td></tr>")
    if not ranked:
        p.append('<tr><td colspan="12" class="empty">No eligible '
                 'opportunities in this window.</td></tr>')
    p.append('</tbody></table></div></section>')

    # Deep dive (top opportunities)
    p.append('<section><h2>Deep Dive — Top Opportunities</h2>')
    for r in ranked[:8]:
        p.append('<div class="dive">')
        p.append(f'<div class="dive-h"><span class="chip" '
                 f'style="background:{_RAG_HEX[r["win_band"]]}">{r["win_band"]} '
                 f'· {r["win_score"]}</span> <strong>{_html_escape(r["title"])}</strong>'
                 f' <span class="muted">({_html_escape(r["solicitation"])})</span></div>')
        # Capture line — role, fulfillment model, margin, and priority up front.
        p.append(f'<p><b>Capture:</b> {_html_escape(r.get("role", ""))} · '
                 f'{_html_escape(r.get("fulfillment_model", ""))} · '
                 f'est. GM {_html_escape(r.get("margin_est", ""))} · '
                 f'priority {r.get("priority", 0)}'
                 + (f' · <span class="muted">labor: '
                    f'{_html_escape(r.get("labor_plan", ""))}</span>'
                    if r.get("labor_plan") else '') + '.</p>')
        if r.get("passthrough_risk") == "high":
            p.append('<p class="urgent" style="font-size:12.5px"><b>⚠ '
                     'Ostensible-subcontractor risk:</b> scope is one '
                     'indivisible specialty — structure real PRG management '
                     '(QC, invoicing, ≥ the self-perform floor) or it reads as '
                     'a pass-through.</p>')
        scope = r["naics_desc"] or "professional services"
        p.append(f'<p><b>Scope:</b> {_html_escape(r["title"])} — '
                 f'{_html_escape(r["agency"])} ({_html_escape(scope)}).</p>')
        if r.get("buyer_is_international"):
            p.append('<p><b>Buyer:</b> <span class="muted">'
                     + _html_escape(r["buyer_type"])
                     + ' — recognizes no US SDVOSB / small-business preference, '
                     'so PRG\'s structural edge does not apply. Not first-contract '
                     'material.</span></p>')
        # Personnel roster
        if r["poc"]:
            roster = []
            for c in r["poc"][:3]:
                bits = [b for b in [c["name"], c["email"], c["phone"]] if b]
                if bits:
                    role = f'{c["type"]}: ' if c["type"] else ""
                    roster.append(role + " · ".join(_html_escape(b) for b in bits))
            if roster:
                p.append('<p><b>Gov. Contacts:</b> ' + " &nbsp;|&nbsp; ".join(roster)
                         + '</p>')
        else:
            p.append('<p><b>Gov. Contacts:</b> <span class="muted">not listed in '
                     'notice — check the SAM.gov posting.</span></p>')
        # Risk / action
        risks = [f"Respond by {str(r['response_deadline']).split('T')[0]}"]
        if r["is_sdvosb"]:
            risks.append("keep SDVOSB certification active through award")
        if r["fte_num"] and r["fte_num"] >= 1:
            risks.append("50% self-performance applies (FAR 52.219-14)")
        if "inherit" in r["staffing_type"].lower():
            risks.append("incumbent present — differentiate to unseat")
        p.append('<p><b>Risk / Action:</b> ' + "; ".join(_html_escape(x) for x in risks)
                 + '.</p>')
        if _is_http(r["link"]):
            p.append(f'<p><a href="{_html_escape(r["link"])}" target="_blank" '
                     f'rel="noopener">View full notice on SAM.gov →</a></p>')
        p.append('</div>')
    if not ranked:
        p.append('<p class="empty">Nothing to deep-dive this run.</p>')
    p.append('</section>')

    # International opportunities — its own category (overseas consulting).
    intl = sorted([r for r in eligible if r["is_international"]],
                  key=lambda r: r["win_score"], reverse=True)
    p.append('<section><h2>🌍 International Opportunities (Overseas Consulting)</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Eligible opportunities with an overseas place of performance — '
             'kept as a separate category from domestic (CONUS) work.</p>')
    if intl:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Rating</th><th>Win</th><th>Solicitation #</th><th>Agency</th>'
                 '<th>Location</th><th>Est. Value</th><th>Set-Aside</th>'
                 '<th>Solo</th><th>Respond By</th></tr></thead><tbody>')
        for r in intl:
            chip = (f'<span class="chip" style="background:'
                    f'{_RAG_HEX[r["win_band"]]}">{r["win_band"]}</span>')
            sol = _html_escape(r["solicitation"])
            if _is_http(r["link"]):
                sol = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                       f'rel="noopener">{sol}</a>')
            p.append("<tr>"
                     f"<td>{chip}</td><td class='num'>{r['win_score']}</td>"
                     f"<td>{sol}</td><td>{_html_escape(r['agency'])}</td>"
                     f"<td>{_html_escape(r['location'])}</td>"
                     f"<td class='num'>{_html_escape(r['value_display'])}</td>"
                     f"<td>{_html_escape(r['setaside'])}</td>"
                     f"<td>{'Yes' if r['is_solo'] else '—'}</td>"
                     f"<td>{str(r['response_deadline']).split('T')[0]}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No eligible international opportunities in '
                 'this window — most notices are domestic (CONUS).</p>')
    p.append('</section>')

    # Vibe-code lane: light web/software builds PRG can ship solo + AI.
    vibecode = sorted([r for r in results if r.get("is_vibecode")],
                      key=lambda r: r["win_score"], reverse=True)
    p.append('<section><h2>🧑‍💻 Vibe-Code — Web &amp; Software (Solo + AI)'
             '</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             + _html_escape(VIBECODE_DELIVERY_NOTE) + '</p>')
    if vibecode:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Rating</th><th>Win</th><th>Title</th><th>Agency</th>'
                 '<th>What to Build</th><th>Est. Value</th><th>Set-Aside</th>'
                 '<th>Respond By</th><th>Solicitation #</th></tr></thead><tbody>')
        for r in vibecode[:100]:
            chip = (f'<span class="chip" style="background:'
                    f'{_RAG_HEX[r["win_band"]]}">{r["win_band"]}</span>')
            sol = _html_escape(r["solicitation"])
            if _is_http(r["link"]):
                sol = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                       f'rel="noopener">{sol}</a>')
            p.append("<tr>"
                     f"<td>{chip}</td><td class='num'>{r['win_score']}</td>"
                     f"<td>{_html_escape(_clean(r['title'], 55))}</td>"
                     f"<td>{_html_escape(r['agency'])}</td>"
                     f"<td>{_html_escape(r['vibe_kw_display'])}</td>"
                     f"<td class='num'>{_html_escape(r['value_display'])}</td>"
                     f"<td>{_html_escape(r['setaside'])}</td>"
                     f"<td>{str(r['response_deadline']).split('T')[0]}</td>"
                     f"<td>{sol}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No web/software builds matched this window. '
                 'These are seasonal — re-run with a wider --days window.</p>')
    p.append('</section>')

    # Priority-agency focus (VA / HHS / DTRA): OSDBU portals + live counts.
    p.append('<section><h2>🏛️ Priority Agencies — OSDBU Focus</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'PRG-target agencies with active SDVOSB outreach. Their '
             'solicitations post to SAM.gov (pulled below) and their awards to '
             'USASpending (Recompete Radar). The OSDBU sites themselves are '
             'forecast / outreach portals — review them by hand for the '
             'pre-SAM.gov pipeline.</p>')
    p.append('<div class="scroll"><table><thead><tr>'
             '<th>Agency</th><th>Live Opportunities (this pull)</th>'
             '<th>Upcoming Recompetes</th><th>OSDBU / Acquisition Portal</th>'
             '</tr></thead><tbody>')
    for a in AGENCY_TARGETS:
        live = [r for r in eligible if r.get("agency_bucket") == a["key"]]
        rc = [r for r in (recompetes or []) if r.get("agency_bucket") == a["key"]]
        portal = _html_escape(a["portal"])
        p.append(
            f"<tr><td><b>{_html_escape(a['label'])}</b></td>"
            f"<td class='num'>{len(live)}</td>"
            f"<td class='num'>{len(rc)}</td>"
            f'<td><a href="{portal}" target="_blank" rel="noopener">'
            f"{portal}</a><br><span class='muted' style='font-size:12px'>"
            f"{_html_escape(a['portal_note'])}</span></td></tr>")
    p.append('</tbody></table></div></section>')

    # Resource Directory — the whole federal-contracting ecosystem, one click away.
    p.append('<section><h2>📚 Resource Directory — Every Place to Hunt &amp; '
             'Get Help</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'This tool searches SAM.gov + USASpending for you. These are the '
             'other places to hunt, research incumbents, get certified, and get '
             '<b>free</b> expert help. Everything below is a real government or '
             'SBA-partner resource — no logins sold, no fees.</p>')
    _cat_order, _by_cat = [], {}
    for row in _resource_rows():
        if row["category"] not in _by_cat:
            _by_cat[row["category"]] = []
            _cat_order.append(row["category"])
        _by_cat[row["category"]].append(row)
    _cat_emoji = {
        "Find Opportunities": "🔎", "Market Intel": "📊",
        "Certify & Register": "✅", "Free Help": "🤝", "Agency OSDBU": "🏛️",
    }
    for cat in _cat_order:
        p.append(f'<h3 style="margin:16px 0 6px">{_cat_emoji.get(cat, "•")} '
                 f'{_html_escape(cat)}</h3>')
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Resource</th><th>What it\'s for</th><th>Link</th>'
                 '</tr></thead><tbody>')
        for row in _by_cat[cat]:
            u = _html_escape(row["url"])
            p.append(
                f"<tr><td><b>{_html_escape(row['name'])}</b></td>"
                f"<td>{_html_escape(row['note'])}</td>"
                f'<td><a href="{u}" target="_blank" rel="noopener">{u}</a></td>'
                "</tr>")
        p.append('</tbody></table></div>')
    p.append('</section>')

    # Recompete Radar — upcoming rebids (from USASpending.gov).
    p.append('<section><h2>🔭 Recompete Radar — Upcoming Rebids</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Existing contracts in your NAICS that expire soon — position now '
             'to compete the rebid (or team with the incumbent). This is where '
             'most contracts are actually won.</p>')
    if recompetes:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Ends In</th><th>PoP End</th>'
                 '<th>Incumbent (beat / team)</th><th>Award Amount</th>'
                 '<th>Agency</th><th>Priority</th><th>NAICS</th><th>Award ID</th>'
                 '</tr></thead><tbody>')
        for r in recompetes:
            soon = r["months_left"] <= 6
            ends = (f'<b class="urgent">{r["months_left"]} mo 🔴</b>' if soon
                    else f'{r["months_left"]} mo')
            p.append(f"<tr><td>{ends}</td><td>{_html_escape(r['end_date'])}</td>"
                     f"<td>{_html_escape(r['recipient'])}</td>"
                     f"<td class='num'>{_html_escape(r['amount_display'])}</td>"
                     f"<td>{_html_escape(_clean(r['agency'], 34))}</td>"
                     f"<td>{_html_escape(r.get('agency_bucket') or '—')}</td>"
                     f"<td>{_html_escape(str(r['naics']))}</td>"
                     f"<td>{_html_escape(r['award_id'])}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No expiring contracts found (or USASpending '
                 'was unreachable this run). Widen --recompete-months and retry.</p>')
    p.append('</section>')

    # WATCH-TEMPLATE — Type A individual-credential opportunities (hire-to-win).
    templates = sorted([r for r in results if r.get("watch_template")],
                       key=lambda r: r["win_score"], reverse=True)
    p.append('<section><h2>🧩 WATCH-TEMPLATE — Hire-to-Win (individual '
             'credential)</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'These need a licensed/certified <b>person</b> PRG doesn\'t have on '
             'staff — but that person is <b>hireable W-2</b>, so their labor '
             'counts toward the 50% self-performance rule and PRG project-'
             'manages. Winnable IF you line up a contingent (on-award) commitment '
             'letter from the credential-holder <b>before</b> the deadline. Not a '
             'kill — a template to build a team around.</p>')
    if templates:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Credential Needed</th><th>Title</th><th>Agency</th>'
                 '<th>Set-Aside</th><th>NAICS</th><th>Respond By</th>'
                 '<th>ACTION</th><th>Solicitation #</th></tr></thead><tbody>')
        for r in templates[:100]:
            sol = _html_escape(r["solicitation"])
            if _is_http(r["link"]):
                sol = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                       f'rel="noopener">{sol}</a>')
            p.append(f"<tr><td><b>{_html_escape(r['credential_needed'])}</b></td>"
                     f"<td>{_html_escape(_clean(r['title'], 55))}</td>"
                     f"<td>{_html_escape(r['agency'])}</td>"
                     f"<td>{_html_escape(r['setaside'])}</td>"
                     f"<td>{_html_escape(r['naics'])}</td>"
                     f"<td>{str(r['response_deadline']).split('T')[0]}</td>"
                     f"<td>{_html_escape(r['template_action'])}</td>"
                     f"<td>{sol}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No individual-credential (hire-to-win) '
                 'opportunities this run.</p>')
    p.append('</section>')

    # WATCHLIST — FUTURE items (pre-sols / sources sought that fit): prep, don't bid.
    p.append('<section><h2>👀 Watchlist — Prep, Don’t Bid Yet</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Pre-solicitations and Sources Sought that fit PRG — not biddable '
             'now, but respond to shape the requirement and be ready when the '
             'RFP drops.</p>')
    if watchlist:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Notice Type</th><th>Title</th><th>Agency</th><th>NAICS</th>'
                 '<th>Set-Aside</th><th>Fit</th><th>Respond By</th>'
                 '<th>Solicitation #</th></tr></thead><tbody>')
        for r in watchlist[:100]:
            sol = _html_escape(r["solicitation"])
            if _is_http(r["link"]):
                sol = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                       f'rel="noopener">{sol}</a>')
            p.append(f"<tr><td>{_html_escape(r['notice_type'])}</td>"
                     f"<td>{_html_escape(_clean(r['title'], 55))}</td>"
                     f"<td>{_html_escape(r['agency'])}</td>"
                     f"<td>{_html_escape(r['naics'])}</td>"
                     f"<td>{_html_escape(r['setaside'])}</td>"
                     f"<td>{_html_escape(r['fit'])}</td>"
                     f"<td>{str(r['response_deadline']).split('T')[0]}</td>"
                     f"<td>{sol}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No fitting pre-solicitations this run.</p>')
    p.append('</section>')

    # CYCLE CALENDAR — recurring programs (SBIR/STTR): anticipate next release.
    cycles = {}
    for r in results:
        if r.get("cycle_program"):
            cycles.setdefault(r["cycle_program"], r)
    p.append('<section><h2>🗓️ Cycle Calendar — Recurring Programs</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Annual programs (SBIR/STTR) so next year&#39;s release is '
             'anticipated, not discovered late.</p>')
    if cycles:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Program</th><th>This Cycle (seen)</th>'
                 '<th>Inferred Next Window</th><th>Solicitation #</th>'
                 '</tr></thead><tbody>')
        for prog, r in cycles.items():
            this_due = str(r.get("response_deadline") or "").split("T")[0] or "n/a"
            p.append(f"<tr><td>{_html_escape(prog)}</td>"
                     f"<td>due {_html_escape(this_due)}</td>"
                     f"<td>{_html_escape(_next_cycle_window(r.get('response_deadline')))}</td>"
                     f"<td>{_html_escape(r['solicitation'])}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No recurring-program notices this run.</p>')
    p.append('</section>')

    # KILL LOG — auto-screened-out opportunities (deep-screen gates).
    killed = sorted([r for r in results if r["disqualified"]
                     and not r["is_awarded"] and not r["is_expired"]],
                    key=lambda r: r["kill_gate"])
    p.append('<section><h2>🛑 Kill Log — Auto-Screened Out</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Opportunities removed from the lists above by the deep-screen '
             "gates (self-performance, coverage, scope). This is a "
             '<b>keyword-level first pass</b> — the definitive screen still '
             'requires reading each notice’s PWS/attachments. Shown so the '
             'screen can be audited.</p>')
    if killed:
        p.append(f'<p style="font-size:13px;margin:0 0 10px"><b>{len(killed)}</b> '
                 'screened out.</p>')
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Failed Gate</th><th>Why</th><th>Title</th><th>Agency</th>'
                 '<th>NAICS</th><th>Solicitation #</th></tr></thead><tbody>')
        for r in killed[:150]:
            sol = _html_escape(r["solicitation"])
            if _is_http(r["link"]):
                sol = (f'<a href="{_html_escape(r["link"])}" target="_blank" '
                       f'rel="noopener">{sol}</a>')
            p.append(f"<tr><td>{_html_escape(r['kill_gate'])}</td>"
                     f"<td>{_html_escape(r['kill_reason'])}</td>"
                     f"<td>{_html_escape(_clean(r['title'], 55))}</td>"
                     f"<td>{_html_escape(r['agency'])}</td>"
                     f"<td>{_html_escape(r['naics'])}</td><td>{sol}</td></tr>")
        p.append('</tbody></table></div>')
        if len(killed) > 150:
            p.append(f'<p class="muted" style="font-size:12px">…and '
                     f'{len(killed) - 150} more (full list in the Excel KILL LOG '
                     'tab).</p>')
    else:
        p.append('<p class="empty">Nothing auto-screened-out this run.</p>')
    p.append('</section>')

    # NAICS gap — wide-net payoff: codes to add to SAM registration.
    gap = _naics_gap_rows(results)
    p.append('<section><h2>🧭 NAICS Gap — Codes to Add to Your SAM '
             'Registration</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'From the WIDE net (all eligible set-asides across ALL NAICS): these '
             'codes are NOT in PRG\'s current profile but carry set-aside-eligible, '
             'un-killed opportunities. Add the high-count ones to your SAM.gov Reps '
             '&amp; Certs so agency market research finds you and you match more '
             'searches.</p>')
    if gap:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>NAICS</th><th># Winnable</th><th>Best Win</th>'
                 '<th>Top Set-Aside</th><th>Example Opportunity</th>'
                 '<th>Agency</th></tr></thead><tbody>')
        for g in gap[:40]:
            p.append(f"<tr><td><b>{_html_escape(g['naics'])}</b></td>"
                     f"<td class='num'>{g['count']}</td>"
                     f"<td class='num'>{g['best_win']}</td>"
                     f"<td>{_html_escape(g['setaside'])}</td>"
                     f"<td>{_html_escape(_clean(g['example_title'], 55))}</td>"
                     f"<td>{_html_escape(g['example_agency'])}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No out-of-profile NAICS with winnable work '
                 'this run (or run without --narrow to cast the wide net).</p>')
    p.append('</section>')

    # Other federal opportunity banks with free APIs — Grants.gov + SBIR/STTR.
    p.append('<section><h2>💰 Grants.gov — Assistance &amp; Global-Health Awards'
             '</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Federal <b>grants &amp; cooperative agreements</b> (a different '
             'instrument than contracts) pulled live from the keyless Grants.gov '
             'API — including State Dept global-health APS and assistance awards '
             'that never touch SAM.gov.</p>')
    if grants:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Title</th><th>Agency</th><th>Status</th><th>Closes</th>'
                 '<th>ALN</th><th>Opportunity #</th></tr></thead><tbody>')
        for g in grants[:100]:
            t = _html_escape(_clean(g["title"], 60))
            if _is_http(g["link"]):
                t = (f'<a href="{_html_escape(g["link"])}" target="_blank" '
                     f'rel="noopener">{t}</a>')
            p.append(f"<tr><td>{t}</td><td>{_html_escape(g['agency'])}</td>"
                     f"<td>{_html_escape(g['status'])}</td>"
                     f"<td>{_html_escape(g['close_date'])}</td>"
                     f"<td>{_html_escape(g['aln'])}</td>"
                     f"<td>{_html_escape(g['number'])}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No grants pulled this run (Grants.gov '
                 'unreachable, or run without --no-grants).</p>')
    p.append('</section>')

    p.append('<section><h2>🔬 SBIR / STTR — Open R&amp;D Solicitations</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Open small-business R&amp;D innovation solicitations from the '
             'keyless SBIR.gov API — annual agency cycles, squarely in PRG\'s '
             'research lane.</p>')
    if sbir:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Solicitation</th><th>Program</th><th>Agency</th>'
                 '<th>Closes</th><th>Topics</th></tr></thead><tbody>')
        for s in sbir[:100]:
            t = _html_escape(_clean(s["title"], 60))
            if _is_http(s["link"]):
                t = (f'<a href="{_html_escape(s["link"])}" target="_blank" '
                     f'rel="noopener">{t}</a>')
            p.append(f"<tr><td>{t}</td><td>{_html_escape(s['program'])}</td>"
                     f"<td>{_html_escape(s['agency'])}</td>"
                     f"<td>{_html_escape(s['close_date'])}</td>"
                     f"<td>{_html_escape(s['topics'])}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No open SBIR/STTR solicitations pulled this '
                 'run.</p>')
    p.append('</section>')

    # ReliefWeb — live international short-term consultancies (free API).
    p.append('<section><h2>🌍 ReliefWeb — Live International Consultancies</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'UN &amp; humanitarian short-term postings pulled live from the '
             'keyless ReliefWeb API, filtered to PRG\'s clinical/health/research '
             'lane — the one international bank that <b>does</b> have an open API.'
             '</p>')
    if reliefweb:
        p.append('<div class="scroll"><table><thead><tr>'
                 '<th>Title</th><th>Organization</th><th>Country</th>'
                 '<th>Category</th><th>Closes</th></tr></thead><tbody>')
        for j in reliefweb[:100]:
            t = _html_escape(_clean(j["title"], 60))
            if _is_http(j["link"]):
                t = (f'<a href="{_html_escape(j["link"])}" target="_blank" '
                     f'rel="noopener">{t}</a>')
            p.append(f"<tr><td>{t}</td><td>{_html_escape(j['org'])}</td>"
                     f"<td>{_html_escape(j['country'])}</td>"
                     f"<td>{_html_escape(j['category'])}</td>"
                     f"<td>{_html_escape(j['closing'])}</td></tr>")
        p.append('</tbody></table></div>')
    else:
        p.append('<p class="empty">No ReliefWeb postings pulled this run '
                 '(unreachable, or run without --no-reliefweb).</p>')
    p.append('</section>')

    # International opportunity sources — where the short-term global gigs live.
    p.append('<section><h2>🌐 International Sources — Where the Short-Term Gigs '
             'Live</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'Short-term international consultancies are scattered across portals '
             'that don\'t talk to each other and mostly need a login (no open API '
             'to pull) — so this is a <b>register-here directory</b>. Fit-notes '
             'are tuned to PRG\'s founder (ACRP clinical creds, Spanish, veteran + '
             'evacuation-ops). <b>Best first move: register on UNGM + 2–3 '
             'implementing-partner rosters this week — it costs nothing.</b></p>')
    _isrc_by_cat, _isrc_order = {}, []
    for row in _intl_source_rows():
        if row["category"] not in _isrc_by_cat:
            _isrc_by_cat[row["category"]] = []
            _isrc_order.append(row["category"])
        _isrc_by_cat[row["category"]].append(row)
    for cat in _isrc_order:
        p.append(f'<h3 style="margin:16px 0 6px">{_html_escape(cat)}</h3>')
        p.append('<div class="scroll"><table><thead><tr><th>Source</th>'
                 '<th>What it covers / fit for PRG</th><th>Register</th>'
                 '</tr></thead><tbody>')
        for row in _isrc_by_cat[cat]:
            u = _html_escape(row["url"])
            note = _html_escape(row["note"]).replace("**", "")
            p.append(f"<tr><td><b>{_html_escape(row['name'])}</b></td>"
                     f"<td>{note}</td>"
                     f'<td><a href="{u}" target="_blank" rel="noopener">{u}</a>'
                     "</td></tr>")
        p.append('</tbody></table></div>')
    p.append('</section>')

    # SLED + marketplace sources — the biggest pool this report does NOT pull
    # live (state/local portals need per-agency accounts; no open APIs).
    p.append('<section><h2>🏛️ SLED &amp; Marketplaces — The Pool This Report '
             'Doesn\'t Pull</h2>')
    p.append('<p class="muted" style="font-size:13px;margin:0 0 12px">'
             'State/local/education portals and federal marketplaces post an '
             'estimated 50–90K open solicitations nationwide at any moment, '
             'but each needs its own vendor registration and none offer an '
             'open API — so this is a <b>register-here directory</b>, '
             'priority-ordered for PRG. <b>Best first moves: the VA forecast '
             '(SDVOSB rule-of-two), then Cal eProcure + the CA DVBE '
             'certification (3% state participation goals = built-in '
             'demand).</b></p>')
    _sled_by_cat, _sled_order = {}, []
    for row in _sled_source_rows():
        if row["category"] not in _sled_by_cat:
            _sled_by_cat[row["category"]] = []
            _sled_order.append(row["category"])
        _sled_by_cat[row["category"]].append(row)
    for cat in _sled_order:
        p.append(f'<h3 style="margin:16px 0 6px">{_html_escape(cat)}</h3>')
        p.append('<div class="scroll"><table><thead><tr><th>Source</th>'
                 '<th>Why / how it fits PRG</th><th>Register</th>'
                 '</tr></thead><tbody>')
        for row in _sled_by_cat[cat]:
            u = _html_escape(row["url"])
            note = _html_escape(row["note"]).replace("**", "")
            p.append(f"<tr><td><b>{_html_escape(row['name'])}</b></td>"
                     f"<td>{note}</td>"
                     f'<td><a href="{u}" target="_blank" rel="noopener">{u}</a>'
                     "</td></tr>")
        p.append('</tbody></table></div>')
    p.append('</section>')

    # Pipeline insights
    p.append('<section><h2>Growth &amp; Pipeline Insights</h2><div class="card">')
    if agency_counts:
        foot = ", ".join(f"{_html_escape(a)} ({n})"
                         for a, n in agency_counts.most_common(3))
        p.append(f'<p><b>Strongest agency footprints:</b> {foot}.</p>')
        targets = [a for a, _ in agency_counts.most_common(3)]
        p.append('<p><b>Recommended target agencies for future bids:</b> '
                 + _html_escape(", ".join(targets)) + '. Concentrate capability '
                 'statements and Sources Sought responses here to build past '
                 'performance where PRG already sees demand.</p>')
        naics_counts = Counter(r["naics"] for r in eligible)
        strong_naics = ", ".join(f"{n} ({c})" for n, c in naics_counts.most_common(3))
        p.append(f'<p><b>Strongest NAICS demand:</b> {_html_escape(strong_naics)}.</p>')
    else:
        p.append('<p class="empty">Not enough eligible data yet — widen the '
                 'search window and re-run.</p>')
    p.append('</div></section>')

    p.append('<footer>Heuristic planning report generated from live SAM.gov '
             'data. Win scores and estimates are decision aids, not guarantees. '
             'Verify scope, value, and contacts in each official notice before '
             f'bidding. · {COMPANY_NAME} · {dt.date.today().isoformat()}</footer>')
    p.append('<script>' + _REPORT_JS + '</script>')
    p.append('</div></body></html>')

    if not path.lower().endswith(".html"):
        path += ".html"
    html = "".join(p)
    try:
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(html)
    except PermissionError:
        alt = _timestamped_alt(path)
        with open(alt, "w", encoding="utf-8") as fh:
            fh.write(html)
        sys.stderr.write(
            f"NOTE: '{os.path.basename(path)}' was open/locked — saved as "
            f"'{os.path.basename(alt)}' instead.\n"
        )
        return alt
    return path


_REPORT_CSS = """
:root{--paper:#F4F6F9;--surface:#fff;--ink:#14263C;--muted:#5A6B7E;
--line:#DCE1E8;--primary:#1E3A5F;--accent:#9A6B1C;}
*{box-sizing:border-box}
body{margin:0;background:var(--paper);color:var(--ink);
font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Helvetica,Arial,sans-serif;
line-height:1.55;-webkit-font-smoothing:antialiased}
.wrap{max-width:1000px;margin:0 auto;padding:36px 22px 70px}
header{border-bottom:3px solid var(--accent);padding-bottom:18px;margin-bottom:24px}
.eyebrow{font-size:11px;letter-spacing:.16em;text-transform:uppercase;
color:var(--accent);font-weight:700;margin:0 0 8px}
h1{font-family:Georgia,"Times New Roman",serif;font-size:clamp(24px,4vw,34px);
margin:0 0 8px;letter-spacing:-.01em}
.sub{color:var(--muted);font-size:13.5px;margin:0}
h2{font-family:Georgia,serif;font-size:21px;margin:34px 0 12px;
border-bottom:1px solid var(--line);padding-bottom:6px}
h3{font-size:14px;margin:0 0 10px;color:var(--ink)}
.kpis{display:grid;grid-template-columns:repeat(6,1fr);gap:12px;margin-bottom:26px}
@media(max-width:820px){.kpis{grid-template-columns:repeat(3,1fr)}}
@media(max-width:520px){.kpis{grid-template-columns:repeat(2,1fr)}}
.kpi{background:var(--surface);border:1px solid var(--line);border-left:4px solid var(--primary);
border-radius:10px;padding:14px 16px}
.kpi-v{font-family:Georgia,serif;font-size:24px;font-weight:700;color:var(--primary);
font-variant-numeric:tabular-nums;line-height:1.1}
.kpi-l{font-size:12px;font-weight:700;margin-top:4px}
.kpi-s{font-size:11px;color:var(--muted)}
.charts{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:8px}
@media(max-width:820px){.charts{grid-template-columns:1fr}}
.card{background:var(--surface);border:1px solid var(--line);border-radius:12px;padding:18px 20px}
.chart{width:100%;height:auto}
.c-lab{fill:var(--ink);font-size:12px}
.c-val{fill:var(--muted);font-size:12px;font-weight:700}
.scroll{overflow-x:auto;border:1px solid var(--line);border-radius:10px}
table{border-collapse:collapse;width:100%;min-width:820px;font-size:12.5px}
thead th{background:#EEF1F5;text-align:left;padding:9px 11px;font-size:10.5px;
letter-spacing:.06em;text-transform:uppercase;border-bottom:1px solid var(--line);white-space:nowrap}
tbody td{padding:9px 11px;border-bottom:1px solid var(--line);vertical-align:top}
tbody tr:last-child td{border-bottom:none}
.num{font-variant-numeric:tabular-nums;text-align:right;white-space:nowrap}
.chip{display:inline-block;color:#fff;font-size:10.5px;font-weight:700;
padding:2px 8px;border-radius:20px;white-space:nowrap}
.dive{background:var(--surface);border:1px solid var(--line);border-radius:10px;
padding:14px 18px;margin-bottom:12px}
.dive-h{margin-bottom:6px;font-size:15px}
.dive p{margin:4px 0;font-size:13px}
.muted{color:var(--muted)}
a{color:#1E5F9E;font-weight:600;text-decoration:none}
a:hover{text-decoration:underline}
.urgent{color:#C0392B}
.findings{background:#FBF6EC;border:1px solid var(--accent);border-radius:12px;
padding:8px 20px 16px;margin:8px 0 26px}
.findings h2{border:none;margin:14px 0 8px;font-size:19px}
.find{display:flex;gap:10px;padding:7px 0;border-top:1px solid #EADFC6;font-size:13px}
.find:first-of-type{border-top:none}
.find-i{font-size:16px;flex:none;width:22px;text-align:center}
.empty{color:var(--muted);font-style:italic;padding:14px}
footer{margin-top:34px;padding-top:16px;border-top:1px solid var(--line);
font-size:11.5px;color:var(--muted)}
.kpi.clickable{cursor:pointer;transition:box-shadow .12s,transform .12s}
.kpi.clickable:hover{box-shadow:0 4px 14px rgba(20,38,60,.14);transform:translateY(-1px)}
.kpi.clickable:focus-visible{outline:3px solid var(--accent);outline-offset:2px}
.kpi-go{color:var(--accent);font-weight:700;white-space:nowrap}
.filterbar{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin:0 0 12px}
.fb{font:inherit;font-size:12.5px;font-weight:600;color:var(--ink);
background:var(--surface);border:1px solid var(--line);border-radius:20px;
padding:6px 13px;cursor:pointer}
.fb:hover{border-color:var(--primary)}
.fb.active{background:var(--primary);color:#fff;border-color:var(--primary)}
.fsearch{font:inherit;font-size:13px;padding:6px 12px;border:1px solid var(--line);
border-radius:20px;min-width:190px;flex:1 1 190px;max-width:320px}
.count{font-size:12.5px;color:var(--muted);font-weight:600;white-space:nowrap}
@media print{body{background:#fff}.wrap{max-width:100%;padding:0}
.card,.dive,.kpi{break-inside:avoid}.filterbar{display:none}}
"""


_REPORT_JS = r"""
(function(){
  var body=document.getElementById('matrixBody');
  if(!body) return;
  var rows=[].slice.call(body.querySelectorAll('tr[data-band]'));
  var countEl=document.getElementById('count');
  var qEl=document.getElementById('q');
  var buttons=[].slice.call(document.querySelectorAll('.fb'));
  var state={filter:'all',q:''};

  function matchFilter(tr,f){
    if(f==='all') return true;
    if(f==='solo') return tr.getAttribute('data-solo')==='1';
    if(f==='intl') return tr.getAttribute('data-intl')==='1';
    if(f==='sdvosb') return tr.getAttribute('data-sdvosb')==='1';
    if(f.indexOf('band:')===0) return tr.getAttribute('data-band')===f.slice(5);
    if(f.indexOf('role:')===0) return tr.getAttribute('data-role')===f.slice(5);
    return true;
  }
  function apply(){
    var shown=0;
    rows.forEach(function(tr){
      var ok=matchFilter(tr,state.filter) &&
        (!state.q || (tr.getAttribute('data-text')||'').indexOf(state.q)>-1);
      tr.style.display=ok?'':'none';
      if(ok) shown++;
    });
    if(countEl) countEl.textContent='Showing '+shown+' of '+rows.length;
    buttons.forEach(function(b){
      b.classList.toggle('active',b.getAttribute('data-filter')===state.filter);
    });
  }
  function setFilter(f){ state.filter=f; apply();
    document.getElementById('matrix').scrollIntoView({behavior:'smooth',block:'start'}); }

  buttons.forEach(function(b){
    b.addEventListener('click',function(){ state.filter=b.getAttribute('data-filter'); apply(); });
  });
  document.querySelectorAll('.kpi.clickable').forEach(function(k){
    var go=function(){ setFilter(k.getAttribute('data-filter')); };
    k.addEventListener('click',go);
    k.addEventListener('keydown',function(e){ if(e.key==='Enter'||e.key===' '){e.preventDefault();go();} });
  });
  if(qEl) qEl.addEventListener('input',function(){ state.q=qEl.value.toLowerCase().trim(); apply(); });
  apply();
})();
"""


# ---------------------------------------------------------------------------
# SELF-TESTS — verify the kill rules fire on two known NO-BID solicitations.
# (Representative payloads: the SAM.gov description/attachments aren't fetched
# here, so these mirror the disqualifying language each notice is known for.)
# ---------------------------------------------------------------------------

_SELFTEST_CASES = [
    {
        "label": "VA 36C26026Q0674 — medical gas (ASSE-cert techs + equipment = "
                 "capital/workforce required → KILL on Improvement 2)",
        "opp": {
            "solicitationNumber": "36C26026Q0674",
            "noticeId": "st1",
            "title": "Medical Gas System Inspection, Testing and Certification",
            "description": ("Contractor shall provide medical gas system testing "
                            "and certification. Personnel must hold ASSE 6030 and "
                            "ASSE 6040 medical gas certification per NFPA 99."),
            "naicsCode": "811210",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_truthy": ["capital_required"],
    },
    {
        "label": "VA 36C26126Q0797 — Medical Physicist (Type A individual "
                 "credential, ABR board-certified → WATCH-TEMPLATE, hireable)",
        "opp": {
            "solicitationNumber": "36C26126Q0797",
            "noticeId": "st1a",
            "title": "Medical Physicist Services — Radiation Therapy QA",
            "description": ("The contractor shall provide the services of a "
                            "board-certified medical physicist (certified by the "
                            "American Board of Radiology) to perform annual "
                            "quality-assurance surveys and calibration of the "
                            "radiation therapy system. Key personnel: the "
                            "physicist shall be board certified."),
            "naicsCode": "541380",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "WATCH-TEMPLATE",
        "expect_truthy": ["watch_template", "credential_needed"],
        "expect_falsy": ["disqualified"],
    },
    {
        "label": "VA 36C24426Q0491 — Erie VAMC Metasys/JCI (Type B corporate OEM "
                 "authorization → KILL + capital-required)",
        "opp": {
            "solicitationNumber": "36C24426Q0491",
            "noticeId": "st1b",
            "title": "Building Automation System Maintenance — Metasys",
            "description": ("Maintenance and service of the Johnson Controls "
                            "Metasys building automation system. The offeror must "
                            "be an authorized JCI service provider with written "
                            "authorization from the manufacturer to service the "
                            "proprietary system."),
            "naicsCode": "238220",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_truthy": ["capital_required"],
    },
    {
        "label": "Navy N0040626Q0335 — Fluke calibration (Type B OEM calibration "
                 "authorization → KILL)",
        "opp": {
            "solicitationNumber": "N0040626Q0335",
            "noticeId": "st1c",
            "title": "Calibration of Fluke Test Equipment",
            "description": ("Calibration and repair of Fluke test equipment. Work "
                            "shall be performed by a Fluke-authorized calibration "
                            "provider to manufacturer specification."),
            "naicsCode": "811210",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "Department of the Navy",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_truthy": ["capital_required"],
    },
    {
        "label": "FDA 75F40126R00051 — NCTR O&M (CBA + workforce + PP wall)",
        "opp": {
            "solicitationNumber": "75F40126R00051",
            "noticeId": "st2",
            "title": "NCTR Operations and Maintenance Services",
            "description": ("Operations and maintenance of facilities. A collective "
                            "bargaining agreement applies and Service Contract Act "
                            "wage determinations are incorporated. Offerors must "
                            "provide a phase-in plan and demonstrate recent and "
                            "relevant past performance of similar size and scope; "
                            "CPARS will be evaluated."),
            "naicsCode": "561210",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "Department of Health and Human Services",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
    },
    {
        "label": "BLM 140L3726Q0100 — Class III Cultural Resource Inventory "
                 "(CRUP + SOI standards + fieldwork)",
        "opp": {
            "solicitationNumber": "140L3726Q0100",
            "noticeId": "st3",
            "title": ("Class III Cultural Resource Inventory, Uncompahgre "
                      "Field Office"),
            "description": ("Contractor shall conduct a Class III pedestrian "
                            "cultural resource inventory. The contractor must "
                            "hold a current BLM Cultural Resource Use Permit and "
                            "all supervisory personnel must meet the Secretary of "
                            "the Interior's Professional Qualification Standards "
                            "in archaeology. Fieldwork includes pedestrian survey "
                            "along transects with a high-clearance vehicle."),
            "naicsCode": "541620",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "Department of the Interior",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
    },
    {
        "label": "USSOCOM H9242126QE036 — SaaS subscription resale "
                 "(rubric v2 Gate 0: resale, no services labor → KILL)",
        "opp": {
            "solicitationNumber": "H9242126QE036",
            "noticeId": "st4",
            "title": "Software as a Service (SaaS) Subscription Licenses",
            "description": ("Procurement of commercial SaaS subscription "
                            "licenses. The offeror may be required to provide an "
                            "authorized reseller letter of supply from the OEM."),
            "naicsCode": "513210",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "US Special Operations Command",
            "type": "Combined Synopsis/Solicitation",
        },
        "deadline_offset_days": 3,
        "expect_verdict": "NO-BID",
        "expect_equals": {"kill_gate": "Gate 0 (resale, no value-add)",
                          "win_band": "Red", "win_score": 0},
    },
    {
        "label": "GE MAC 5500 ECG PM — OEM pass-through (brand equipment + "
                 "manufacturer spec, killed regardless of SDVOSB set-aside)",
        "opp": {
            "solicitationNumber": "OEM-ECG-TEST",
            "noticeId": "st4b",
            "title": "GE MAC 5500 ECG Preventive Maintenance and Calibration",
            "description": ("Contractor shall provide preventive maintenance and "
                            "calibration of GE MAC 5500 ECG units. All work shall "
                            "be performed by an authorized service center in "
                            "accordance with the manufacturer specification."),
            "naicsCode": "811210",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
    },
    {
        "label": "ED/IES SBIR 91990026R0005 — expired pre-sol (EXPIRED at stage 0, "
                 "cycle entry)",
        "opp": {
            "solicitationNumber": "91990026R0005",
            "noticeId": "st5",
            "title": "ED/IES SBIR Direct to Phase II",
            "description": ("Department of Education Institute of Education "
                            "Sciences SBIR program. Small Business Innovation "
                            "Research Direct to Phase II solicitation."),
            "naicsCode": "541715",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "Department of Education",
            "type": "Presolicitation",
        },
        "deadline_offset_days": -8,
        "expect_verdict": "EXPIRED",
        "expect_truthy": ["cycle_program"],
    },
    {
        "label": "Synthetic Sources Sought (NAICS 541611) — WATCH + "
                 "RESPOND-RECOMMENDED, no bid-queue entry",
        "opp": {
            "solicitationNumber": "SS-541611-TEST",
            "noticeId": "st6",
            "title": "Sources Sought — Management Consulting Support",
            "description": ("Sources sought seeking capable small businesses for "
                            "management and program support consulting."),
            "naicsCode": "541611",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "General Services Administration",
            "type": "Sources Sought",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "WATCH",
        "expect_truthy": ["respond_recommended"],
    },
    {
        "label": "NCIA Sources Sought (in-lane consulting, international buyer) — "
                 "WATCH but NOT respond-recommended, scored down (flag not kill)",
        "opp": {
            "solicitationNumber": "NCIA-SS-TEST",
            "noticeId": "st7",
            "title": "Sources Sought — Program Support Consulting",
            "description": ("NATO Communications and Information Agency (NCIA) "
                            "seeks capable firms for management and program "
                            "support consulting."),
            "naicsCode": "541611",
            "typeOfSetAside": "",
            "typeOfSetAsideDescription": "",
            "fullParentPathName": "NATO Communications and Information Agency",
            "type": "Sources Sought",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "WATCH",
        "expect_truthy": ["buyer_is_international"],
        "expect_falsy": ["respond_recommended"],
    },
    {
        "label": "State/Embassy Manila 19RP3826Q0057 — medical test-kit SUPPLY "
                 "(stale pre-sol + overseas post + commodity supply + unrestricted)",
        "opp": {
            "solicitationNumber": "19RP3826Q0057",
            "noticeId": "st8",
            "title": ("Supply and Delivery of Medical Testing Kits — "
                      "U.S. Embassy Manila"),
            "description": ("The U.S. Embassy in Manila, Philippines intends to "
                            "issue a solicitation for the supply and delivery of "
                            "medical testing kits and diagnostic laboratory "
                            "supplies. Award will be made on a lowest price "
                            "technically acceptable (LPTA) basis. Anticipated "
                            "release date: April 2026."),
            "naicsCode": "339113",
            "classificationCode": "6550",
            "typeOfSetAside": "",
            "typeOfSetAsideDescription": "",
            "organizationName": "American Embassy Manila",
            "fullParentPathName": "STATE, DEPARTMENT OF",
            "placeOfPerformance": {"country": {"code": "PHL",
                                               "name": "Philippines"}},
            "type": "Presolicitation",
        },
        "deadline_offset_days": -30,       # April 2026 window already passed
        "expect_verdict": "EXPIRED",
        "expect_truthy": ["disqualified", "buyer_is_international"],
        "expect_falsy": ["tech_match"],
        "expect_equals": {"setaside": "None"},
    },
    {
        "label": "Synthetic construction NAICS 236220 — small roof job, no "
                 "bonding (capture v3: prime via trade sub, NOT killed)",
        "opp": {
            "solicitationNumber": "KC1-NAICS-TEST",
            "noticeId": "st9",
            "title": "Roof Replacement, Building 204",
            "description": ("Replacement of the roof membrane and flashing on "
                            "Building 204."),
            "naicsCode": "236220",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "BID",
        "expect_falsy": ["disqualified"],
        "expect_equals": {"role": "PRIME WITH TEAMING PARTNER",
                          "fulfillment_model": "trade-sub management"},
    },
    {
        "label": "Synthetic construction WITH performance bond — Miller Act "
                 "bonding stays a hard kill (flagged 'bonding required')",
        "opp": {
            "solicitationNumber": "KC1-BOND-TEST",
            "noticeId": "st10",
            "title": "Parking Lot Paving and Striping",
            "description": ("Asphalt paving and striping of the north parking "
                            "lot. Performance and payment bonds are required "
                            "in accordance with FAR 52.228-15."),
            "naicsCode": "237310",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_truthy": ["bonding_required"],
        "expect_contains": {"kill_gate": "bonding"},
    },
    {
        "label": "Synthetic $25M program support — wrong scale "
                 "(KC5 → KILL, single-member LLC can't self-perform 50%)",
        "opp": {
            "solicitationNumber": "KC5-SCALE-TEST",
            "noticeId": "st11",
            "title": "Enterprise Program Management Support Services",
            "description": ("Program management and administrative support "
                            "services. Estimated ceiling: $25,000,000 over a "
                            "five-year ordering period."),
            "naicsCode": "541611",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "General Services Administration",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_equals": {"kill_gate": "Gate 0 (KC5 wrong scale)"},
    },
    # --- Rubric v2 regression set (from the last report's known-bad ratings) ---
    {
        "label": "REGRESSION trash can replacement — product purchase "
                 "(Gate 0, product → Red; old rubric scored this 72/Green)",
        "opp": {
            "solicitationNumber": "REG-TRASHCAN",
            "noticeId": "rg1",
            "title": "Replacement Trash Can Receptacles",
            "description": ("Furnish and deliver replacement trash can "
                            "receptacles to the facility loading dock. Brand "
                            "name or equal."),
            "naicsCode": "332999",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_equals": {"win_band": "Red", "win_score": 0},
    },
    {
        "label": "REGRESSION EBSCO subscriptions — resale (Gate 0 → Red)",
        "opp": {
            "solicitationNumber": "REG-EBSCO",
            "noticeId": "rg2",
            "title": "EBSCO Research Database Subscriptions Renewal",
            "description": ("Annual renewal of EBSCO research database "
                            "subscriptions for the medical library."),
            "naicsCode": "519130",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "Department of Health and Human Services",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_equals": {"kill_gate": "Gate 0 (resale, no value-add)",
                          "win_band": "Red", "win_score": 0},
    },
    {
        "label": "REGRESSION vacuum pump replacement — capture v3: small "
                 "unbonded trade job = prime with mechanical sub (not killed)",
        "opp": {
            "solicitationNumber": "REG-VACPUMP",
            "noticeId": "rg3",
            "title": "Medical Vacuum Pump Replacement, Building 1",
            "description": ("Removal of the existing medical vacuum pumps and "
                            "installation of new pumps, including associated "
                            "piping and electrical work."),
            "naicsCode": "238220",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "BID",
        "expect_falsy": ["disqualified"],
        "expect_equals": {"role": "PRIME WITH TEAMING PARTNER",
                          "fulfillment_model": "trade-sub management"},
    },
    {
        "label": "CAPTURE janitorial services — delegable trade, deep local "
                 "sub market (not killed; prime with teaming)",
        "opp": {
            "solicitationNumber": "CAP-JANITORIAL",
            "noticeId": "cp1",
            "title": "Janitorial Services, Federal Building",
            "description": ("Janitorial and custodial services for the "
                            "federal building, five days per week. Base year "
                            "plus four option years."),
            "naicsCode": "561720",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "General Services Administration",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "BID",
        "expect_falsy": ["disqualified"],
        "expect_equals": {"role": "PRIME WITH TEAMING PARTNER",
                          "fulfillment_model": "trade-sub management"},
    },
    {
        "label": "CAPTURE value-added supply — furniture supply AND "
                 "installation (NMR flag, not killed)",
        "opp": {
            "solicitationNumber": "CAP-SUPPLY-VA",
            "noticeId": "cp2",
            "title": "Exam Room Furniture — Supply and Installation",
            "description": ("Furnish and deliver exam room furniture "
                            "including installation, assembly, and removal "
                            "and disposal of existing furniture."),
            "naicsCode": "337127",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "BID",
        "expect_falsy": ["disqualified"],
        "expect_equals": {"fulfillment_model": "value-added supply"},
        "expect_contains": {"soft_flags_display": "nonmanufacturer"},
    },
    {
        "label": "CAPTURE guard services — firm-level PPO licensure stays a "
                 "kill (not delegable without ostensible-sub risk)",
        "opp": {
            "solicitationNumber": "CAP-GUARDS",
            "noticeId": "cp3",
            "title": "Unarmed Security Guard Services",
            "description": ("Unarmed guard services at the facility, two "
                            "posts, day shift."),
            "naicsCode": "561612",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Veterans Affairs",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_equals": {"kill_gate": "Gate 0 (firm licensure)"},
    },
    {
        "label": "REGRESSION duct bank survey — firm land-surveying licensure "
                 "with no key-personnel allowance (Gate 0 → Red)",
        "opp": {
            "solicitationNumber": "REG-DUCTBANK",
            "noticeId": "rg4",
            "title": "Underground Utility Duct Bank Survey",
            "description": ("Survey and mapping of the underground utility "
                            "duct bank. The firm must be licensed to perform "
                            "land surveying; all deliverables shall be sealed "
                            "by a licensed land surveyor."),
            "naicsCode": "541370",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "General Services Administration",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "NO-BID",
        "expect_equals": {"kill_gate": "Gate 0 (firm licensure)",
                          "win_band": "Red", "win_score": 0},
    },
    {
        "label": "REGRESSION DOJ Communication Support (SDVOSB) — hirable "
                 "commodity labor, founder can supervise (Green, W2 plan)",
        "opp": {
            "solicitationNumber": "REG-DOJ-COMMS",
            "noticeId": "rg5",
            "title": "Communication Support Services",
            "description": ("The contractor shall provide communication "
                            "support services including communications "
                            "specialists to develop outreach materials, "
                            "newsletters, and stakeholder engagement reports. "
                            "Base year plus four option years."),
            "naicsCode": "541611",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Justice",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "BID",
        "expect_equals": {"win_band": "Green"},
        "expect_contains": {"labor_plan": "W2 communications specialist"},
    },
    {
        "label": "REGRESSION HHS clinical modeling (Total SB) — founder "
                 "domain, deliverable-based, small (Green, Solo: Yes)",
        "opp": {
            "solicitationNumber": "REG-HHS-MODEL",
            "noticeId": "rg6",
            "title": "Statistical Modeling of Clinical Trial Outcomes",
            "description": ("The contractor shall perform statistical "
                            "modeling of clinical trial outcomes data and "
                            "deliver a final report with model documentation. "
                            "A biostatistician-level analysis is required. "
                            "Estimated value: $180,000."),
            "naicsCode": "541715",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "Department of Health and Human Services",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "BID",
        "expect_truthy": ["is_solo"],
        "expect_equals": {"win_band": "Green", "pp_value": "High"},
        "expect_contains": {"labor_plan": "1099 biostatistician"},
    },
    {
        "label": "REGRESSION environmental compliance — founder cannot judge "
                 "quality (Gate 3 low → Yellow at best, never Green)",
        "opp": {
            "solicitationNumber": "REG-ENVCOMP",
            "noticeId": "rg7",
            "title": "Environmental Compliance Support Services",
            "description": ("Environmental compliance support including "
                            "environmental sampling, air monitoring, and "
                            "preparation of annual compliance reports by "
                            "qualified environmental scientists. Base year "
                            "plus two option years."),
            "naicsCode": "541620",
            "typeOfSetAside": "SBA",
            "typeOfSetAsideDescription": "Total Small Business",
            "fullParentPathName": "General Services Administration",
            "type": "Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "forbid_verdict": "NO-BID",
        "expect_falsy": ["is_solo"],
        "expect_contains": {"win_band": "ellow"},  # Yellow (at best) — never Green
    },
    {
        "label": "State/Vienna 19AU9026Q0012 — heritage-collection appraisal "
                 "(appraiser = hireable credential → WATCH-TEMPLATE, not Red)",
        "opp": {
            "solicitationNumber": "19AU9026Q0012",
            "noticeId": "st13",
            "title": "Appraisal of Heritage Collection — U.S. Embassy Vienna",
            "description": ("Professional appraisal services to evaluate 33 "
                            "pieces of contemporary and historic works of art, "
                            "sculptures, textiles, and furniture comprising "
                            "the Cultural Heritage collection. An acknowledged "
                            "subject matter expert in fine art appraisal is "
                            "required; a qualified appraiser shall document "
                            "condition and establish fair market values. "
                            "Report and valuation written in English."),
            "naicsCode": "541990",
            "classificationCode": "R707",
            "typeOfSetAside": "",
            "typeOfSetAsideDescription": "",
            "organizationName": "U.S. Tri-Mission Vienna",
            "fullParentPathName": "STATE, DEPARTMENT OF",
            "placeOfPerformance": {"country": {"code": "AUT",
                                               "name": "Austria"}},
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "expect_verdict": "WATCH-TEMPLATE",
        "expect_truthy": ["watch_template", "credential_needed",
                          "buyer_is_international"],
        "expect_falsy": ["disqualified"],
        "expect_equals": {"role": "PRIME WITH TEAMING PARTNER"},
    },
    {
        "label": "LEAD TIME — janitorial (teaming needed) with a 4-day fuse "
                 "→ killed at Gate 0 lead time (needs 10d)",
        "opp": {
            "solicitationNumber": "LEAD-JAN-4D",
            "noticeId": "lt1",
            "title": "Janitorial Services, Field Office",
            "description": ("Janitorial and custodial services for the field "
                            "office, five days per week."),
            "naicsCode": "561720",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "General Services Administration",
            "type": "Solicitation",
        },
        "deadline_offset_days": 4,
        "expect_verdict": "NO-BID",
        "expect_equals": {"kill_gate": "Gate 0 (lead time)", "win_score": 0},
    },
    {
        "label": "LEAD TIME — founder-deliverable clinical analysis with a "
                 "3-day fuse → still actionable (SHORT-FUSE, not killed)",
        "opp": {
            "solicitationNumber": "LEAD-SOLO-3D",
            "noticeId": "lt2",
            "title": "Rapid Literature Review — Clinical Trial Retention",
            "description": ("The contractor shall deliver a literature review "
                            "and final report on clinical trial participant "
                            "retention strategies. Estimated value: $45,000."),
            "naicsCode": "541715",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "Department of Health and Human Services",
            "type": "Combined Synopsis/Solicitation",
        },
        "deadline_offset_days": 3,
        "expect_verdict": "SHORT-FUSE",
        "expect_truthy": ["is_solo"],
        "expect_falsy": ["disqualified"],
    },
    {
        "label": "Synthetic size-standard mention — '$34.5 million size "
                 "standard' is NOT a contract value (KC5 must not fire)",
        "opp": {
            "solicitationNumber": "KC5-SIZESTD-TEST",
            "noticeId": "st12",
            "title": "Management Consulting Support",
            "description": ("Management and program support consulting. The "
                            "small business size standard is $34.5 million."),
            "naicsCode": "541611",
            "typeOfSetAside": "SDVOSBC",
            "typeOfSetAsideDescription": "SDVOSB Set-Aside",
            "fullParentPathName": "General Services Administration",
            "type": "Combined Synopsis/Solicitation",
            "responseDeadLine": "2026-12-01",
        },
        "forbid_verdict": "NO-BID",
        "expect_falsy": ["disqualified"],
    },
]


def run_selftests():
    """Evaluate the known cases and assert the verdicts (and any required
    fields) fire. Returns 0 on all-pass, 1 otherwise."""
    ok = True
    today = dt.date.today()
    print("PRG screen self-tests:\n")
    for case in _SELFTEST_CASES:
        opp = dict(case["opp"])
        if "deadline_offset_days" in case:
            opp["responseDeadLine"] = (
                today + dt.timedelta(days=case["deadline_offset_days"])).isoformat()
        r = evaluate(opp)
        if "expect_verdict" in case:
            passed = (r["verdict"] == case["expect_verdict"])
            want = f"== {case['expect_verdict']}"
        else:
            passed = (r["verdict"] != case["forbid_verdict"])
            want = f"!= {case['forbid_verdict']}"
        for field in case.get("expect_truthy", []):
            if not r.get(field):
                passed = False
                want += f" & {field}"
        for field in case.get("expect_falsy", []):
            if r.get(field):
                passed = False
                want += f" & !{field}"
        for field, val in case.get("expect_equals", {}).items():
            if r.get(field) != val:
                passed = False
                want += f" & {field}=={val!r}"
        for field, sub in case.get("expect_contains", {}).items():
            if sub.lower() not in str(r.get(field, "")).lower():
                passed = False
                want += f" & {sub!r} in {field}"
        ok = ok and passed
        print(f"  [{'PASS' if passed else 'FAIL'}] {case['label']}")
        extra = ""
        report_fields = (case.get("expect_truthy", [])
                         + case.get("expect_falsy", [])
                         + list(case.get("expect_equals", {}))
                         + list(case.get("expect_contains", {})))
        if report_fields:
            extra = " | " + ", ".join(
                f"{f}={r.get(f)!r}" for f in report_fields)
        print(f"         verdict={r['verdict']} (want {want}), "
              f"gate={r['kill_gate'] or '-'}{extra}")
    print("\nResult:", "ALL PASS ✅" if ok else "SOME FAILED ❌")
    return 0 if ok else 1


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def parse_args(argv):
    parser = argparse.ArgumentParser(
        description="SAM.gov capability matcher for Pacific Research Group LLC.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent(
            """\
            Examples:
              python3 samgov_opportunity_matcher.py
              python3 samgov_opportunity_matcher.py --days 60 --limit 200
            """
        ),
    )
    parser.add_argument("--days", type=int, default=30,
                        help="Look-back window in days (default: 30).")
    parser.add_argument("--limit", type=int, default=100,
                        help="Max opportunities per NAICS query (default: 100).")
    parser.add_argument("--api-key", default=None,
                        help="Override the SAM.gov API key.")
    parser.add_argument("--excel", default=None, metavar="FILE",
                        help="Path for the Excel workbook. By default it is "
                             "saved as PRG_Contracts_YYYY-MM-DD.xlsx (the search "
                             "date) on your Desktop. Falls back to CSV files if "
                             "openpyxl is missing.")
    parser.add_argument("--no-excel", action="store_true",
                        help="Skip writing the spreadsheet (console output only).")
    parser.add_argument("--report", default=None, metavar="FILE",
                        help="Path for the executive HTML report. By default it "
                             "is saved as PRG_Executive_Report_YYYY-MM-DD.html "
                             "(the search date) on your Desktop.")
    parser.add_argument("--no-report", action="store_true",
                        help="Skip writing the executive HTML report.")
    parser.add_argument("--outdir", default=None, metavar="FOLDER",
                        help="Save date-stamped Excel + HTML into this folder "
                             "(creates it if needed). Used by the scheduled "
                             "weekly/monthly automation to keep a history.")
    parser.add_argument("--no-print", action="store_true",
                        help="Suppress the console Markdown report.")
    parser.add_argument("--no-psc", action="store_true",
                        help="Skip the extra PSC (Product/Service Code) queries.")
    parser.add_argument("--narrow", action="store_true",
                        help="Restrict to PRG's target NAICS/PSC only. By DEFAULT "
                             "the tool casts the WIDE net — every set-aside PRG is "
                             "eligible for across ALL NAICS — and reports which "
                             "NAICS to add to SAM. Use --narrow for a fast, "
                             "profile-only run.")
    parser.add_argument("--no-recompetes", action="store_true",
                        help="Skip the Recompete Radar (USASpending.gov) lookup.")
    parser.add_argument("--no-grants", action="store_true",
                        help="Skip the Grants.gov assistance-opportunity pull.")
    parser.add_argument("--no-sbir", action="store_true",
                        help="Skip the SBIR.gov open-solicitation pull.")
    parser.add_argument("--no-reliefweb", action="store_true",
                        help="Skip the ReliefWeb international-consultancy pull.")
    parser.add_argument("--recompete-months", type=int, default=18, metavar="N",
                        help="Recompete horizon: contracts expiring within N "
                             "months (default: 18).")
    parser.add_argument("--selftest", action="store_true",
                        help="Run the built-in NO-BID rule self-tests and exit.")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv if argv is not None else sys.argv[1:])

    if args.selftest:
        return run_selftests()

    api_key = args.api_key or API_KEY

    if not api_key or api_key.startswith("SAM-") is False:
        sys.stderr.write(
            "WARNING: API key looks malformed. Set --api-key or SAM_API_KEY.\n"
        )

    posted_from, posted_to = _date_range(args.days)
    sys.stderr.write(
        f"Querying SAM.gov for active opportunities "
        f"({posted_from} – {posted_to}) across "
        f"{len(TARGET_NAICS)} core + {len(LOW_BARRIER_NAICS)} low-barrier "
        f"NAICS codes...\n"
    )

    # Pull opportunities per NAICS (core + low-barrier) and de-dupe by notice id.
    seen = set()
    all_opps = []
    for naics in ALL_QUERY_NAICS:
        batch = fetch_opportunities(
            api_key, naics, posted_from, posted_to, args.limit
        )
        sys.stderr.write(f"  NAICS {naics}: {len(batch)} record(s)\n")
        for opp in batch:
            key = opp.get("noticeId") or opp.get("solicitationNumber") or id(opp)
            if key in seen:
                continue
            seen.add(key)
            all_opps.append(opp)

    # Also query by target PSC (Product/Service) codes — more specific than
    # NAICS, catches work classified by service type.
    if not args.no_psc:
        for psc in TARGET_PSC:
            batch = fetch_opportunities(
                api_key, psc, posted_from, posted_to, args.limit, code_param="ccode"
            )
            sys.stderr.write(f"  PSC {psc}: {len(batch)} record(s)\n")
            for opp in batch:
                key = opp.get("noticeId") or opp.get("solicitationNumber") or id(opp)
                if key in seen:
                    continue
                seen.add(key)
                all_opps.append(opp)

    # WIDE NET — sweep EVERY set-aside PRG is eligible for across ALL NAICS, not
    # just the target list. This is where opportunities in codes PRG hasn't
    # registered surface; the NAICS-gap report then flags which to add to SAM.
    if not args.narrow:
        sys.stderr.write("  Wide net: sweeping all eligible set-asides across "
                         "ALL NAICS (this is the big one)...\n")
        for sa in SETASIDE_QUERY_CODES:
            batch = fetch_opportunities(
                api_key, sa, posted_from, posted_to, args.limit,
                code_param="typeOfSetAside",
            )
            new = 0
            for opp in batch:
                key = opp.get("noticeId") or opp.get("solicitationNumber") or id(opp)
                if key in seen:
                    continue
                seen.add(key)
                all_opps.append(opp)
                new += 1
            sys.stderr.write(f"  set-aside {sa}: {len(batch)} record(s) "
                             f"({new} new)\n")

    # Recompete Radar — upcoming rebids from USASpending.gov (keyless API).
    recompetes = []
    if not args.no_recompetes:
        sys.stderr.write("  Recompete Radar: querying USASpending.gov...\n")
        recompetes = fetch_recompetes(
            list(TARGET_NAICS), months_ahead=args.recompete_months
        )
        # Targeted pulls for priority agencies (VA / HHS / DTRA) so their
        # incumbents + recompetes surface even outside the general page window.
        seen_awards = {r["award_id"] for r in recompetes}
        for a in AGENCY_TARGETS:
            sys.stderr.write(f"    + {a['key']} awarded-contract pull...\n")
            extra = fetch_recompetes(
                list(TARGET_NAICS), months_ahead=args.recompete_months,
                max_pages=5, agency=a,
            )
            for r in extra:
                if r["award_id"] not in seen_awards:
                    seen_awards.add(r["award_id"])
                    recompetes.append(r)
        recompetes.sort(key=lambda r: r["end_date"])
        sys.stderr.write(f"    {len(recompetes)} contract(s) expiring soon\n")

    # Other opportunity banks with free, keyless APIs (grants + SBIR/STTR).
    grants = []
    if not args.no_grants:
        sys.stderr.write("  Grants.gov: querying assistance opportunities...\n")
        grants = fetch_grants()
        sys.stderr.write(f"    {len(grants)} grant opportunity(ies)\n")
    sbir = []
    if not args.no_sbir:
        sys.stderr.write("  SBIR.gov: querying open R&D solicitations...\n")
        sbir = fetch_sbir()
        sys.stderr.write(f"    {len(sbir)} open SBIR/STTR solicitation(s)\n")
    reliefweb = []
    if not args.no_reliefweb:
        sys.stderr.write("  ReliefWeb: querying international consultancies...\n")
        reliefweb = fetch_reliefweb()
        sys.stderr.write(f"    {len(reliefweb)} international posting(s)\n")

    # Evaluate every opportunity.
    results = [evaluate(opp) for opp in all_opps]

    # FIX 6 — short-circuit anything already vetted this week (checksum match).
    seen_before = apply_vet_cache(results)
    if seen_before:
        sys.stderr.write(f"  ({seen_before} already vetted in a prior run)\n")

    # Sort: eligible first, then by score, then SDVOSB set-asides on top.
    results.sort(
        key=lambda r: (r["eligible"], r["score"], r["is_sdvosb"]),
        reverse=True,
    )

    if not args.no_print:
        render_report(results)

    # Resolve output paths. --outdir writes date-stamped files into a folder
    # (for the scheduled automation history); otherwise write to the Desktop.
    stamp = dt.date.today().isoformat()
    if args.outdir:
        os.makedirs(args.outdir, exist_ok=True)
        xlsx_path = os.path.join(args.outdir, f"PRG_Contracts_{stamp}.xlsx")
        report_path = os.path.join(args.outdir, f"PRG_Executive_Report_{stamp}.html")
    else:
        # Default Desktop files carry the search date so each run is its own
        # dated snapshot (e.g. PRG_Contracts_2026-07-07.xlsx) instead of
        # overwriting the last one. An explicit --excel/--report still wins.
        xlsx_path = args.excel or _desktop_path(f"PRG_Contracts_{stamp}.xlsx")
        report_path = args.report or _desktop_path(
            f"PRG_Executive_Report_{stamp}.html")

    saved = []
    if not args.no_excel:
        try:
            saved.append(export_spreadsheet(results, xlsx_path, recompetes,
                                            grants, sbir, reliefweb))
        except Exception as exc:  # never let one save abort the other
            sys.stderr.write(f"WARNING: could not save spreadsheet: {exc}\n")
    if not args.no_report:
        try:
            saved.append(export_html_report(results, report_path, args.days,
                                            recompetes, grants, sbir, reliefweb))
        except Exception as exc:
            sys.stderr.write(f"WARNING: could not save HTML report: {exc}\n")

    if _RUN_STATE["rate_limited"]:
        sys.stderr.write(
            "\n*** NOTE: SAM.gov rate-limited this run — the results are "
            "PARTIAL. Re-run later (or with a lower --limit) for full coverage. "
            "A daily quota applies to each API key. ***\n"
        )

    if saved:
        sys.stderr.write(
            "\n============================================================\n"
            "  FILES SAVED:\n" + "".join(f"  {s}\n" for s in saved)
            + "  (Open the .html report in your browser; the .xlsx in Excel.)\n"
            "============================================================\n"
        )

    # Keep the window open when run interactively (e.g. double-clicked on
    # Windows) so the output and saved-file path don't vanish instantly.
    # --outdir marks a scheduled/automation run: a Task Scheduler console
    # still counts as a TTY, and pausing there would hang the task forever.
    try:
        if sys.stdout.isatty() and not args.outdir:
            input("\nDone. Press Enter to close this window...")
    except (EOFError, KeyboardInterrupt):
        pass

    return 0


if __name__ == "__main__":
    sys.exit(main())
